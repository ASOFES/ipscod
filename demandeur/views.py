from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from core.models import Course, ActionTraceur, Utilisateur, Message
from .forms import DemandeForm
from notifications.utils import notify_user, send_sms, send_whatsapp
import datetime
from django.http import HttpResponse
from django.template.loader import get_template, render_to_string
# from weasyprint import HTML  # Temporairement comment√© pour le d√©ploiement
import openpyxl
from openpyxl.styles import Font
import os
import base64
from django.contrib.staticfiles import finders
from django.core.mail import send_mail
import xlwt
from django.views.decorators.http import require_POST

@login_required
def dashboard(request):
    """Vue pour le tableau de bord du demandeur de missions"""
    # R√©cup√©rer les demandes de l'utilisateur connect√©
    if request.user.role == 'admin' or request.user.is_superuser:
        # Pour les admins et superusers, montrer toutes les demandes o√π ils sont demandeur OU de leur √©tablissement
        demandes = Course.objects.filter(
            Q(demandeur=request.user) | Q(etablissement=request.user.etablissement)
        ).order_by('-date_demande')
    else:
        # Pour les demandeurs normaux, montrer seulement leurs demandes
        demandes = Course.objects.filter(demandeur=request.user).order_by('-date_demande')
    
    # Filtres
    statut = request.GET.get('statut')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if statut:
        demandes = demandes.filter(statut=statut)
    
    if date_debut:
        date_debut = datetime.datetime.strptime(date_debut, '%Y-%m-%d').date()
        demandes = demandes.filter(date_demande__date__gte=date_debut)
    
    if date_fin:
        date_fin = datetime.datetime.strptime(date_fin, '%Y-%m-%d').date()
        demandes = demandes.filter(date_demande__date__lte=date_fin)
    
    # Pagination
    paginator = Paginator(demandes, 12)  # 12 demandes par page
    page_number = request.GET.get('page')
    demandes_page = paginator.get_page(page_number)
    
    # Statistiques
    if request.user.role == 'admin' or request.user.is_superuser:
        stats = {
            'total': Course.objects.count(),
            'en_attente': Course.objects.filter(statut='en_attente').count(),
            'validees': Course.objects.filter(statut='validee').count(),
            'refusees': Course.objects.filter(statut='refusee').count(),
        }
    else:
        stats = {
            'total': Course.objects.filter(demandeur=request.user).count(),
            'en_attente': Course.objects.filter(demandeur=request.user, statut='en_attente').count(),
            'validees': Course.objects.filter(demandeur=request.user, statut='validee').count(),
            'refusees': Course.objects.filter(demandeur=request.user, statut='refusee').count(),
        }
    
    context = {
        'demandes_page': demandes_page,
        'stats': stats,
        'user': request.user,
    }
    
    return render(request, 'demandeur/dashboard.html', context)

@login_required
def nouvelle_demande(request):
    """Vue pour cr√©er une nouvelle demande de mission"""
    if request.method == 'POST':
        form = DemandeForm(request.POST)
        if form.is_valid():
            demande = form.save(commit=False)
            demande.demandeur = request.user
            demande.statut = 'en_attente'
            demande.etablissement = request.user.etablissement
            demande.save()
            
            # Cr√©er une entr√©e dans l'historique des actions
            ActionTraceur.objects.create(
                utilisateur=request.user,
                action="Cr√©ation de demande de mission",
                details=f"Demande #{demande.id} - {demande.point_embarquement} ‚Üí {demande.destination}"
            )
            
            # R√©cup√©rer les administrateurs et les dispatchers
            admins_dispatchers = Utilisateur.objects.filter(
                Q(role='admin') | Q(role='dispatch'),
                is_active=True
            ).distinct()
            
            # Message de notification
            notification_title = f"Nouvelle demande de course #{demande.id}"
            notification_message = f"{request.user.get_full_name()} a cr√©√© une nouvelle demande de course de {demande.point_embarquement} √† {demande.destination}."
            
            # Envoi des notifications √† chaque admin/dispatcher
            for user in admins_dispatchers:
                # Notification interne
                notify_user(
                    user,
                    notification_title,
                    notification_message,
                    notification_type='all',
                    course=demande
                )
                
                # Envoi SMS/WhatsApp (d√©sactiv√© pour le moment)
                if user.telephone:
                    send_sms(user.telephone, notification_message)
                    send_whatsapp(user.telephone, notification_message)
            
            # R√©cup√©rer l'utilisateur syst√®me pour les messages syst√®me
            system_user = Utilisateur.objects.get_or_create(
                username='system',
                defaults={
                    'email': 'system@example.com',
                    'password': 'system_password',  # Le mot de passe sera ignor√© car l'utilisateur existe d√©j√†
                    'role': 'admin',
                    'is_staff': True,
                    'is_superuser': True,
                    'first_name': 'Syst√®me',
                    'last_name': 'MAMO'
                }
            )[0]
            
            # Cr√©er un message pour chaque admin/dispatcher
            for user in admins_dispatchers:
                chat_message = f"""üöó Nouvelle demande de course #{demande.id}
- üë§ Demandeur: {demande.demandeur.get_full_name()}
- üìÖ Date souhait√©e: {demande.date_souhaitee.strftime('%d/%m/%Y √† %H:%M')}
- üö© D√©part: {demande.point_embarquement}
- üèÅ Destination: {demande.destination}
- üë• Nombre de passagers: {demande.nombre_passagers}
- üìù Statut: En attente de validation"""
                
                # Cr√©er le message dans le chat interne
                Message.objects.create(
                    sender=system_user,  # Utiliser l'utilisateur syst√®me comme exp√©diteur
                    recipient=user,
                    content=chat_message,
                    is_read=False
                )
            
            messages.success(request, 'Votre demande de mission a √©t√© cr√©√©e avec succ√®s et est en attente de validation.')
            return redirect('demandeur:detail_demande', demande.id)
    else:
        form = DemandeForm()
    
    return render(request, 'demandeur/nouvelle_demande.html', {'form': form})

@login_required
def detail_demande(request, demande_id):
    """Vue pour afficher les d√©tails d'une demande de mission"""
    # Pour les admins et superusers, permettre l'acc√®s √† toutes les demandes
    if request.user.role == 'admin' or request.user.is_superuser:
        demande = get_object_or_404(Course, id=demande_id)
    else:
        demande = get_object_or_404(Course, id=demande_id, demandeur=request.user)
    
    # R√©cup√©rer l'historique des actions li√©es √† cette demande
    historique = ActionTraceur.objects.filter(
        Q(details__icontains=f"Demande #{demande.id}") |
        Q(details__icontains=f"Course {demande.id}")
    ).order_by('-date_action')
    
    context = {
        'demande': demande,
        'historique': historique,
    }
    
    return render(request, 'demandeur/detail_demande.html', context)

@login_required
def modifier_demande(request, demande_id):
    """Vue pour modifier une demande de mission"""
    # Pour les admins et superusers, permettre l'acc√®s √† toutes les demandes
    if request.user.role == 'admin' or request.user.is_superuser:
        demande = get_object_or_404(Course, id=demande_id, statut='en_attente')
    else:
        demande = get_object_or_404(Course, id=demande_id, demandeur=request.user, statut='en_attente')
    
    if request.method == 'POST':
        form = DemandeForm(request.POST, instance=demande)
        if form.is_valid():
            form.save()
            
            # Cr√©er une entr√©e dans l'historique des actions
            ActionTraceur.objects.create(
                utilisateur=request.user,
                action="Modification de demande de mission",
                details=f"Demande #{demande.id} - {demande.point_embarquement} ‚Üí {demande.destination}"
            )
            
            messages.success(request, 'Votre demande de mission a √©t√© modifi√©e avec succ√®s.')
            return redirect('demandeur:detail_demande', demande.id)
    else:
        form = DemandeForm(instance=demande)
    
    return render(request, 'demandeur/nouvelle_demande.html', {'form': form, 'demande': demande})

@login_required
def annuler_demande(request, demande_id):
    """Vue pour annuler une demande de mission"""
    # Pour les admins et superusers, permettre l'acc√®s √† toutes les demandes
    if request.user.role == 'admin' or request.user.is_superuser:
        demande = get_object_or_404(Course, id=demande_id, statut='en_attente')
    else:
        demande = get_object_or_404(Course, id=demande_id, demandeur=request.user, statut='en_attente')
    
    demande.statut = 'annulee'
    demande.save()
    
    # Cr√©er une entr√©e dans l'historique des actions
    ActionTraceur.objects.create(
        utilisateur=request.user,
        action="Annulation de demande de mission",
        details=f"Demande #{demande.id} - {demande.point_embarquement} ‚Üí {demande.destination}"
    )
    
    messages.success(request, 'Votre demande de mission a √©t√© annul√©e avec succ√®s.')
    return redirect('demandeur:dashboard')

@login_required
def demande_pdf(request, demande_id):
    demande = get_object_or_404(Course, id=demande_id)
    template = get_template('demandeur/demande_pdf.html')
    html = template.render({
        'demande': demande,
        'request': request,
        'now': timezone.now()
    })

    pdf_bytes = HTML(string=html).write_pdf()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="demande_{demande_id}.pdf"'
    return response

@login_required
def demande_excel(request, demande_id):
    demande = get_object_or_404(Course, id=demande_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Demande {demande.id}"

    # En-t√™tes
    ws['A1'] = "Champ"
    ws['B1'] = "Valeur"
    ws['A1'].font = ws['B1'].font = Font(bold=True)

    # Donn√©es
    data = [
        ("ID", demande.id),
        ("Date de demande", demande.date_demande.strftime("%d/%m/%Y %H:%M") if demande.date_demande else ""),
        ("Date souhait√©e", demande.date_souhaitee.strftime("%d/%m/%Y %H:%M") if demande.date_souhaitee else ""),
        ("Demandeur", demande.demandeur.get_full_name() if demande.demandeur else ""),
        ("D√©partement du demandeur", demande.demandeur.etablissement.nom if demande.demandeur and demande.demandeur.etablissement else "-"),
        ("T√©l√©phone demandeur", demande.demandeur.telephone if demande.demandeur and demande.demandeur.telephone else "-"),
        ("Destination", demande.destination),
        ("Point d'embarquement", demande.point_embarquement),
        ("Motif", demande.motif),
        ("Nombre de passagers", demande.nombre_passagers),
        ("Statut", demande.get_statut_display()),
        ("Chauffeur", demande.chauffeur.get_full_name() if demande.chauffeur else "Non assign√©"),
        ("D√©partement du v√©hicule", demande.vehicule.etablissement.nom if demande.vehicule and demande.vehicule.etablissement else "-"),
        ("V√©hicule", demande.vehicule.immatriculation if demande.vehicule else "Non assign√©"),
        ("Marque v√©hicule", demande.vehicule.marque if demande.vehicule else "-"),
        ("Mod√®le v√©hicule", demande.vehicule.modele if demande.vehicule else "-"),
        ("Num√©ro de ch√¢ssis", demande.vehicule.numero_chassis if demande.vehicule else "-"),
        ("Kilom√©trage d√©part", demande.kilometrage_depart if demande.kilometrage_depart else "-"),
        ("Kilom√©trage fin", demande.kilometrage_fin if demande.kilometrage_fin else "-"),
        ("Distance parcourue", demande.distance_parcourue if demande.distance_parcourue else "-"),
        ("Date d√©part", demande.date_depart.strftime("%d/%m/%Y %H:%M") if demande.date_depart else ""),
        ("Date fin", demande.date_fin.strftime("%d/%m/%Y %H:%M") if demande.date_fin else ""),
        ("Dispatcher", demande.dispatcher.get_full_name() if demande.dispatcher else "-"),
        ("Date validation", demande.date_validation.strftime("%d/%m/%Y %H:%M") if demande.date_validation else ""),
    ]
    for i, (champ, valeur) in enumerate(data, start=2):
        ws[f'A{i}'] = champ
        ws[f'B{i}'] = valeur

    # G√©n√©ration de la r√©ponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=demande_{demande.id}.xlsx'
    wb.save(response)
    return response

@login_required
def demandes_pdf(request):
    # R√©cup√©rer les demandes selon le r√¥le
    if request.user.role == 'admin' or request.user.is_superuser:
        demandes = Course.objects.filter(etablissement=request.user.etablissement).order_by('-date_demande')
    else:
        demandes = Course.objects.filter(demandeur=request.user, etablissement=request.user.etablissement).order_by('-date_demande')

    # Utiliser finders pour localiser le logo dans les fichiers statiques
    logo_path = finders.find('images/logo_mamo.png')
    logo_data_uri = ''
    if logo_path and os.path.exists(logo_path):
        with open(logo_path, "rb") as image_file:
            logo_data_uri = "data:image/png;base64," + base64.b64encode(image_file.read()).decode('utf-8')

    template = get_template('demandeur/demandes_pdf.html')
    html = template.render({
        'demandes': demandes, 
        'user': request.user, 
        'logo_data_uri': logo_data_uri,
        'now': timezone.now(),
        'request': request
    })

    # G√©n√©rer le PDF directement en m√©moire
    pdf_bytes = HTML(string=html).write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'filename="demandes.pdf"'
    return response

@login_required
def demandes_excel(request):
    # R√©cup√©rer les demandes selon le r√¥le
    if request.user.role == 'admin' or request.user.is_superuser:
        demandes = Course.objects.filter(etablissement=request.user.etablissement).order_by('-date_demande')
    else:
        demandes = Course.objects.filter(demandeur=request.user, etablissement=request.user.etablissement).order_by('-date_demande')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demandes"

    # En-t√™tes
    headers = [
        "ID", "Date de demande", "Date souhait√©e", "Demandeur", "D√©partement du demandeur", "T√©l√©phone demandeur", "Destination", "Point d'embarquement", "Motif", "Nombre de passagers", "Statut", "Chauffeur", "D√©partement du v√©hicule", "V√©hicule", "Marque v√©hicule", "Mod√®le v√©hicule", "Num√©ro de ch√¢ssis", "Kilom√©trage d√©part", "Kilom√©trage fin", "Distance parcourue", "Date d√©part", "Date fin", "Dispatcher", "Date validation"
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    # Donn√©es
    for row, demande in enumerate(demandes, start=2):
        ws.cell(row=row, column=1, value=demande.id)
        ws.cell(row=row, column=2, value=demande.date_demande.strftime("%d/%m/%Y %H:%M") if demande.date_demande else "")
        ws.cell(row=row, column=3, value=demande.date_souhaitee.strftime("%d/%m/%Y %H:%M") if demande.date_souhaitee else "")
        ws.cell(row=row, column=4, value=demande.demandeur.get_full_name() if demande.demandeur else "")
        ws.cell(row=row, column=5, value=demande.demandeur.etablissement.nom if demande.demandeur and demande.demandeur.etablissement else "-")
        ws.cell(row=row, column=6, value=demande.demandeur.telephone if demande.demandeur and demande.demandeur.telephone else "-")
        ws.cell(row=row, column=7, value=demande.destination)
        ws.cell(row=row, column=8, value=demande.point_embarquement)
        ws.cell(row=row, column=9, value=demande.motif)
        ws.cell(row=row, column=10, value=demande.nombre_passagers)
        ws.cell(row=row, column=11, value=demande.get_statut_display())
        ws.cell(row=row, column=12, value=demande.chauffeur.get_full_name() if demande.chauffeur else "Non assign√©")
        ws.cell(row=row, column=13, value=demande.vehicule.etablissement.nom if demande.vehicule and demande.vehicule.etablissement else "-")
        ws.cell(row=row, column=14, value=demande.vehicule.immatriculation if demande.vehicule else "Non assign√©")
        ws.cell(row=row, column=15, value=demande.vehicule.marque if demande.vehicule else "-")
        ws.cell(row=row, column=16, value=demande.vehicule.modele if demande.vehicule else "-")
        ws.cell(row=row, column=17, value=demande.vehicule.numero_chassis if demande.vehicule else "-")
        ws.cell(row=row, column=18, value=demande.kilometrage_depart if demande.kilometrage_depart else "-")
        ws.cell(row=row, column=19, value=demande.kilometrage_fin if demande.kilometrage_fin else "-")
        ws.cell(row=row, column=20, value=demande.distance_parcourue if demande.distance_parcourue else "-")
        ws.cell(row=row, column=21, value=demande.date_depart.strftime("%d/%m/%Y %H:%M") if demande.date_depart else "")
        ws.cell(row=row, column=22, value=demande.date_fin.strftime("%d/%m/%Y %H:%M") if demande.date_fin else "")
        ws.cell(row=row, column=23, value=demande.dispatcher.get_full_name() if demande.dispatcher else "-")
        ws.cell(row=row, column=24, value=demande.date_validation.strftime("%d/%m/%Y %H:%M") if demande.date_validation else "")

    # G√©n√©ration de la r√©ponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=demandes.xlsx'
    wb.save(response)
    return response

@login_required
@require_POST
def notifier_chauffeurs(request):
    chauffeurs = Utilisateur.objects.filter(role='chauffeur', is_active=True)
    emails = [c.email for c in chauffeurs if c.email]
    # Notification interne (message Django)
    for c in chauffeurs:
        # Ici, tu peux cr√©er un objet Message ou Notification si tu as un mod√®le d√©di√©
        pass
    # Notification externe (email)
    if emails:
        send_mail(
            'Notification MAMO',
            'Vous avez une nouvelle notification sur la plateforme MAMO.',
            'no-reply@mamo.com',
            emails,
            fail_silently=True
        )
    messages.success(request, "Tous les chauffeurs ont √©t√© notifi√©s (interne et email si disponible).")
    return redirect('demandeur:dashboard')

@login_required
def export_courses_jour_pdf(request):
    today = timezone.localdate()
    courses = Course.objects.filter(date_souhaitee__date=today)
    
    # Pr√©parer le contexte avec les donn√©es n√©cessaires
    context = {
        'courses': courses, 
        'date': today,
        'now': timezone.now(),
        'user': request.user
    }
    
    # Utiliser render_to_pdf qui inclut automatiquement le logo
    from core.utils import render_to_pdf
    return render_to_pdf('demandeur/courses_jour_pdf.html', context, f'courses_{today}.pdf')

@login_required
def export_courses_jour_excel(request):
    today = timezone.localdate()
    courses = Course.objects.filter(date_souhaitee__date=today)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="courses_{today}.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Courses du jour')
    row_num = 0
    columns = ['ID', 'Demandeur', 'D√©part', 'Destination', 'Chauffeur', 'V√©hicule', 'Statut', 'Date souhait√©e']
    for col_num, column_title in enumerate(columns):
        ws.write(row_num, col_num, column_title)
    for course in courses:
        row_num += 1
        ws.write(row_num, 0, course.id)
        ws.write(row_num, 1, str(course.demandeur))
        ws.write(row_num, 2, course.point_embarquement)
        ws.write(row_num, 3, course.destination)
        ws.write(row_num, 4, str(course.chauffeur) if course.chauffeur else '-')
        ws.write(row_num, 5, str(course.vehicule) if course.vehicule else '-')
        ws.write(row_num, 6, course.get_statut_display())
        ws.write(row_num, 7, course.date_souhaitee.strftime('%d/%m/%Y %H:%M') if course.date_souhaitee else '-')
    wb.save(response)
    return response
