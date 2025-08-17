import logging
from datetime import datetime, timedelta
from decimal import Decimal
import os

from django.db.models import (
    Sum, F, Q, ExpressionWrapper, DecimalField, 
    Count, Avg, Value, DurationField
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.template.loader import render_to_string
from django.utils import timezone
from django.conf import settings
from django.core.paginator import Paginator
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook
import io
import base64
import csv

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Models
from core.models import Vehicule, Course, Utilisateur
from ravitaillement.models import Ravitaillement
from entretien.models import Entretien
from core.models import HistoriqueKilometrage

logger = logging.getLogger(__name__)

def is_admin_or_dispatch_or_superuser(user):
    return user.is_authenticated and (user.role in ['admin', 'dispatch'] or user.is_superuser)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def dashboard(request):
    """Vue pour le tableau de bord des rapports."""
    # Statistiques pour le tableau de bord
    total_vehicules = Vehicule.objects.count()
    total_chauffeurs = Utilisateur.objects.filter(role='chauffeur').count()
    total_missions_terminees = Course.objects.filter(statut='terminee').count()
    total_entretiens = Entretien.objects.count()
    
    context = {
        'total_vehicules': total_vehicules,
        'total_chauffeurs': total_chauffeurs,
        'total_missions_terminees': total_missions_terminees,
        'total_entretiens': total_entretiens
    }
    return render(request, 'rapport/dashboard.html', context)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_vehicules(request):
    """Rapport détaillé sur les véhicules avec système de scoring avancé."""
    # Récupération des paramètres de filtre
    immatriculation = request.GET.get('immatriculation')
    marque = request.GET.get('marque')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Filtrage initial
    vehicules = Vehicule.objects.all()
    
    # Application des filtres
    if immatriculation:
        vehicules = vehicules.filter(immatriculation__iexact=immatriculation)
    if marque:
        vehicules = vehicules.filter(marque__icontains=marque)
    if date_debut:
        vehicules = vehicules.filter(date_creation__date__gte=date_debut)
    if date_fin:
        vehicules = vehicules.filter(date_creation__date__lte=date_fin)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(vehicules, 12)  # 12 véhicules par page

    stats_vehicules = []
    scores_data = []  # Pour calculer les moyennes de scoring
    
    for v in vehicules:
        # 1. Distance totale depuis la mise en service
        distance_totale = v.kilometrage_actuel if v.kilometrage_actuel is not None else 0
        # 2. Distance depuis le dernier entretien
        if v.kilometrage_actuel is not None and v.kilometrage_dernier_entretien is not None:
            distance_apres_entretien = v.kilometrage_actuel - v.kilometrage_dernier_entretien
            if distance_apres_entretien < 0:
                distance_apres_entretien = 0
        else:
            distance_apres_entretien = None
        # 3. Distance parcourue (somme des courses terminées)
        from core.models import Course
        courses_terminees = Course.objects.filter(vehicule=v, statut='terminee')
        distance_parcourue_courses = courses_terminees.aggregate(total=Sum('distance_parcourue'))['total'] or 0
        # 4. Distance estimée (kilométrage)
        plus_ancien_depart = courses_terminees.order_by('kilometrage_depart').first()
        if v.kilometrage_actuel is not None and plus_ancien_depart and plus_ancien_depart.kilometrage_depart is not None:
            distance_estimee_km = v.kilometrage_actuel - plus_ancien_depart.kilometrage_depart
            if distance_estimee_km < 0:
                distance_estimee_km = 0
        else:
            distance_estimee_km = None
        # 5. Alerte incohérence si écart > 10%
        alerte_incoherence = False
        if distance_parcourue_courses and distance_estimee_km:
            ecart = abs(distance_parcourue_courses - distance_estimee_km)
            seuil = 0.1 * max(distance_parcourue_courses, distance_estimee_km)
            if ecart > seuil:
                alerte_incoherence = True

        entretiens = Entretien.objects.filter(vehicule=v)
        if date_debut:
            entretiens = entretiens.filter(date_creation__date__gte=date_debut)
        if date_fin:
            entretiens = entretiens.filter(date_creation__date__lte=date_fin)
        nb_entretiens = entretiens.count()
        types_entretiens = ', '.join(sorted(set(e.get_type_entretien_display() for e in entretiens)))
        cout_entretiens = entretiens.aggregate(Sum('cout'))['cout__sum'] or 0

        ravs = Ravitaillement.objects.filter(vehicule=v)
        if date_debut:
            ravs = ravs.filter(date_ravitaillement__date__gte=date_debut)
        if date_fin:
            ravs = ravs.filter(date_ravitaillement__date__lte=date_fin)
        total_litres = ravs.aggregate(Sum('litres'))['litres__sum'] or 0
        total_cout_carburant = ravs.aggregate(Sum('cout_total'))['cout_total__sum'] or 0

        budget_total = cout_entretiens + total_cout_carburant
        
        # === SYSTÈME DE SCORING AVANCÉ POUR VÉHICULES ===
        score_total = 0
        score_details = {}
        
        # 1. Score de fiabilité (35% du total)
        # Basé sur la cohérence des données et la fréquence des entretiens
        score_fiabilite = 100
        if alerte_incoherence:
            score_fiabilite -= 30  # Pénalité pour incohérence
        if distance_apres_entretien and distance_apres_entretien > 10000:
            score_fiabilite -= 20  # Pénalité pour entretien tardif
        if nb_entretiens == 0 and distance_totale > 5000:
            score_fiabilite -= 40  # Pénalité pour absence d'entretien
        score_fiabilite = max(0, score_fiabilite)
        score_details['fiabilite'] = score_fiabilite
        score_total += score_fiabilite * 0.35
        
        # 2. Score d'efficacité énergétique (25% du total)
        # Basé sur la consommation moyenne
        if distance_parcourue_courses > 0 and total_litres > 0:
            conso_moyenne = (total_litres * 100) / distance_parcourue_courses
            if conso_moyenne <= 8:  # Excellente consommation
                score_efficacite = 100
            elif conso_moyenne <= 10:  # Bonne consommation
                score_efficacite = 80
            elif conso_moyenne <= 12:  # Consommation moyenne
                score_efficacite = 60
            elif conso_moyenne <= 15:  # Consommation élevée
                score_efficacite = 40
            else:  # Consommation très élevée
                score_efficacite = 20
        else:
            score_efficacite = 50  # Score neutre si pas de données
        score_details['efficacite'] = score_efficacite
        score_total += score_efficacite * 0.25
        
        # 3. Score de rentabilité (25% du total)
        # Basé sur le coût par km
        if distance_parcourue_courses > 0 and budget_total > 0:
            cout_km = budget_total / distance_parcourue_courses
            if cout_km <= 0.5:  # Très rentable
                score_rentabilite = 100
            elif cout_km <= 1.0:  # Rentable
                score_rentabilite = 80
            elif cout_km <= 1.5:  # Moyennement rentable
                score_rentabilite = 60
            elif cout_km <= 2.0:  # Peu rentable
                score_rentabilite = 40
            else:  # Non rentable
                score_rentabilite = 20
        else:
            score_rentabilite = 50  # Score neutre si pas de données
        score_details['rentabilite'] = score_rentabilite
        score_total += score_rentabilite * 0.25
        
        # 4. Score d'utilisation (15% du total)
        # Basé sur l'utilisation du véhicule
        if distance_parcourue_courses > 0:
            # Score basé sur l'utilisation (plus de km = meilleur score)
            utilisation_score = min(100, (distance_parcourue_courses / 1000) * 10)
        else:
            utilisation_score = 0
        score_details['utilisation'] = utilisation_score
        score_total += utilisation_score * 0.15
        
        # Classification du véhicule
        if score_total >= 85:
            classification = "Excellent"
            badge_class = "badge-success"
        elif score_total >= 70:
            classification = "Bon"
            badge_class = "badge-primary"
        elif score_total >= 55:
            classification = "Moyen"
            badge_class = "badge-warning"
        elif score_total >= 40:
            classification = "À améliorer"
            badge_class = "badge-danger"
        else:
            classification = "Critique"
            badge_class = "badge-dark"
        
        # Recommandations basées sur les scores
        recommandations = []
        if score_details['fiabilite'] < 60:
            recommandations.append("Vérifier la cohérence des données kilométriques")
        if distance_apres_entretien and distance_apres_entretien > 10000:
            recommandations.append("Programmer un entretien urgent")
        if score_details['efficacite'] < 60:
            recommandations.append("Vérifier la consommation et l'état du moteur")
        if score_details['rentabilite'] < 60:
            recommandations.append("Optimiser les coûts d'exploitation")
        if score_details['utilisation'] < 30:
            recommandations.append("Augmenter l'utilisation du véhicule")
        
        # Calcul de l'âge du véhicule
        age_vehicule = 0
        if v.date_immatriculation:
            age_vehicule = (timezone.now().date() - v.date_immatriculation.date()).days // 365
        
        # Calcul de la consommation moyenne
        conso_moyenne = (total_litres * 100) / distance_parcourue_courses if distance_parcourue_courses > 0 else 0
        cout_km = budget_total / distance_parcourue_courses if distance_parcourue_courses > 0 else 0
        
        stats_vehicules.append({
            'vehicule': v,
            'distance_totale': distance_totale,
            'distance_apres_entretien': distance_apres_entretien,
            'distance_parcourue_courses': distance_parcourue_courses,
            'distance_estimee_km': distance_estimee_km,
            'alerte_incoherence': alerte_incoherence,
            'nb_entretiens': nb_entretiens,
            'types_entretiens': types_entretiens,
            'cout_entretiens': cout_entretiens,
            'budget_total': budget_total,
            'total_litres': total_litres,
            'total_cout_carburant': total_cout_carburant,
            # Nouvelles métriques de scoring
            'score_total': round(score_total, 1),
            'score_details': score_details,
            'classification': classification,
            'badge_class': badge_class,
            'recommandations': recommandations,
            'age_vehicule': age_vehicule,
            'conso_moyenne': conso_moyenne,
            'cout_km': cout_km,
            'nb_courses': courses_terminees.count(),
            'derniere_course': courses_terminees.order_by('-date_fin').first()
        })
        
        scores_data.append(score_total)

    # Calcul des totaux
    total_distance = sum(v['distance_totale'] for v in stats_vehicules if v['distance_totale'] is not None)
    total_distance_apres_entretien = sum(v['distance_apres_entretien'] for v in stats_vehicules if v['distance_apres_entretien'] is not None)
    total_distance_parcourue_courses = sum(v['distance_parcourue_courses'] for v in stats_vehicules)
    total_distance_estimee_km = sum(v['distance_estimee_km'] for v in stats_vehicules if v['distance_estimee_km'] is not None)
    total_entretiens = sum(v['nb_entretiens'] for v in stats_vehicules)
    total_cout_entretiens = sum(v['cout_entretiens'] for v in stats_vehicules)
    total_consommation = sum(v['total_litres'] for v in stats_vehicules)
    total_cout_carburant = sum(v['total_cout_carburant'] for v in stats_vehicules)
    total_budget = sum(v['budget_total'] for v in stats_vehicules)
    
    # Calcul de la moyenne des scores
    score_moyen = sum(scores_data) / len(scores_data) if scores_data else 0
    
    # Préparation du contexte
    context = {
        'stats_vehicules': stats_vehicules,
        'vehicules': paginator.get_page(page),
        'immatriculation': immatriculation,
        'marque': marque,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'total_distance': total_distance,
        'total_distance_apres_entretien': total_distance_apres_entretien,
        'total_distance_parcourue_courses': total_distance_parcourue_courses,
        'total_distance_estimee_km': total_distance_estimee_km,
        'total_entretiens': total_entretiens,
        'total_cout_entretiens': total_cout_entretiens,
        'total_consommation': total_consommation,
        'total_cout_carburant': total_cout_carburant,
        'total_budget': total_budget,
        'score_moyen': round(score_moyen, 1),
        'date_generation': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'user': request.user,
    }
    
    # Gestion de l'export PDF
    export_format = request.GET.get('export')
    if export_format == 'pdf':
        return generate_pdf_rapport_vehicules(request, stats_vehicules, date_debut, date_fin, context)
    elif export_format == 'excel':
        return generate_excel_rapport_vehicules(request, stats_vehicules, date_debut, date_fin, context)
    
    return render(request, 'rapport/rapport_vehicules.html', context)
    
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_missions(request):
    """Rapport détaillé sur les missions avec système de scoring avancé."""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    chauffeur_id = request.GET.get('chauffeur')
    vehicule_id = request.GET.get('vehicule')
    demandeur_id = request.GET.get('demandeur')
    export_format = request.GET.get('export')

    courses = Course.objects.all()
    
    if not request.user.is_superuser:
        courses = courses.filter(vehicule__etablissement=request.user.etablissement)
    
    if date_debut:
        courses = courses.filter(date_depart__date__gte=date_debut)
    if date_fin:
        courses = courses.filter(date_depart__date__lte=date_fin)
    statut_filter = request.GET.get('statut')
    if statut_filter:
        courses = courses.filter(statut=statut_filter)
    if chauffeur_id:
        courses = courses.filter(chauffeur_id=chauffeur_id)
    if vehicule_id:
        courses = courses.filter(vehicule_id=vehicule_id)
    if demandeur_id:
        courses = courses.filter(demandeur_id=demandeur_id)

    # Calcul de la distance en km et de la durée en heures
    courses = courses.annotate(
        distance_km=ExpressionWrapper(
            F('distance_parcourue'),  # Utilisation du champ distance_parcourue existant
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    
    # Calcul de la durée en heures avec une valeur par défaut de 0
    # On va d'abord annoter avec None, puis calculer en Python
    courses = courses.annotate(
        duree_heures=Value(None, output_field=DecimalField(max_digits=10, decimal_places=2))
    )
    
    # === SYSTÈME DE SCORING AVANCÉ POUR MISSIONS ===
    missions_scored = []
    scores_data = []  # Pour calculer les moyennes de scoring
    
    for course in courses:
        # Calculer la durée en Python pour chaque cours
        if course.date_depart and course.date_fin:
            duree = (course.date_fin - course.date_depart).total_seconds() / 3600.0
            course.duree_heures = round(duree, 2)
        else:
            course.duree_heures = 0
    
        # === CALCUL DU SCORE DE LA MISSION ===
        score_total = 0
        score_details = {}
        
        # 1. Score de ponctualité (30% du total)
        # Basé sur le respect des horaires et la durée de la mission
        score_ponctualite = 100
        if course.date_depart and course.date_fin:
            duree_mission = course.duree_heures
            distance_mission = float(course.distance_parcourue) if course.distance_parcourue else 0
            
            # Calcul de la vitesse moyenne (km/h)
            if duree_mission > 0:
                vitesse_moyenne = distance_mission / duree_mission
                # Score basé sur la vitesse moyenne (objectif : 60-80 km/h)
                if 60 <= vitesse_moyenne <= 80:
                    score_ponctualite = 100
                elif 50 <= vitesse_moyenne < 60 or 80 < vitesse_moyenne <= 90:
                    score_ponctualite = 80
                elif 40 <= vitesse_moyenne < 50 or 90 < vitesse_moyenne <= 100:
                    score_ponctualite = 60
                elif vitesse_moyenne < 40 or vitesse_moyenne > 100:
                    score_ponctualite = 40
            else:
                score_ponctualite = 50  # Score neutre si pas de durée
        else:
            score_ponctualite = 30  # Pénalité si pas de dates
        
        # Pénalité pour missions annulées ou en retard
        if course.statut == 'annulee':
            score_ponctualite -= 50
        elif course.statut == 'en_retard':
            score_ponctualite -= 30
        
        score_ponctualite = max(0, score_ponctualite)
        score_details['ponctualite'] = score_ponctualite
        score_total += score_ponctualite * 0.30
        
        # 2. Score d'efficacité (25% du total)
        # Basé sur la distance parcourue et l'optimisation du trajet
        score_efficacite = 100
        if course.distance_parcourue:
            distance = float(course.distance_parcourue)
            # Score basé sur la distance (missions plus longues = meilleur score)
            if distance >= 100:  # Mission longue
                score_efficacite = 100
            elif distance >= 50:  # Mission moyenne
                score_efficacite = 80
            elif distance >= 20:  # Mission courte
                score_efficacite = 60
            else:  # Très courte mission
                score_efficacite = 40
        else:
            score_efficacite = 0
        
        # Bonus pour missions terminées
        if course.statut == 'terminee':
            score_efficacite += 20
            score_efficacite = min(100, score_efficacite)
        
        score_details['efficacite'] = score_efficacite
        score_total += score_efficacite * 0.25
        
        # 3. Score de rentabilité (25% du total)
        # Basé sur le coût par km et la valeur de la mission
        score_rentabilite = 100
        if course.distance_parcourue:
            distance = float(course.distance_parcourue)
            cout_km = 0.5  # Coût fixe de 0.5$/km
            cout_total = distance * cout_km
            
            # Score basé sur la rentabilité (missions plus rentables = meilleur score)
            if cout_total <= 25:  # Mission très rentable
                score_rentabilite = 100
            elif cout_total <= 50:  # Mission rentable
                score_rentabilite = 80
            elif cout_total <= 100:  # Mission moyennement rentable
                score_rentabilite = 60
            else:  # Mission coûteuse
                score_rentabilite = 40
        else:
            score_rentabilite = 50
        
        score_details['rentabilite'] = score_rentabilite
        score_total += score_rentabilite * 0.25
        
        # 4. Score de qualité (20% du total)
        # Basé sur la satisfaction client et la qualité du service
        score_qualite = 100
        
        # Pénalités pour différents statuts
        if course.statut == 'annulee':
            score_qualite -= 80
        elif course.statut == 'en_retard':
            score_qualite -= 40
        elif course.statut == 'en_cours':
            score_qualite -= 20
        elif course.statut == 'terminee':
            score_qualite += 10  # Bonus pour mission terminée
            score_qualite = min(100, score_qualite)
        
        # Bonus pour missions avec passagers
        if hasattr(course, 'nombre_passagers') and course.nombre_passagers and course.nombre_passagers > 0:
            score_qualite += 10
            score_qualite = min(100, score_qualite)
        
        score_qualite = max(0, score_qualite)
        score_details['qualite'] = score_qualite
        score_total += score_qualite * 0.20
        
        # Classification de la mission
        if score_total >= 85:
            classification = "Excellent"
            badge_class = "badge-success"
        elif score_total >= 70:
            classification = "Bon"
            badge_class = "badge-primary"
        elif score_total >= 55:
            classification = "Moyen"
            badge_class = "badge-warning"
        elif score_total >= 40:
            classification = "À améliorer"
            badge_class = "badge-danger"
        else:
            classification = "Critique"
            badge_class = "badge-dark"
        
        # Recommandations basées sur les scores
        recommandations = []
        if score_details['ponctualite'] < 60:
            recommandations.append("Améliorer la ponctualité et optimiser les trajets")
        if score_details['efficacite'] < 60:
            recommandations.append("Optimiser les distances et les itinéraires")
        if score_details['rentabilite'] < 60:
            recommandations.append("Réduire les coûts d'exploitation")
        if score_details['qualite'] < 60:
            recommandations.append("Améliorer la qualité du service")
        if course.statut == 'annulee':
            recommandations.append("Analyser les causes d'annulation")
        
        # Ajouter les métriques calculées à la mission
        course.score_total = round(score_total, 1)
        course.score_details = score_details
        course.classification = classification
        course.badge_class = badge_class
        course.recommandations = recommandations
        course.cout_km = 0.5  # Coût fixe
        course.cout_total = float(course.distance_parcourue * 0.5) if course.distance_parcourue else 0
        
        missions_scored.append(course)
        scores_data.append(score_total)
    
    total_missions = len(missions_scored)
    total_distance = sum(float(c.distance_km) for c in missions_scored if c.distance_km is not None)
    total_duree = sum(float(c.duree_heures) for c in missions_scored if c.duree_heures is not None)
    # Calcul du coût basé sur la distance (0.5$ par km)
    COUT_PAR_KM = Decimal('0.5')  # 0.5$ par km
    total_cout = float(sum(Decimal(str(c.distance_km)) * COUT_PAR_KM for c in missions_scored if c.distance_km is not None))
    
    # Calcul de la moyenne des scores
    score_moyen = sum(scores_data) / len(scores_data) if scores_data else 0
    
    stats_par_statut = {}
    for statut_choice in Course.STATUS_CHOICES:
        statut_code = statut_choice[0]
        missions_statut = [c for c in missions_scored if c.statut == statut_code]
        count = len(missions_statut)
        distance = sum(float(m.distance_parcourue) for m in missions_statut if m.distance_parcourue is not None)
        duree = sum(float(m.duree_heures) for m in missions_statut if m.duree_heures is not None)
        cout = float(sum(Decimal(str(m.distance_parcourue)) * Decimal('0.5') for m in missions_statut if m.distance_parcourue is not None))  # 0.5$ par km
        
        stats_par_statut[statut_code] = {
            'count': count,
            'distance': distance,
            'duree': duree,
            'cout': cout,
            'label': dict(Course.STATUS_CHOICES)[statut_code]
        }

    if total_missions > 0:
        moyennes = {
            'distance_moyenne': total_distance / total_missions,
            'duree_moyenne': total_duree / total_missions,
            'cout_moyen': total_cout / total_missions,
            'cout_km': total_cout / total_distance if total_distance > 0 else 0,
            'score_moyen': round(score_moyen, 1)
        }
    else:
        moyennes = {
            'distance_moyenne': 0,
            'duree_moyenne': timedelta(),
            'cout_moyen': 0,
            'cout_km': 0,
            'score_moyen': 0
        }

    destinations_data = {}
    for course in missions_scored:
        if not course.destination:
            continue
            
        if course.destination not in destinations_data:
            destinations_data[course.destination] = {
                'count': 0,
                'total_distance': 0,
                'total_cout': 0,
                'score_moyen': 0,
                'scores': []
            }
            
        destinations_data[course.destination]['count'] += 1
        if course.distance_parcourue:
            distance = float(course.distance_parcourue)
            destinations_data[course.destination]['total_distance'] += distance
            destinations_data[course.destination]['total_cout'] = float(Decimal(str(destinations_data[course.destination]['total_cout'])) + (Decimal(str(distance)) * Decimal('0.5')))  # 0.5$ par km
        
        destinations_data[course.destination]['scores'].append(course.score_total)
    
    # Calculer les scores moyens par destination
    for dest_data in destinations_data.values():
        if dest_data['scores']:
            dest_data['score_moyen'] = sum(dest_data['scores']) / len(dest_data['scores'])
    
    # Trier par nombre de missions et prendre les 5 premières
    top_destinations = [
        {
            'destination': dest,
            'count': data['count'],
            'total_distance': data['total_distance'],
            'total_cout': data['total_cout'],
            'score_moyen': round(data['score_moyen'], 1)
        }
        for dest, data in sorted(
            destinations_data.items(),
            key=lambda x: x[1]['count'],
            reverse=True
        )[:5]
    ]

    # Récupérer les statistiques de base des véhicules
    stats_vehicules_base = courses.values('vehicule__immatriculation', 'vehicule__id').annotate(
        count=Count('id'),
        total_distance=Coalesce(Sum('distance_parcourue'), 0, output_field=DecimalField())
    )

    stats_vehicules = []
    for stat_base in stats_vehicules_base:
        vehicule_id_stat = stat_base['vehicule__id']
        ravitaillements = Ravitaillement.objects.filter(vehicule_id=vehicule_id_stat)
        if date_debut:
            ravitaillements = ravitaillements.filter(date_ravitaillement__date__gte=date_debut)
        if date_fin:
            ravitaillements = ravitaillements.filter(date_ravitaillement__date__lte=date_fin)
        
        total_litres = ravitaillements.aggregate(Sum('litres'))['litres__sum'] or 0
        total_distance_for_conso = stat_base['total_distance']
        conso_moyenne = (total_litres * 100 / total_distance_for_conso) if total_litres > 0 and total_distance_for_conso > 0 else None
        
        # Calcul du coût basé sur la distance (0.5$ par km)
        total_cout = float((Decimal(str(stat_base['total_distance'])) * Decimal('0.5')).quantize(Decimal('0.01'))) if stat_base['total_distance'] else 0
        
        stats_vehicules.append({
            'vehicule__immatriculation': stat_base['vehicule__immatriculation'],
            'count': stat_base['count'],
            'total_distance': stat_base['total_distance'],
            'total_cout': total_cout,
            'conso_moyenne': conso_moyenne,
        })
    
    stats_vehicules = sorted(stats_vehicules, key=lambda x: x['count'], reverse=True)

    # Récupérer les statistiques de base des chauffeurs
    stats_chauffeurs_base = courses.values('chauffeur__first_name', 'chauffeur__last_name').annotate(
        count=Count('id'),
        total_distance=Coalesce(Sum('distance_parcourue'), 0, output_field=DecimalField()),
        missions_terminees=Count('id', filter=Q(statut='terminee'))
    ).order_by('-count')
    
    # Calculer le coût pour chaque chauffeur
    stats_chauffeurs = []
    for chauffeur in stats_chauffeurs_base:
        total_cout = float((Decimal(str(chauffeur['total_distance'])) * Decimal('0.5')).quantize(Decimal('0.01'))) if chauffeur['total_distance'] else 0
        stats_chauffeurs.append({
            'chauffeur__first_name': chauffeur['chauffeur__first_name'],
            'chauffeur__last_name': chauffeur['chauffeur__last_name'],
            'count': chauffeur['count'],
            'total_distance': chauffeur['total_distance'],
            'total_cout': total_cout,
            'missions_terminees': chauffeur['missions_terminees']
        })

    if export_format:
        if export_format == 'pdf':
            return generate_pdf_rapport_missions(missions_scored, stats_par_statut, moyennes, top_destinations, stats_vehicules, stats_chauffeurs)
        elif export_format == 'excel':
            return generate_excel_rapport_missions(missions_scored, stats_par_statut, moyennes, top_destinations, stats_vehicules, stats_chauffeurs)

    chauffeurs = Utilisateur.objects.filter(role='chauffeur')
    if not request.user.is_superuser:
        chauffeurs = chauffeurs.filter(etablissement=request.user.etablissement)
    
    vehicules = Vehicule.objects.filter(etablissement=request.user.etablissement) if not request.user.is_superuser else Vehicule.objects.all()
    
    roles_demandeurs = ['demandeur', 'admin', 'dispatch']
    if not request.user.is_superuser:
        demandeurs = Utilisateur.objects.filter(
            etablissement=request.user.etablissement,
            role__in=roles_demandeurs
        )
    else:
        demandeurs = Utilisateur.objects.filter(role__in=roles_demandeurs)

    logger.info(f"Consultation du rapport des missions par {request.user.username}")

    context = {
        'courses': missions_scored,
        'chauffeurs': chauffeurs,
        'vehicules': vehicules,
        'demandeurs': demandeurs,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'selected_statut': statut_filter,
        'selected_chauffeur': chauffeur_id,
        'selected_vehicule': vehicule_id,
        'selected_demandeur': demandeur_id,
        'total_missions': total_missions,
        'total_distance': total_distance,
        'total_duree': total_duree,
        'total_cout': total_cout,
        'score_moyen': round(score_moyen, 1),
        'stats_par_statut': stats_par_statut,
        'moyennes': moyennes,
        'top_destinations': top_destinations,
        'stats_vehicules': stats_vehicules,
        'stats_chauffeurs': stats_chauffeurs,
        'statut_choices': Course.STATUS_CHOICES,
    }
    return render(request, 'rapport/rapport_missions.html', context)

def generate_pdf_rapport_missions(courses, stats_par_statut, moyennes, top_destinations, stats_vehicules, stats_chauffeurs):
    """Génère un rapport PDF des missions en utilisant pdfkit."""
    try:
        # Préparer le contexte pour le template
        context = {
            'courses': courses,
            'stats_par_statut': stats_par_statut,
            'moyennes': moyennes,
            'top_destinations': top_destinations,
            'stats_vehicules': stats_vehicules,
            'stats_chauffeurs': stats_chauffeurs,
            'date_generation': timezone.now().strftime('%d/%m/%Y à %H:%M'),
        }
        
        # Utiliser la nouvelle fonction PDF compatible Python 3.13
        from core.pdf_utils import render_to_pdf
        return render_to_pdf(
            'rapport/rapport_missions_pdf.html',
            context,
            'rapport_missions.pdf'
        )
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération PDF des missions: {e}")
        return HttpResponse(
            f"Erreur lors de la génération du PDF: {str(e)}",
            content_type='text/plain',
            status=500
        )

def generate_excel_rapport_missions(courses, stats_par_statut, moyennes, top_destinations, stats_vehicules, stats_chauffeurs):
    """Génère un rapport Excel des missions."""
    buffer = BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#D7E4BC',
        'border': 1
    })
    
    # Feuille 1: Missions
    worksheet = workbook.add_worksheet('Missions')
    
    # En-têtes
    headers = ['ID', 'Chauffeur', 'Véhicule', 'Destination', 'Date Départ', 'Date Fin', 'Statut', 'Distance', 'Coût']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Données
    for row, course in enumerate(courses, 1):
        worksheet.write(row, 0, course.id)
        worksheet.write(row, 1, f"{course.chauffeur.first_name} {course.chauffeur.last_name}")
        worksheet.write(row, 2, course.vehicule.immatriculation)
        worksheet.write(row, 3, course.destination)
        worksheet.write(row, 4, course.date_depart.strftime('%d/%m/%Y %H:%M') if course.date_depart else '')
        worksheet.write(row, 5, course.date_fin.strftime('%d/%m/%Y %H:%M') if course.date_fin else '')
        worksheet.write(row, 6, course.get_statut_display())
        worksheet.write(row, 7, course.distance_parcourue or 0)
        worksheet.write(row, 8, course.distance_parcourue * 0.5 if course.distance_parcourue else 0)
    
    # Ajuster la largeur des colonnes
    for col in range(len(headers)):
        worksheet.set_column(col, col, 15)
    
    # Feuille 2: Statistiques
    stats_worksheet = workbook.add_worksheet('Statistiques')
    
    # Statistiques par statut
    stats_worksheet.write(0, 0, 'Statistiques par Statut', header_format)
    stats_worksheet.write(1, 0, 'Statut', header_format)
    stats_worksheet.write(1, 1, 'Nombre', header_format)
    stats_worksheet.write(1, 2, 'Distance Totale', header_format)
    stats_worksheet.write(1, 3, 'Coût Total', header_format)
    
    row = 2
    for stats in stats_par_statut.values():
        stats_worksheet.write(row, 0, stats['label'])
        stats_worksheet.write(row, 1, stats['count'])
        stats_worksheet.write(row, 2, stats['distance'])
        stats_worksheet.write(row, 3, stats['cout'])
        row += 1
    
    # Top destinations
    if top_destinations:
        stats_worksheet.write(row + 1, 0, 'Top Destinations', header_format)
        stats_worksheet.write(row + 2, 0, 'Destination', header_format)
        stats_worksheet.write(row + 2, 1, 'Nombre de Missions', header_format)
        stats_worksheet.write(row + 2, 2, 'Distance Totale', header_format)
        stats_worksheet.write(row + 2, 3, 'Coût Total', header_format)
        
        for i, dest in enumerate(top_destinations):
            stats_worksheet.write(row + 3 + i, 0, dest['destination'])
            stats_worksheet.write(row + 3 + i, 1, dest['count'])
            stats_worksheet.write(row + 3 + i, 2, dest['total_distance'])
            stats_worksheet.write(row + 3 + i, 3, dest['total_cout'])
    
    workbook.close()
    buffer.seek(0)
    
    response = FileResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_missions.xlsx"'
    return response

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_entretiens(request):
    """Rapport sur les entretiens."""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    vehicule_id = request.GET.get('vehicule')
    export_format = request.GET.get('export')
    
    entretiens = Entretien.objects.all()
    
    if not request.user.is_superuser:
        entretiens = entretiens.filter(vehicule__etablissement=request.user.etablissement)
    
    if date_debut:
        entretiens = entretiens.filter(date_creation__date__gte=date_debut)
    if date_fin:
        entretiens = entretiens.filter(date_creation__date__lte=date_fin)
    if vehicule_id:
        entretiens = entretiens.filter(vehicule_id=vehicule_id)
    
    # Gestion de l'exportation
    if export_format == 'pdf':
        return generate_pdf_rapport_entretiens(request, entretiens, date_debut, date_fin)
    elif export_format == 'excel':
        return generate_excel_rapport_entretiens(request, entretiens, date_debut, date_fin)
    
    # Si pas d'export, on prépare les données pour l'affichage normal
    total_entretiens = entretiens.count()
    total_cout = entretiens.aggregate(Sum('cout'))['cout__sum'] or 0
    
    stats_par_type = entretiens.values('type_entretien').annotate(
        count=Count('id'),
        total_cout=Sum('cout')
    ).order_by('-count')
    
    stats_par_vehicule = entretiens.values('vehicule__immatriculation').annotate(
        count=Count('id'),
        total_cout=Sum('cout')
    ).order_by('-total_cout')
    
    vehicules = Vehicule.objects.filter(etablissement=request.user.etablissement) if not request.user.is_superuser else Vehicule.objects.all()
    
    context = {
        'entretiens': entretiens,
        'total_entretiens': total_entretiens,
        'total_cout': total_cout,
        'stats_par_type': stats_par_type,
        'stats_par_vehicule': stats_par_vehicule,
        'vehicules': vehicules,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'selected_vehicule': vehicule_id,
    }
    return render(request, 'rapport/rapport_entretiens.html', context)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_carburant(request):
    """Rapport sur les ravitaillements en carburant."""
    # Récupération des paramètres de filtre
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    vehicule_id = request.GET.get('vehicule')
    station = request.GET.get('station')
    montant_min = request.GET.get('montant_min')
    
    # Initialisation du queryset de base
    ravitaillements = Ravitaillement.objects.select_related('vehicule').all()
    
    # Filtrage par établissement si l'utilisateur n'est pas superutilisateur
    if not request.user.is_superuser:
        ravitaillements = ravitaillements.filter(vehicule__etablissement=request.user.etablissement)
    
    # Application des filtres
    if date_debut:
        ravitaillements = ravitaillements.filter(date_ravitaillement__date__gte=date_debut)
    if date_fin:
        ravitaillements = ravitaillements.filter(date_ravitaillement__date__lte=date_fin)
    if vehicule_id:
        ravitaillements = ravitaillements.filter(vehicule_id=vehicule_id)
    if station:
        ravitaillements = ravitaillements.filter(nom_station__icontains=station)
    if montant_min:
        try:
            montant_min = float(montant_min)
            ravitaillements = ravitaillements.filter(cout_total__gte=montant_min)
        except (ValueError, TypeError):
            pass
    
    # Calcul des totaux généraux
    total_ravitaillements = ravitaillements.count()
    total_litres = ravitaillements.aggregate(Sum('litres'))['litres__sum'] or 0
    total_cout = ravitaillements.aggregate(Sum('cout_total'))['cout_total__sum'] or 0
    
    # Calcul des statistiques par véhicule avec plus de détails
    stats_vehicules = []
    vehicules_distincts = ravitaillements.values('vehicule').distinct()
    
    for v in vehicules_distincts:
        vehicule_ravitaillements = ravitaillements.filter(vehicule_id=v['vehicule'])
        vehicule = Vehicule.objects.get(id=v['vehicule'])
        
        # Calcul de la distance totale parcourue
        distance_totale = 0
        consommation_moyenne = 0
        
        if vehicule_ravitaillements.count() > 1:
            # Trier par date pour obtenir l'ordre chronologique
            ravs_ordered = vehicule_ravitaillements.order_by('date_ravitaillement')
            
            # Calculer la distance entre le premier et le dernier ravitaillement
            premier = ravs_ordered.first()
            dernier = ravs_ordered.last()
            
            if premier and dernier and dernier.kilometrage_apres > premier.kilometrage_avant:
                distance_totale = dernier.kilometrage_apres - premier.kilometrage_avant
                
                # Calculer la consommation moyenne (L/100km)
                total_litres_vehicule = ravitaillements.filter(vehicule_id=v['vehicule']).aggregate(Sum('litres'))['litres__sum'] or 0
                if distance_totale > 0:
                    consommation_moyenne = (total_litres_vehicule * 100) / distance_totale
        
        stats_vehicules.append({
            'vehicule': vehicule,
            'nb_ravitaillements': vehicule_ravitaillements.count(),
            'total_litres': vehicule_ravitaillements.aggregate(Sum('litres'))['litres__sum'] or 0,
            'total_cout': vehicule_ravitaillements.aggregate(Sum('cout_total'))['cout_total__sum'] or 0,
            'total_distance': distance_totale,
            'consommation_moyenne': consommation_moyenne
        })
    
    # Calcul des totaux globaux pour les statistiques
    total_litres_stats = sum(v['total_litres'] for v in stats_vehicules)
    total_cout_stats = sum(v['total_cout'] for v in stats_vehicules)
    total_distance_stats = sum(v['total_distance'] for v in stats_vehicules)
    
    # Calcul du coût moyen par litre global
    cout_moyen_par_litre_global = total_cout_stats / total_litres_stats if total_litres_stats > 0 else 0
    
    # Calcul de la consommation moyenne globale (L/100km)
    global_average_consumption_per_100km = 0
    if total_distance_stats > 0 and total_litres_stats > 0:
        global_average_consumption_per_100km = (total_litres_stats * 100) / total_distance_stats
    
    # Récupération de la liste des véhicules pour le filtre
    vehicules = Vehicule.objects.filter(etablissement=request.user.etablissement) if not request.user.is_superuser else Vehicule.objects.all()
    
    # Statistiques par station de ravitaillement
    stats_station = ravitaillements.values('nom_station').annotate(
        count=Count('id'),
        total_litres=Sum('litres'),
        total_cout=Sum('cout_total')
    ).order_by('-total_litres')
    
    context = {
        'ravitaillements': ravitaillements.order_by('-date_ravitaillement'),
        'total_ravitaillements': total_ravitaillements,
        'total_litres': total_litres,
        'total_cout': total_cout,
        'stats_vehicules': stats_vehicules,
        'stats_station': stats_station,
        'vehicules': vehicules,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'selected_vehicule': vehicule_id,
        'selected_station': station,
        'selected_montant_min': montant_min,
        'total_litres_stats': total_litres_stats,
        'total_cout_stats': total_cout_stats,
        'total_distance_stats': total_distance_stats,
        'cout_moyen_par_litre_global': cout_moyen_par_litre_global,
        'global_average_consumption_per_100km': global_average_consumption_per_100km,
    }
    
    # Gestion de l'export
    export_format = request.GET.get('export')
    if export_format == 'pdf':
        return generate_pdf_rapport_carburant(ravitaillements, context)
    elif export_format == 'excel':
        return generate_excel_rapport_carburant(ravitaillements, context)
    
    return render(request, 'rapport/rapport_carburant.html', context)

def generate_pdf_rapport_carburant(ravitaillements, context):
    """Génère un PDF du rapport de carburant."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rapport_carburant.pdf"'
    
    # Création du document PDF
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Titre du rapport
    title = Paragraph("Rapport de Consommation de Carburant", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Période du rapport
    if context.get('date_debut') or context.get('date_fin'):
        periode = f"Période: {context.get('date_debut', 'Début')} au {context.get('date_fin', 'Fin')}"
        elements.append(Paragraph(periode, styles['Normal']))
    
    # Filtres appliqués
    filters = []
    if context.get('selected_vehicule'):
        vehicule = Vehicule.objects.get(id=context['selected_vehicule'])
        filters.append(f"Véhicule: {vehicule.immatriculation}")
    if context.get('selected_station'):
        filters.append(f"Station: {context['selected_station']}")
    if context.get('selected_montant_min'):
        filters.append(f"Montant minimum: {context['selected_montant_min']} $")
    
    if filters:
        elements.append(Paragraph("Filtres appliqués: " + ", ".join(filters), styles['Normal']))
    
    elements.append(Spacer(1, 20))
    
    # Statistiques globales
    elements.append(Paragraph("Statistiques Globales", styles['Heading2']))
    
    # Tableau des statistiques globales
    stats_data = [
        ["Total des ravitaillements", f"{context['total_ravitaillements']}"],
        ["Total des litres consommés", f"{context['total_litres']:.2f} L"],
        ["Coût total", f"{context['total_cout']:.2f} $"],
        ["Consommation moyenne", f"{context.get('global_average_consumption_per_100km', 0):.2f} L/100km"],
    ]
    
    stats_table = Table(stats_data, colWidths=[200, 200])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Détail des ravitaillements
    elements.append(Paragraph("Détail des Ravitaillements", styles['Heading2']))
    
    # En-têtes du tableau de détail
    detail_headers = [
        'Date', 'Véhicule', 'Station', 'Litres', 'Prix/Litre', 'Coût Total',
        'Km Avant', 'Km Après', 'Distance', 'Consommation (L/100km)'
    ]
    
    detail_data = [detail_headers]
    
    # Données du tableau de détail
    for ravitaillement in ravitaillements:
        consommation = (ravitaillement.litres * 100) / (ravitaillement.kilometrage_apres - ravitaillement.kilometrage_avant) if ravitaillement.kilometrage_apres > ravitaillement.kilometrage_avant else 0
        
        detail_data.append([
            ravitaillement.date_ravitaillement.strftime('%Y-%m-%d %H:%M'),
            str(ravitaillement.vehicule),
            ravitaillement.nom_station or '-',
            f"{ravitaillement.litres:.2f}",
            f"{ravitaillement.cout_unitaire:.3f}",
            f"{ravitaillement.cout_total:.2f}",
            ravitaillement.kilometrage_avant or '-',
            ravitaillement.kilometrage_apres or '-',
            f"{ravitaillement.kilometrage_apres - ravitaillement.kilometrage_avant if ravitaillement.kilometrage_apres and ravitaillement.kilometrage_avant else 0}",
            f"{consommation:.2f}" if consommation > 0 else '-',
        ])
    
    # Création du tableau de détail
    col_widths = [80, 80, 100, 50, 50, 60, 50, 50, 50, 60]
    detail_table = Table(detail_data, colWidths=col_widths)
    
    # Style du tableau de détail
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
    ])
    
    # Ajout du style au tableau
    detail_table.setStyle(style)
    
    # Ajout du tableau aux éléments
    elements.append(detail_table)
    
    # Génération du PDF
    doc.build(elements)
    return response

def generate_excel_rapport_carburant(ravitaillements, context):
    """Génère un fichier Excel du rapport de carburant."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rapport_carburant.xlsx'
    
    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Carburant"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # En-tête du rapport
    ws.append(["Rapport de Consommation de Carburant"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Période du rapport
    if context.get('date_debut') or context.get('date_fin'):
        ws.append([f"Période: {context.get('date_debut', 'Début')} au {context.get('date_fin', 'Fin')}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    
    # Filtres appliqués
    filters = []
    if context.get('selected_vehicule'):
        vehicule = Vehicule.objects.get(id=context['selected_vehicule'])
        filters.append(f"Véhicule: {vehicule.immatriculation}")
    if context.get('selected_station'):
        filters.append(f"Station: {context['selected_station']}")
    if context.get('selected_montant_min'):
        filters.append(f"Montant minimum: {context['selected_montant_min']} $")
    
    if filters:
        ws.append(["Filtres appliqués: " + ", ".join(filters)])
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
    
    # Statistiques globales
    ws.append([""])
    ws.append(["Statistiques Globales"])
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    
    stats_data = [
        ["Total des ravitaillements", context['total_ravitaillements']],
        ["Total des litres consommés", f"{context['total_litres']:.2f} L"],
        ["Coût total", f"{context['total_cout']:.2f} $"],
        ["Consommation moyenne", f"{context.get('global_average_consumption_per_100km', 0):.2f} L/100km"],
    ]
    
    for row in stats_data:
        ws.append(row)
    
    # Détail des ravitaillements
    ws.append([""])
    ws.append(["Détail des Ravitaillements"])
    ws.merge_cells(start_row=len(stats_data) + 7, start_column=1, end_row=len(stats_data) + 7, end_column=10)
    
    # En-têtes du tableau de détail
    headers = [
        'Date', 'Véhicule', 'Station', 'Litres', 'Prix/Litre', 'Coût Total',
        'Km Avant', 'Km Après', 'Distance', 'Consommation (L/100km)'
    ]
    
    ws.append(headers)
    
    # Style des en-têtes
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données du tableau de détail
    for ravitaillement in ravitaillements:
        consommation = (ravitaillement.litres * 100) / (ravitaillement.kilometrage_apres - ravitaillement.kilometrage_avant) if ravitaillement.kilometrage_apres > ravitaillement.kilometrage_avant else 0
        
        ws.append([
            ravitaillement.date_ravitaillement.strftime('%Y-%m-%d %H:%M'),
            str(ravitaillement.vehicule),
            ravitaillement.nom_station or '-',
            ravitaillement.litres,
            ravitaillement.cout_unitaire,
            ravitaillement.cout_total,
            ravitaillement.kilometrage_avant or '-',
            ravitaillement.kilometrage_apres or '-',
            ravitaillement.kilometrage_apres - ravitaillement.kilometrage_avant if ravitaillement.kilometrage_apres and ravitaillement.kilometrage_apres > ravitaillement.kilometrage_avant else 0,
            consommation if consommation > 0 else '-',
        ])
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        # Trouver la première cellule non fusionnée
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
                break
        if column_letter:
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Format des nombres
    for row in ws.iter_rows(min_row=ws.max_row - ravitaillements.count() + 1, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '0.00'
    
    for row in ws.iter_rows(min_row=ws.max_row - ravitaillements.count() + 1, max_row=ws.max_row, min_col=9, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    
    # Sauvegarde du classeur
    wb.save(response)
    return response

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_evaluation_chauffeurs(request):
    """Rapport d'évaluation des chauffeurs avec système de scoring avancé."""
    # Récupération des paramètres de filtre
    chauffeur_id = request.GET.get('chauffeur')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Récupération de la liste des chauffeurs pour le menu déroulant
    chauffeurs = Utilisateur.objects.filter(role='chauffeur')
    if not request.user.is_superuser:
        chauffeurs = chauffeurs.filter(etablissement=request.user.etablissement)
    
    # Requête de base pour les courses terminées uniquement, et distance raisonnable
    courses = Course.objects.filter(statut='terminee', distance_parcourue__lte=1000)
    
    # Filtrage par chauffeur
    if chauffeur_id:
        courses = courses.filter(chauffeur_id=chauffeur_id)
    
    # Filtrage par date
    if date_debut:
        courses = courses.filter(date_depart__date__gte=date_debut)
    if date_fin:
        courses = courses.filter(date_fin__date__lte=date_fin)
    
    # Calcul des statistiques des chauffeurs avec les véhicules associés
    stats_chauffeurs = courses.values(
        'chauffeur__id',
        'chauffeur__first_name', 
        'chauffeur__last_name',
        'chauffeur__date_joined',
        'vehicule__immatriculation',
        'vehicule__marque',
        'vehicule__modele'
    ).annotate(
        missions_terminees=Count('id'),
        distance_totale=Coalesce(Sum('distance_parcourue'), 0),
        duree_moyenne=Avg(ExpressionWrapper(
            F('date_fin') - F('date_depart'),
            output_field=DurationField()
        ))
    ).order_by('-missions_terminees')
    
    # Récupération des dépenses de carburant et d'entretien
    for chauffeur in stats_chauffeurs:
        vehicule_id = chauffeur.get('vehicule__id')
        if vehicule_id:
            # Dépenses de carburant et consommation
            ravitaillements = Ravitaillement.objects.filter(
                vehicule_id=vehicule_id,
                date_ravitaillement__date__range=[date_debut, date_fin] if date_debut and date_fin else None
            ).aggregate(
                total_carburant=Coalesce(Sum('cout_total'), 0),
                total_litres=Coalesce(Sum('litres'), 0)
            )
            chauffeur['depenses_carburant'] = ravitaillements['total_carburant']
            chauffeur['litres_consommes'] = ravitaillements['total_litres']
            
            # Dépenses d'entretien
            depenses_entretien = Entretien.objects.filter(
                vehicule_id=vehicule_id,
                date__date__range=[date_debut, date_fin] if date_debut and date_fin else None
            ).aggregate(
                total_entretien=Coalesce(Sum('cout'), 0)
            )
            chauffeur['depenses_entretien'] = depenses_entretien['total_entretien']
            
            # Coût total
            chauffeur['cout_total'] = chauffeur['depenses_carburant'] + chauffeur['depenses_entretien']
        else:
            chauffeur['depenses_carburant'] = 0
            chauffeur['depenses_entretien'] = 0
            chauffeur['cout_total'] = 0
    
    # Récupération des destinations fréquentes par chauffeur
    destinations_par_chauffeur = {}
    for chauffeur in stats_chauffeurs:
        destinations = (
            courses.filter(chauffeur_id=chauffeur['chauffeur__id'])
                  .values('destination')
                  .annotate(count=Count('id'))
                  .order_by('-count')[:3]  # Top 3 destinations
        )
        destinations_par_chauffeur[chauffeur['chauffeur__id']] = [
            f"{d['destination']} ({d['count']}x)" for d in destinations
        ]
    
    # Récupération des notes existantes depuis la session
    notes_chauffeurs = request.session.get('notes_chauffeurs', {})
    
    # Calcul des statistiques supplémentaires et du scoring pour chaque chauffeur
    evaluations = []
    scores_data = []  # Pour calculer les moyennes de scoring
    
    for chauffeur in stats_chauffeurs:
        # Calcul du nombre de jours prestés
        jours_prestes = courses.filter(chauffeur__id=chauffeur['chauffeur__id']).values('date_depart__date').distinct().count()
        # Calcul de l'ancienneté en jours
        anciennete = (timezone.now().date() - chauffeur['chauffeur__date_joined'].date()).days
        # Calcul de la distance moyenne par mission
        distance_moyenne = chauffeur['distance_totale'] / chauffeur['missions_terminees'] if chauffeur['missions_terminees'] > 0 else 0
        # Calcul du nombre de missions par jour
        missions_par_jour = chauffeur['missions_terminees'] / jours_prestes if jours_prestes > 0 else 0
        # Date de la dernière mission terminée
        last_mission = courses.filter(chauffeur__id=chauffeur['chauffeur__id']).order_by('-date_fin').first()
        date_mission = last_mission.date_fin if last_mission and last_mission.date_fin else None
        
        # Calcul de la consommation moyenne
        conso_moyenne = (chauffeur.get('litres_consommes', 0) * 100) / chauffeur['distance_totale'] if chauffeur['distance_totale'] > 0 else 0
        # Calcul du coût par km
        cout_km = chauffeur['distance_totale'] / chauffeur.get('cout_total', 1) if chauffeur.get('cout_total', 0) > 0 else 0
        
        # === SYSTÈME DE SCORING AVANCÉ ===
        score_total = 0
        score_details = {}
        
        # 1. Score de productivité (40% du total)
        if missions_par_jour > 0:
            score_productivite = min(100, (missions_par_jour / 2) * 100)  # 2 missions/jour = 100%
        else:
            score_productivite = 0
        score_details['productivite'] = score_productivite
        score_total += score_productivite * 0.4
        
        # 2. Score d'efficacité énergétique (25% du total)
        if conso_moyenne > 0:
            if conso_moyenne <= 6:  # Excellente consommation
                score_efficacite = 100
            elif conso_moyenne <= 8:  # Bonne consommation
                score_efficacite = 80
            elif conso_moyenne <= 10:  # Consommation moyenne
                score_efficacite = 60
            elif conso_moyenne <= 12:  # Consommation élevée
                score_efficacite = 40
            else:  # Consommation très élevée
                score_efficacite = 20
        else:
            score_efficacite = 0
        score_details['efficacite'] = score_efficacite
        score_total += score_efficacite * 0.25
        
        # 3. Score de rentabilité (20% du total)
        if cout_km > 0:
            if cout_km <= 50:  # Très rentable
                score_rentabilite = 100
            elif cout_km <= 100:  # Rentable
                score_rentabilite = 80
            elif cout_km <= 150:  # Moyennement rentable
                score_rentabilite = 60
            elif cout_km <= 200:  # Peu rentable
                score_rentabilite = 40
            else:  # Non rentable
                score_rentabilite = 20
        else:
            score_rentabilite = 0
        score_details['rentabilite'] = score_rentabilite
        score_total += score_rentabilite * 0.20
        
        # 4. Score de régularité (15% du total)
        if jours_prestes > 0:
            taux_regularite = (jours_prestes / 30) * 100  # Basé sur 30 jours
            score_regularite = min(100, taux_regularite)
        else:
            score_regularite = 0
        score_details['regularite'] = score_regularite
        score_total += score_regularite * 0.15
        
        # Classification du chauffeur
        if score_total >= 85:
            classification = "Excellent"
            badge_class = "badge-success"
        elif score_total >= 70:
            classification = "Bon"
            badge_class = "badge-primary"
        elif score_total >= 55:
            classification = "Moyen"
            badge_class = "badge-warning"
        elif score_total >= 40:
            classification = "À améliorer"
            badge_class = "badge-danger"
        else:
            classification = "Critique"
            badge_class = "badge-dark"
        
        # Recommandations basées sur les scores
        recommandations = []
        if score_details['productivite'] < 60:
            recommandations.append("Augmenter le nombre de missions par jour")
        if score_details['efficacite'] < 60:
            recommandations.append("Améliorer la conduite pour réduire la consommation")
        if score_details['rentabilite'] < 60:
            recommandations.append("Optimiser les coûts d'exploitation")
        if score_details['regularite'] < 60:
            recommandations.append("Améliorer l'assiduité")
        
        chauffeur_id = chauffeur['chauffeur__id']
        evaluations.append({
            'chauffeur': {
                'id': chauffeur_id,
                'first_name': chauffeur['chauffeur__first_name'],
                'last_name': chauffeur['chauffeur__last_name'],
                'get_full_name': f"{chauffeur['chauffeur__first_name']} {chauffeur['chauffeur__last_name']}"
            },
            'vehicule': {
                'immatriculation': chauffeur.get('vehicule__immatriculation', 'Non affecté'),
                'marque': chauffeur.get('vehicule__marque', 'Inconnue'),
                'modele': chauffeur.get('vehicule__modele', '')
            },
            'nb_courses': chauffeur['missions_terminees'],
            'nb_courses_terminees': chauffeur['missions_terminees'],
            'nb_jours_prestes': jours_prestes,
            'distance_totale': chauffeur['distance_totale'],
            'anciennete': anciennete,
            'distance_moyenne': distance_moyenne,
            'missions_par_jour': missions_par_jour,
            'depenses_carburant': chauffeur.get('depenses_carburant', 0),
            'depenses_entretien': chauffeur.get('depenses_entretien', 0),
            'cout_total': chauffeur.get('cout_total', 0),
            'cout_km': cout_km,
            'conso_moyenne': conso_moyenne,
            'top_destinations': destinations_par_chauffeur.get(chauffeur_id, []),
            'notes': notes_chauffeurs.get(str(chauffeur_id), ''),
            'date_mission': date_mission,
            # Nouvelles métriques de scoring
            'score_total': round(score_total, 1),
            'score_details': score_details,
            'classification': classification,
            'badge_class': badge_class,
            'recommandations': recommandations,
            'tendance': 'stable'  # À implémenter avec historique
        })
        
        scores_data.append(score_total)
    
    # Calcul des moyennes globales
    if date_debut and date_fin:
        try:
            date_debut_dt = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d').date()
            total_jours = (date_fin_dt - date_debut_dt).days + 1
        except (ValueError, TypeError):
            total_jours = 1
    else:
        # Si pas de date spécifiée, on prend la période depuis la première mission
        premiere_mission = courses.order_by('date_depart').first()
        if premiere_mission:
            total_jours = (timezone.now().date() - premiere_mission.date_depart.date()).days + 1
        else:
            total_jours = 1
    
    total_missions = sum(eval_['nb_courses'] for eval_ in evaluations)
    total_distance = sum(eval_['distance_totale'] for eval_ in evaluations)
    total_jours_prestes = sum(eval_['nb_jours_prestes'] for eval_ in evaluations if eval_['nb_jours_prestes'] > 0)
    
    # Calcul des totaux pour les moyennes globales
    total_litres_consommes = sum(eval_.get('litres_consommes', 0) for eval_ in evaluations)
    total_cout = sum(eval_.get('cout_total', 0) for eval_ in evaluations)
    
    # Calcul de la moyenne des scores
    score_moyen = sum(scores_data) / len(scores_data) if scores_data else 0
    
    moyennes = {
        'missions_par_jour': total_missions / total_jours if total_jours > 0 else 0,
        'distance_moyenne': total_distance / total_missions if total_missions > 0 else 0,
        'conso_moyenne': (total_litres_consommes * 100) / total_distance if total_distance > 0 else 0,
        'cout_km': total_cout / total_distance if total_distance > 0 else 0,
        'total_chauffeurs': len(evaluations),
        'total_distance': total_distance,
        'moyenne_jours_prestes': total_jours_prestes / len(evaluations) if evaluations else 0,
        'score_moyen': round(score_moyen, 1)
    }
    
    # Pagination
    paginator = Paginator(evaluations, 12)  # 12 évaluations par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'evaluations': page_obj,
        'chauffeurs': chauffeurs,
        'selected_chauffeur': int(chauffeur_id) if chauffeur_id else None,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'moyennes': moyennes,
        'request': request
    }
    
    # Gestion de la sauvegarde des notes
    if request.method == 'POST' and 'save_notes' in request.POST:
        notes_chauffeurs = request.session.get('notes_chauffeurs', {})
        chauffeur_id = request.POST.get('chauffeur_id')
        notes = request.POST.get('notes', '').strip()
        
        if chauffeur_id:
            if notes:
                notes_chauffeurs[chauffeur_id] = notes
            elif chauffeur_id in notes_chauffeurs:
                del notes_chauffeurs[chauffeur_id]
            
            request.session['notes_chauffeurs'] = notes_chauffeurs
            messages.success(request, 'Notes enregistrées avec succès.')
            return redirect(request.path + '?' + request.GET.urlencode())
    
    # Gestion de l'export
    export_format = request.GET.get('export')
    if export_format == 'pdf':
        return generate_pdf_rapport_evaluation_chauffeurs(request, evaluations, context)
    elif export_format == 'excel':
        return generate_excel_rapport_evaluation_chauffeurs(request, evaluations, context)
    
    # Affichage normal du template
    return render(request, 'rapport/rapport_evaluation_chauffeurs.html', context)

def generate_pdf_rapport_evaluation_chauffeurs(request, evaluations, context):
    """Génère un PDF pour le rapport d'évaluation des chauffeurs avec scoring avancé en utilisant pdfkit."""
    try:
        # Préparer le contexte pour le template
        context['evaluations'] = evaluations
        context['date_generation'] = timezone.now().strftime('%d/%m/%Y à %H:%M')
        
        # Utiliser la nouvelle fonction PDF compatible Python 3.13
        from core.pdf_utils import render_to_pdf
        return render_to_pdf(
            'rapport/rapport_evaluation_chauffeurs_pdf.html',
            context,
            'rapport_evaluation_chauffeurs.pdf'
        )
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération PDF des évaluations: {e}")
        return HttpResponse(
            f"Erreur lors de la génération du PDF: {str(e)}",
            content_type='text/plain',
            status=500
        )

def generate_excel_rapport_evaluation_chauffeurs(request, evaluations, context):
    """Génère un fichier Excel pour le rapport d'évaluation des chauffeurs avec scoring avancé."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_evaluation_chauffeurs.xlsx"'
    
    # Création du classeur et de la feuille
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Évaluation chauffeurs")
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#D7E4BC',
        'border': 1
    })
    
    score_format = workbook.add_format({
        'bold': True,
        'fg_color': '#2E86AB',
        'font_color': 'white',
        'border': 1
    })
    
    excellent_format = workbook.add_format({
        'bold': True,
        'fg_color': '#28a745',
        'font_color': 'white',
        'border': 1
    })
    
    bon_format = workbook.add_format({
        'bold': True,
        'fg_color': '#007bff',
        'font_color': 'white',
        'border': 1
    })
    
    moyen_format = workbook.add_format({
        'bold': True,
        'fg_color': '#ffc107',
        'font_color': 'black',
        'border': 1
    })
    
    ameliorer_format = workbook.add_format({
        'bold': True,
        'fg_color': '#dc3545',
        'font_color': 'white',
        'border': 1
    })
    
    critique_format = workbook.add_format({
        'bold': True,
        'fg_color': '#6c757d',
        'font_color': 'white',
        'border': 1
    })
    
    # En-têtes
    headers = [
        'Chauffeur', 'Score', 'Classification', 'Missions terminées', 'Jours prestés', 
        'Distance totale (km)', 'Distance moyenne (km)', 
        'Missions par jour', 'Ancienneté (jours)',
        'Consommation (L/100km)', 'Coût au km (FCFA/km)', 'Top destinations', 'Recommandations'
    ]
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Données
    for row, eval_ in enumerate(evaluations, 1):
        chauffeur_name = f"{eval_['chauffeur']['first_name']} {eval_['chauffeur']['last_name']}"
        score_total = eval_.get('score_total', 0)
        classification = eval_.get('classification', '')
        
        # Format de classification selon le score
        if score_total >= 85:
            class_format = excellent_format
        elif score_total >= 70:
            class_format = bon_format
        elif score_total >= 55:
            class_format = moyen_format
        elif score_total >= 40:
            class_format = ameliorer_format
        else:
            class_format = critique_format
        
        # Top destinations
        destinations = eval_.get('top_destinations', [])
        destinations_str = ', '.join(destinations) if destinations else 'Aucune'
        
        # Recommandations
        recommandations = eval_.get('recommandations', [])
        recommandations_str = ', '.join(recommandations) if recommandations else 'Aucune'
        
        worksheet.write(row, 0, chauffeur_name)
        worksheet.write(row, 1, score_total, score_format)
        worksheet.write(row, 2, classification, class_format)
        worksheet.write(row, 3, eval_['nb_courses'])
        worksheet.write(row, 4, eval_['nb_jours_prestes'])
        worksheet.write(row, 5, eval_['distance_totale'])
        worksheet.write(row, 6, eval_['distance_moyenne'])
        worksheet.write(row, 7, eval_['missions_par_jour'])
        worksheet.write(row, 8, eval_['anciennete'])
        worksheet.write(row, 9, round(eval_.get('conso_moyenne', 0), 2))
        worksheet.write(row, 10, round(eval_.get('cout_km', 0), 2))
        worksheet.write(row, 11, destinations_str)
        worksheet.write(row, 12, recommandations_str)
    
    # Ajustement de la largeur des colonnes
    column_widths = [20, 10, 15, 15, 12, 18, 18, 15, 15, 20, 18, 25, 30]
    for col, width in enumerate(column_widths):
        worksheet.set_column(col, col, width)
    
    # Ajout des moyennes
    row = len(evaluations) + 3
    worksheet.write(row, 0, 'Moyennes', header_format)
    worksheet.write(row + 1, 0, 'Score moyen')
    worksheet.write(row + 1, 1, context['moyennes'].get('score_moyen', 0))
    worksheet.write(row + 2, 0, 'Missions par jour')
    worksheet.write(row + 2, 1, context['moyennes']['missions_par_jour'])
    worksheet.write(row + 3, 0, 'Distance moyenne par mission (km)')
    worksheet.write(row + 3, 1, context['moyennes']['distance_moyenne'])
    worksheet.write(row + 4, 0, 'Consommation moyenne (L/100km)')
    worksheet.write(row + 4, 1, context['moyennes'].get('conso_moyenne', 0))
    worksheet.write(row + 5, 0, 'Coût moyen au km (FCFA/km)')
    worksheet.write(row + 5, 1, context['moyennes'].get('cout_km', 0))
    
    # Ajout d'une feuille avec les détails du scoring
    scoring_sheet = workbook.add_worksheet("Système de scoring")
    
    # En-têtes pour la feuille scoring
    scoring_headers = ['Critère', 'Pondération', 'Objectif', 'Description']
    for col, header in enumerate(scoring_headers):
        scoring_sheet.write(0, col, header, header_format)
    
    # Données du système de scoring
    scoring_data = [
        ['Productivité', '40%', '2 missions/jour = 100%', 'Basé sur le nombre de missions par jour'],
        ['Efficacité énergétique', '25%', '≤6L/100km = 100%', 'Basé sur la consommation moyenne'],
        ['Rentabilité', '20%', '≤50$/km = 100%', 'Basé sur le coût par km'],
        ['Régularité', '15%', '30 jours/mois = 100%', 'Basé sur l\'assiduité']
    ]
    
    for row, data in enumerate(scoring_data, 1):
        for col, value in enumerate(data):
            scoring_sheet.write(row, col, value)
    
    # Ajustement de la largeur des colonnes pour la feuille scoring
    scoring_widths = [20, 15, 25, 40]
    for col, width in enumerate(scoring_widths):
        scoring_sheet.set_column(col, col, width)
    
    # Ajout d'une feuille avec les classifications
    class_sheet = workbook.add_worksheet("Classifications")
    
    # En-têtes pour la feuille classifications
    class_headers = ['Classification', 'Score minimum', 'Score maximum', 'Couleur']
    for col, header in enumerate(class_headers):
        class_sheet.write(0, col, header, header_format)
    
    # Données des classifications
    class_data = [
        ['Excellent', 85, 100, 'Vert'],
        ['Bon', 70, 84, 'Bleu'],
        ['Moyen', 55, 69, 'Jaune'],
        ['À améliorer', 40, 54, 'Rouge'],
        ['Critique', 0, 39, 'Gris']
    ]
    
    for row, data in enumerate(class_data, 1):
        for col, value in enumerate(data):
            class_sheet.write(row, col, value)
    
    # Ajustement de la largeur des colonnes pour la feuille classifications
    class_widths = [15, 15, 15, 15]
    for col, width in enumerate(class_widths):
        class_sheet.set_column(col, col, width)
    
    workbook.close()
    output.seek(0)
    response.write(output.getvalue())
    
    return response

def rapport_demandeurs(request):
    """Rapport sur les demandeurs."""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    courses = Course.objects.all()
    
    if not request.user.is_superuser:
        courses = courses.filter(vehicule__etablissement=request.user.etablissement)
    
    if date_debut:
        courses = courses.filter(date_depart__date__gte=date_debut)
    if date_fin:
        courses = courses.filter(date_depart__date__lte=date_fin)
    
    stats_demandeurs = courses.values('demandeur__first_name', 'demandeur__last_name').annotate(
        demandes_crees=Count('id'),
        demandes_terminees=Count('id', filter=Q(statut='terminee')),
        demandes_annulees=Count('id', filter=Q(statut='annulee'))
    ).order_by('-demandes_crees')
    
    context = {
        'stats_demandeurs': stats_demandeurs,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'rapport/rapport_demandeurs.html', context)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def rapport_depenses_carburant_entretien(request):
    """Rapport combiné sur les dépenses carburant et entretien."""
    from decimal import ROUND_HALF_UP
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    vehicule_id = request.GET.get('vehicule')
    
    # Fonction pour formater les montants en dollars avec 2 décimales
    def format_currency(amount):
        if amount is None:
            return Decimal('0.00')
        if isinstance(amount, (int, float)):
            amount = Decimal(str(amount))
        return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    # Données carburant
    ravitaillements = Ravitaillement.objects.all()
    if not request.user.is_superuser:
        ravitaillements = ravitaillements.filter(vehicule__etablissement=request.user.etablissement)
    if date_debut:
        ravitaillements = ravitaillements.filter(date_ravitaillement__date__gte=date_debut)
    if date_fin:
        ravitaillements = ravitaillements.filter(date_ravitaillement__date__lte=date_fin)
    if vehicule_id:
        ravitaillements = ravitaillements.filter(vehicule_id=vehicule_id)
    
    total_carburant = format_currency(ravitaillements.aggregate(Sum('cout_total'))['cout_total__sum'])
    
    # Données entretien
    entretiens = Entretien.objects.all()
    if not request.user.is_superuser:
        entretiens = entretiens.filter(vehicule__etablissement=request.user.etablissement)
    if date_debut:
        entretiens = entretiens.filter(date_creation__date__gte=date_debut)
    if date_fin:
        entretiens = entretiens.filter(date_creation__date__lte=date_fin)
    if vehicule_id:
        entretiens = entretiens.filter(vehicule_id=vehicule_id)
    
    total_entretien = format_currency(entretiens.aggregate(Sum('cout'))['cout__sum'])
    
    # Calcul du total général
    total_general = total_carburant + total_entretien
    
    # Préparation des statistiques par véhicule
    stats_vehicules = []
    vehicules = Vehicule.objects.filter(etablissement=request.user.etablissement) if not request.user.is_superuser else Vehicule.objects.all()
    
    for vehicule in vehicules:
        ravs_vehicule = ravitaillements.filter(vehicule=vehicule)
        entretiens_vehicule = entretiens.filter(vehicule=vehicule)
        
        cout_carburant = format_currency(ravs_vehicule.aggregate(Sum('cout_total'))['cout_total__sum'])
        cout_entretien = format_currency(entretiens_vehicule.aggregate(Sum('cout'))['cout__sum'])
        cout_total = cout_carburant + cout_entretien
        
        stats_vehicules.append({
            'immatriculation': vehicule.immatriculation,
            'marque_modele': f"{vehicule.marque} {vehicule.modele}",
            'cout_carburant': cout_carburant,
            'cout_entretien': cout_entretien,
            'cout_total': cout_total,
        })
    
    # Tri par coût total décroissant
    stats_vehicules.sort(key=lambda x: x['cout_total'], reverse=True)
    
    # Préparation du contexte
    context = {
        'total_carburant': total_carburant,
        'total_entretien': total_entretien,
        'total_general': total_general,  # Ajout du total général manquant
        'depenses_par_vehicule': stats_vehicules,  # Pour correspondre au template
        'vehicules': vehicules,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'selected_vehicule': vehicule_id,
    }
    
    # Gestion de l'export
    export_format = request.GET.get('format')
    if export_format == 'pdf':
        return generate_pdf_rapport_depenses(request, context)
    elif export_format == 'excel':
        return generate_excel_rapport_depenses(request, context)
    
    return render(request, 'rapport/rapport_depenses_carburant_entretien.html', context)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def generate_pdf_rapport_entretiens(request, entretiens, date_debut=None, date_fin=None):
    """Génère un rapport PDF des entretiens."""
    # Préparer les données pour le template
    data = []
    for entretien in entretiens:
        data.append({
            'id': entretien.id,
            'date': entretien.date_creation.strftime('%d/%m/%Y') if entretien.date_creation else '-',
            'vehicule': entretien.vehicule.immatriculation if entretien.vehicule else '-',
            'type': entretien.get_type_entretien_display(),
            'garage': entretien.garage or '-',
            'cout': f"{entretien.cout} $" if entretien.cout is not None else '-',
            'kilometrage': entretien.kilometrage or '-',
            'statut': entretien.get_statut_display(),
            'commentaires': entretien.commentaires or '-',
            'departement': entretien.vehicule.etablissement.nom if entretien.vehicule and entretien.vehicule.etablissement else '-',
        })
    
    # Calculer les totaux
    total_cout = sum(entretien.cout or 0 for entretien in entretiens)
    
    # Préparer le contexte
    context = {
        'entretiens': data,
        'total_entretiens': len(data),
        'total_cout': total_cout,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'date_generation': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'STATIC_URL': settings.STATIC_URL,
    }
    
    # Encoder le logo en base64
    import base64
    logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_ips_co.png'
    logo_base64 = ''
    
    try:
        with open(logo_path, 'rb') as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
    except Exception as e:
        logger.warning(f"Impossible de charger le logo: {e}")
    
    context['logo_base64'] = logo_base64
    
    # Gérer l'export PDF
    if request.GET.get('export') == 'pdf':
        try:
            # Utiliser les nouvelles fonctions PDF compatibles Python 3.13
            from core.pdf_utils import render_to_pdf
            return render_to_pdf(
                'rapport/rapport_entretiens_pdf.html',
                context,
                'rapport_entretiens.pdf'
            )
        except Exception as e:
            logger.error(f"Erreur lors de la génération PDF: {e}")
            return HttpResponse(
                f"Erreur lors de la génération du PDF: {str(e)}",
                content_type='text/plain',
                status=500
            )
    
    # Rendre le template HTML
    return render(request, 'rapport/rapport_entretiens_pdf.html', context)

@login_required
@user_passes_test(is_admin_or_dispatch_or_superuser)
def generate_excel_rapport_entretiens(request, entretiens, date_debut=None, date_fin=None):
    """Génère un rapport Excel des entretiens."""
    # Créer un objet BytesIO pour le fichier Excel
    output = BytesIO()
    
    # Créer un classeur Excel et ajouter une feuille de calcul
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Rapport Entretiens')
    
    # Définir les formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    
    # Écrire les en-têtes de colonnes
    headers = [
        'ID', 'Date', 'Véhicule', 'Département', 'Type', 'Garage',
        'Coût ($)', 'Kilométrage', 'Statut', 'Commentaires'
    ]
    
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, header_format)
    
    # Écrire les données
    for row_num, entretien in enumerate(entretiens, 1):
        worksheet.write(row_num, 0, entretien.id)
        worksheet.write(row_num, 1, entretien.date_creation.strftime('%d/%m/%Y') if entretien.date_creation else '-')
        worksheet.write(row_num, 2, entretien.vehicule.immatriculation if entretien.vehicule else '-')
        worksheet.write(row_num, 3, entretien.vehicule.etablissement.nom if entretien.vehicule and entretien.vehicule.etablissement else '-')
        worksheet.write(row_num, 4, entretien.get_type_entretien_display())
        worksheet.write(row_num, 5, entretien.garage or '-')
        worksheet.write(row_num, 6, entretien.cout or 0)
        worksheet.write(row_num, 7, entretien.kilometrage or '-')
        worksheet.write(row_num, 8, entretien.get_statut_display())
        worksheet.write(row_num, 9, entretien.commentaires or '-')
    
    # Ajuster la largeur des colonnes
    worksheet.set_column('A:A', 8)  # ID
    worksheet.set_column('B:B', 12)  # Date
    worksheet.set_column('C:C', 15)  # Véhicule
    worksheet.set_column('D:D', 20)  # Département
    worksheet.set_column('E:E', 15)  # Type
    worksheet.set_column('F:F', 20)  # Garagiste
    worksheet.set_column('G:G', 12)  # Coût
    worksheet.set_column('H:H', 12)  # Kilométrage
    worksheet.set_column('I:I', 15)  # Statut
    worksheet.set_column('J:J', 30)  # Commentaires
    
    # Ajouter un filtre
    worksheet.autofilter(0, 0, len(entretiens), len(headers) - 1)
    
    # Fermer le classeur
    workbook.close()
    
    # Préparer la réponse
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=rapport_entretiens.xlsx'
    
    return response

def generate_pdf_rapport_depenses(request, context):
    """Génère un fichier PDF pour le rapport des dépenses en utilisant pdfkit."""
    try:
        # Préparer le contexte pour le template
        context['date_generation'] = timezone.now().strftime('%d/%m/%Y à %H:%M')
        
        # Utiliser la nouvelle fonction PDF compatible Python 3.13
        from core.pdf_utils import render_to_pdf
        return render_to_pdf(
            'rapport/rapport_depenses_pdf.html',
            context,
            'rapport_depenses.pdf'
        )
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération PDF des dépenses: {e}")
        return HttpResponse(
            f"Erreur lors de la génération du PDF: {str(e)}",
            content_type='text/plain',
            status=500
        )
        pass  # Code reportlab supprimé, remplacé par pdfkit

def generate_excel_rapport_depenses(request, context):
    """Génère un fichier Excel pour le rapport des dépenses."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rapport_depenses.xlsx"'
    
    # Création du classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dépenses"
    
    # Styles
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    number_format = '0.00'
    
    # En-tête
    ws['A1'] = "Rapport des Dépenses Carburant & Entretien"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Période
    date_debut = context.get('date_debut', 'début')
    date_fin = context.get('date_fin', "aujourd'hui")
    ws['A2'] = f"Période: du {date_debut} au {date_fin}"
    ws.merge_cells('A2:E2')
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Ligne vide
    ws.append([])
    
    # En-têtes du tableau des totaux
    ws.append(['Total Carburant ($)', 'Total Entretien ($)', 'Total Général ($)'])
    for col in ['A', 'B', 'C']:
        ws[f'{col}4'].fill = header_fill
        ws[f'{col}4'].font = header_font
        ws[f'{col}4'].alignment = header_alignment
    
    # Données des totaux
    ws.append([
        float(context['total_carburant']),
        float(context['total_entretien']),
        float(context['total_general'])
    ])
    
    # Formatage des nombres dans le tableau des totaux
    for col in ['A', 'B', 'C']:
        ws.column_dimensions[col].width = 20
        ws[f'{col}5'].number_format = number_format
    
    # Ligne vide
    ws.append([])
    
    # En-têtes du tableau détaillé
    if context['depenses_par_vehicule']:
        ws.append([
            'Immatriculation', 
            'Marque/Modèle', 
            'Carburant ($)', 
            'Entretien ($)', 
            'Total ($)'
        ])
        
        # Style des en-têtes
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}7'].fill = header_fill
            ws[f'{col}7'].font = header_font
            ws[f'{col}7'].alignment = header_alignment
        
        # Données détaillées
        row_num = 8
        for item in context['depenses_par_vehicule']:
            ws.append([
                item['immatriculation'],
                item['marque_modele'],
                float(item['cout_carburant']),
                float(item['cout_entretien']),
                float(item['cout_total'])
            ])
            
            # Formatage des nombres
            for col in ['C', 'D', 'E']:
                ws[f'{col}{row_num}'].number_format = number_format
            
            # Alternance des couleurs de ligne
            if row_num % 2 == 0:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row_num}'].fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            row_num += 1
        
        # Ajustement de la largeur des colonnes
        ws.column_dimensions['A'].width = 15  # Immatriculation
        ws.column_dimensions['B'].width = 25  # Marque/Modèle
        ws.column_dimensions['C'].width = 15  # Carburant
        ws.column_dimensions['D'].width = 15  # Entretien
        ws.column_dimensions['E'].width = 15  # Total
    
    # Sauvegarde du classeur
    wb.save(response)
    return response

def generate_pdf_rapport_vehicules(request, vehicules, date_debut, date_fin, context):
    """Génère un rapport PDF des véhicules."""
    # Encoder le logo en base64
    import base64
    
    logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_ips_co.png'
    logo_base64 = ''
    
    try:
        with open(logo_path, 'rb') as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
    except Exception as e:
        logger.warning(f"Impossible de charger le logo: {e}")
    
    context['logo_base64'] = logo_base64
    
    # Gérer l'export PDF
    if request.GET.get('export') == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_vehicules.pdf"'
        
        # Options pour la génération PDF
        options = {
            'encoding': 'UTF-8',
            'quiet': True,
            'enable-javascript': True,
            'javascript-delay': 1000,
            'no-stop-slow-scripts': True,
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '15mm',
            'margin-right': '10mm',
            'margin-bottom': '15mm',
            'margin-left': '10mm',
            'no-outline': None,
            'enable-local-file-access': ''
        }
        
        # Générer le PDF
        try:
            # Utiliser les nouvelles fonctions PDF compatibles Python 3.13
            from core.pdf_utils import render_to_pdf
            return render_to_pdf(
                'rapport/rapport_vehicules_pdf.html',
                context,
                'rapport_vehicules.pdf'
            )
        except Exception as e:
            logger.error(f"Erreur lors de la génération PDF: {e}")
            return HttpResponse(
                f"Erreur lors de la génération du PDF: {str(e)}",
                content_type='text/plain',
                status=500
            )
    
    # Rendre le template HTML
    return render(request, 'rapport/rapport_vehicules_pdf.html', context)

def generate_excel_rapport_vehicules(request, vehicules, date_debut, date_fin, context):
    """Génère un rapport Excel des véhicules avec scoring avancé."""
    # Créer un objet BytesIO pour le fichier Excel
    output = BytesIO()
    
    # Créer un classeur Excel et ajouter une feuille de calcul
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Rapport Véhicules')
    
    # Définir les formats
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#4472C4',
        'font_color': 'white',
        'border': 1,
        'font_size': 10,
    })
    
    score_format = workbook.add_format({
        'bold': True,
        'fg_color': '#2E86AB',
        'font_color': 'white',
        'border': 1
    })
    
    excellent_format = workbook.add_format({
        'bold': True,
        'fg_color': '#28a745',
        'font_color': 'white',
        'border': 1
    })
    
    bon_format = workbook.add_format({
        'bold': True,
        'fg_color': '#007bff',
        'font_color': 'white',
        'border': 1
    })
    
    moyen_format = workbook.add_format({
        'bold': True,
        'fg_color': '#ffc107',
        'font_color': 'black',
        'border': 1
    })
    
    ameliorer_format = workbook.add_format({
        'bold': True,
        'fg_color': '#dc3545',
        'font_color': 'white',
        'border': 1
    })
    
    critique_format = workbook.add_format({
        'bold': True,
        'fg_color': '#6c757d',
        'font_color': 'white',
        'border': 1
    })
    
    # Écrire les en-têtes
    headers = [
        'Véhicule', 'Immatriculation', 'Score', 'Classification', 'Âge (ans)', 'Courses',
        'Distance parcourue (courses)', 'Distance estimée (km)', 'Alerte',
        'Distance totale (km)', 'Distance depuis dernier entretien (km)',
        'Nb entretiens', 'Types entretiens', 'Coût entretiens (FCFA)',
        'Consommation (L)', 'Consommation (L/100km)', 'Coût carburant (FCFA)', 
        'Coût/km ($/km)', 'Budget total (FCFA)', 'Recommandations'
    ]
    
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, header_format)
    
    # Écrire les données
    for row_num, stat in enumerate(vehicules, 1):
        vehicule_name = f"{stat['vehicule'].marque} {stat['vehicule'].modele}"
        score_total = stat.get('score_total', 0)
        classification = stat.get('classification', '')
        
        # Format de classification selon le score
        if score_total >= 85:
            class_format = excellent_format
        elif score_total >= 70:
            class_format = bon_format
        elif score_total >= 55:
            class_format = moyen_format
        elif score_total >= 40:
            class_format = ameliorer_format
        else:
            class_format = critique_format
        
        # Recommandations
        recommandations = stat.get('recommandations', [])
        recommandations_str = ', '.join(recommandations) if recommandations else 'Aucune'
        
        worksheet.write(row_num, 0, vehicule_name)
        worksheet.write(row_num, 1, stat['vehicule'].immatriculation)
        worksheet.write(row_num, 2, score_total, score_format)
        worksheet.write(row_num, 3, classification, class_format)
        worksheet.write(row_num, 4, stat.get('age_vehicule', 0))
        worksheet.write(row_num, 5, stat.get('nb_courses', 0))
        worksheet.write_number(row_num, 6, stat['distance_parcourue_courses'])
        worksheet.write_number(row_num, 7, stat['distance_estimee_km'] if stat['distance_estimee_km'] is not None else 0)
        worksheet.write(row_num, 8, '⚠️ Incohérence' if stat['alerte_incoherence'] else '')
        worksheet.write_number(row_num, 9, stat['distance_totale'])
        worksheet.write_number(row_num, 10, stat['distance_apres_entretien'] if stat['distance_apres_entretien'] is not None else 0)
        worksheet.write_number(row_num, 11, stat['nb_entretiens'])
        worksheet.write(row_num, 12, stat.get('types_entretiens', 'Aucun'))
        worksheet.write_number(row_num, 13, stat.get('cout_entretiens', 0))
        worksheet.write_number(row_num, 14, stat.get('total_litres', 0))
        worksheet.write_number(row_num, 15, stat.get('conso_moyenne', 0))
        worksheet.write_number(row_num, 16, stat.get('total_cout_carburant', 0))
        worksheet.write_number(row_num, 17, stat.get('cout_km', 0))
        worksheet.write_number(row_num, 18, stat.get('budget_total', 0))
        worksheet.write(row_num, 19, recommandations_str)
    
    # Ajouter les totaux
    row_num = len(vehicules) + 1
    worksheet.write(row_num, 0, 'TOTAUX', header_format)
    worksheet.write_formula(row_num, 6, f'SUM(G2:G{row_num})')
    worksheet.write_formula(row_num, 7, f'SUM(H2:H{row_num})')
    worksheet.write_formula(row_num, 9, f'SUM(J2:J{row_num})')
    worksheet.write_formula(row_num, 10, f'SUM(K2:K{row_num})')
    worksheet.write_formula(row_num, 11, f'SUM(L2:L{row_num})')
    worksheet.write_formula(row_num, 13, f'SUM(N2:N{row_num})')
    worksheet.write_formula(row_num, 14, f'SUM(O2:O{row_num})')
    worksheet.write_formula(row_num, 16, f'SUM(Q2:Q{row_num})')
    worksheet.write_formula(row_num, 18, f'SUM(S2:S{row_num})')
    
    # Ajuster la largeur des colonnes
    column_widths = [30, 15, 10, 15, 10, 10, 22, 22, 12, 18, 22, 15, 25, 20, 15, 20, 20, 15, 20, 40]
    for i, width in enumerate(column_widths):
        worksheet.set_column(i, i, width)
    
    # Ajout d'une feuille avec les détails du scoring
    scoring_sheet = workbook.add_worksheet("Système de scoring")
    
    # En-têtes pour la feuille scoring
    scoring_headers = ['Critère', 'Pondération', 'Objectif', 'Description']
    for col, header in enumerate(scoring_headers):
        scoring_sheet.write(0, col, header, header_format)
    
    # Données du système de scoring
    scoring_data = [
        ['Fiabilité', '35%', 'Cohérence des données et entretiens réguliers', 'Basé sur la cohérence des données et la fréquence des entretiens'],
        ['Efficacité énergétique', '25%', '≤8L/100km = 100%', 'Basé sur la consommation moyenne'],
        ['Rentabilité', '25%', '≤0.5$/km = 100%', 'Basé sur le coût par km'],
        ['Utilisation', '15%', 'Plus de km = meilleur score', 'Basé sur l\'utilisation du véhicule']
    ]
    
    for row, data in enumerate(scoring_data, 1):
        for col, value in enumerate(data):
            scoring_sheet.write(row, col, value)
    
    # Ajustement de la largeur des colonnes pour la feuille scoring
    scoring_widths = [20, 15, 35, 50]
    for col, width in enumerate(scoring_widths):
        scoring_sheet.set_column(col, col, width)
    
    # Ajout d'une feuille avec les classifications
    class_sheet = workbook.add_worksheet("Classifications")
    
    # En-têtes pour la feuille classifications
    class_headers = ['Classification', 'Score minimum', 'Score maximum', 'Couleur']
    for col, header in enumerate(class_headers):
        class_sheet.write(0, col, header, header_format)
    
    # Données des classifications
    class_data = [
        ['Excellent', 85, 100, 'Vert'],
        ['Bon', 70, 84, 'Bleu'],
        ['Moyen', 55, 69, 'Jaune'],
        ['À améliorer', 40, 54, 'Rouge'],
        ['Critique', 0, 39, 'Gris']
    ]
    
    for row, data in enumerate(class_data, 1):
        for col, value in enumerate(data):
            class_sheet.write(row, col, value)
    
    # Ajustement de la largeur des colonnes pour la feuille classifications
    class_widths = [15, 15, 15, 15]
    for col, width in enumerate(class_widths):
        class_sheet.set_column(col, col, width)
    
    # Fermer le classeur
    workbook.close()
    
    # Préparer la réponse
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=rapport_vehicules.xlsx'
    
    return response

def generer_rapport(request, type_rapport):
    """Génère un rapport selon le type spécifié."""
    if type_rapport == 'missions':
        return rapport_missions(request)
    elif type_rapport == 'vehicules':
        return rapport_vehicules(request)
    elif type_rapport == 'entretiens':
        return rapport_entretiens(request)
    elif type_rapport == 'carburant':
        return rapport_carburant(request)
    elif type_rapport == 'evaluations':
        return rapport_evaluation_chauffeurs(request)
    elif type_rapport == 'demandeurs':
        return rapport_demandeurs(request)
    elif type_rapport == 'depenses':
        return rapport_depenses_carburant_entretien(request)
    else:
        # Redirection vers le tableau de bord si le type n'est pas reconnu
        return redirect('rapport:dashboard')

@login_required
def rapport_journalier_flotte(request):
    today = timezone.localdate()
    vehicules = Vehicule.objects.all()
    data = []
    total_distance = 0
    total_missions = 0
    total_litres = 0
    total_depense = 0
    total_maintenance = 0
    vehicules_labels = []
    vehicules_distances = []
    for v in vehicules:
        courses = Course.objects.filter(vehicule=v, date_souhaitee__date=today)
        distance = sum([(c.kilometrage_fin or 0) - (c.kilometrage_depart or 0) for c in courses if c.kilometrage_depart and c.kilometrage_fin])
        missions = courses.count()
        ravs = Ravitaillement.objects.filter(vehicule=v, date_ravitaillement__date=today)
        total_litres_v = sum([r.quantite for r in ravs])
        total_depense_v = sum([r.montant for r in ravs])
        entretiens = Entretien.objects.filter(vehicule=v, date_entretien__date=today)
        total_maintenance_v = sum([e.cout for e in entretiens if hasattr(e, 'cout') and e.cout])
        data.append({
            'vehicule': v,
            'chauffeur': courses.first().chauffeur if courses.exists() else None,
            'distance': distance,
            'missions': missions,
            'litres': total_litres_v,
            'depense_carburant': total_depense_v,
            'maintenance': total_maintenance_v,
            'obs': '',
        })
        total_distance += distance
        total_missions += missions
        total_litres += total_litres_v
        total_depense += total_depense_v
        total_maintenance += total_maintenance_v
        vehicules_labels.append(f"{v.marque} {v.modele}")
        vehicules_distances.append(distance)
    # Générer le graphique (histogramme) - Temporairement désactivé pour le déploiement
    # plt.figure(figsize=(8, 4))
    # plt.bar(vehicules_labels, vehicules_distances, color='#3498db')
    # plt.xlabel('Véhicule')
    # plt.ylabel('Distance parcourue (km)')
    # plt.title('Distance parcourue par véhicule - ' + today.strftime('%d/%m/%Y'))
    # plt.xticks(rotation=30, ha='right')
    # plt.tight_layout()
    # buf = io.BytesIO()
    # plt.savefig(buf, format='png')
    # plt.close()
    # buf.seek(0)
    # graph_base64 = base64.b64encode(buf.read()).decode('utf-8')
    # Temporairement désactivé - matplotlib en cours de configuration
    graph_base64 = ""
    context = {
        'data': data,
        'date': today,
        'total_distance': total_distance,
        'total_missions': total_missions,
        'total_litres': total_litres,
        'total_depense': total_depense,
        'total_maintenance': total_maintenance,
        'graph_base64': graph_base64,
    }
    
    # Gérer l'export PDF
    if request.GET.get('export') == 'pdf':
        try:
            # Utiliser les nouvelles fonctions PDF compatibles Python 3.13
            from core.pdf_utils import render_to_pdf
            return render_to_pdf(
                'rapport/rapport_journalier_flotte.html',
                context,
                f'rapport_flotte_{today}.pdf'
            )
        except Exception as e:
            logger.error(f"Erreur lors de la génération PDF: {e}")
            return HttpResponse(
                f"Erreur lors de la génération du PDF: {str(e)}",
                content_type='text/plain',
                status=500
            )
    
    # Rendre le template HTML
    return render(request, 'rapport/rapport_journalier_flotte.html', context)

@login_required
@user_passes_test(lambda u: u.is_authenticated and (u.role in ['admin', 'gestionnaire'] or u.is_superuser))
def rapport_vehicule_advanced(request):
    """Rapport avancé pour un véhicule (synthèse, stats, historiques, alertes, graphiques)"""
    vehicule_id = request.GET.get('vehicule')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    vehicules = Vehicule.objects.all().order_by('immatriculation')
    selected_vehicule = None
    if vehicule_id:
        selected_vehicule = Vehicule.objects.filter(id=vehicule_id).first()
    if not selected_vehicule and vehicules.exists():
        selected_vehicule = vehicules.first()
    # Filtres sur les historiques
    entretiens = Entretien.objects.filter(vehicule=selected_vehicule)
    ravs = Ravitaillement.objects.filter(vehicule=selected_vehicule)
    missions = Course.objects.filter(vehicule=selected_vehicule)
    if date_debut:
        entretiens = entretiens.filter(date_creation__date__gte=date_debut)
        ravs = ravs.filter(date_ravitaillement__date__gte=date_debut)
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        entretiens = entretiens.filter(date_creation__date__lte=date_fin)
        ravs = ravs.filter(date_ravitaillement__date__lte=date_fin)
        missions = missions.filter(date_validation__date__lte=date_fin)
    # Synthèse
    km_actuel = selected_vehicule.kilometrage_actuel if selected_vehicule else 0
    missions_count = missions.count()
    distance_totale = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    total_litres = ravs.aggregate(l=Sum('litres'))['l'] or 0
    conso_moyenne = round((total_litres * 100 / distance_totale), 2) if distance_totale > 0 and total_litres > 0 else None
    cout_total_carburant = ravs.aggregate(c=Sum('cout_total'))['c'] or 0
    incidents = missions.filter(motif__icontains='incident').count()
    prochain_entretien = entretiens.order_by('-date_entretien').first()
    # Alertes
    alertes = []
    if prochain_entretien and km_actuel - (prochain_entretien.kilometrage_apres or 0) > 4500:
        alertes.append("Entretien en retard de plus de 4500 km !")
    if conso_moyenne and conso_moyenne > 8:
        alertes.append(f"Consommation élevée : {conso_moyenne} L/100km")
    # Historique entretiens/ravitaillements
    entretiens_list = entretiens.order_by('-date_entretien')[:10]
    ravs_list = ravs.order_by('-date_ravitaillement')[:10]
    # Préparer les données pour graphiques (mois)
    from django.db.models.functions import TruncMonth
    stats_mois_missions = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    stats_mois_ravs = ravs.annotate(mois=TruncMonth('date_ravitaillement')).values('mois').annotate(
        litres=Sum('litres'),
    ).order_by('mois')
    # Fusionner les stats par mois (missions et ravitaillements)
    stats_mois = []
    mois_set = set([s['mois'] for s in stats_mois_missions] + [s['mois'] for s in stats_mois_ravs])
    for mois in sorted(mois_set):
        missions_data = next((s for s in stats_mois_missions if s['mois'] == mois), {})
        ravs_data = next((s for s in stats_mois_ravs if s['mois'] == mois), {})
        stats_mois.append({
            'mois': mois,
            'missions': missions_data.get('missions', 0),
            'km': missions_data.get('km', 0),
            'litres': ravs_data.get('litres', 0),
        })
    # Historique détaillé fusionné
    historique_corrections = HistoriqueKilometrage.objects.filter(vehicule=selected_vehicule).order_by('-date_modification')
    # Incidents = missions dont motif contient 'incident'
    incidents_list = missions.filter(motif__icontains='incident').order_by('-date_validation')
    # Fusionner tous les événements dans une liste chronologique
    historique_detaille = []
    for e in entretiens:
        historique_detaille.append({
            'type': 'entretien',
            'date': e.date_entretien,
            'libelle': f"Entretien : {e.get_type_entretien_display()}",
            'km': e.kilometrage,
            'cout': e.cout,
            'auteur': e.createur.get_full_name() if hasattr(e, 'createur') else '',
            'obs': e.observations or '',
        })
    for r in ravs:
        historique_detaille.append({
            'type': 'ravitaillement',
            'date': r.date_ravitaillement,
            'libelle': f"Ravitaillement ({r.litres} L)",
            'km': r.kilometrage_apres,
            'cout': r.cout_total,
            'auteur': '',
            'obs': r.station or '',
        })
    for c in historique_corrections:
        historique_detaille.append({
            'type': 'correction_km',
            'date': c.date_modification,
            'libelle': f"Correction km : {c.valeur_avant} → {c.valeur_apres}",
            'km': c.valeur_apres,
            'cout': '',
            'auteur': c.utilisateur.get_full_name() if c.utilisateur else '',
            'obs': c.commentaire or '',
        })
    for i in incidents_list:
        historique_detaille.append({
            'type': 'incident',
            'date': i.date_validation,
            'libelle': f"Incident : {i.motif[:40]}...",
            'km': i.kilometrage_fin,
            'cout': '',
            'auteur': i.chauffeur.get_full_name() if i.chauffeur else '',
            'obs': '',
        })
    historique_detaille = sorted(historique_detaille, key=lambda x: x['date'] or timezone.now(), reverse=True)
    # Filtres dynamiques sur l'historique détaillé
    type_filters = []
    if request.GET.get('type_entretien'): type_filters.append('entretien')
    if request.GET.get('type_ravitaillement'): type_filters.append('ravitaillement')
    if request.GET.get('type_correction_km'): type_filters.append('correction_km')
    if request.GET.get('type_incident'): type_filters.append('incident')
    periode_debut = request.GET.get('periode_debut')
    periode_fin = request.GET.get('periode_fin')
    recherche = request.GET.get('recherche', '').strip().lower()
    filtered_historique = []
    for h in historique_detaille:
        if type_filters and h['type'] not in type_filters:
            continue
        if periode_debut and h['date'] and h['date'] < datetime.strptime(periode_debut, '%Y-%m-%d').date():
            continue
        if periode_fin and h['date'] and h['date'] > datetime.strptime(periode_fin, '%Y-%m-%d').date():
            continue
        if recherche and recherche not in (str(h['libelle']).lower() + str(h['auteur']).lower() + str(h['obs']).lower()):
            continue
        filtered_historique.append(h)
    # Si aucun filtre, on affiche tout
    if not (type_filters or periode_debut or periode_fin or recherche):
        filtered_historique = historique_detaille
    tri = request.GET.get('tri', 'date')
    ordre = request.GET.get('ordre', 'desc')
    def get_sort_key(h):
        if tri == 'date': return h['date'] or timezone.now()
        if tri == 'type': return h['type'] or ''
        if tri == 'km': return h['km'] or 0
        if tri == 'cout': return float(h['cout']) if h['cout'] else 0
        if tri == 'auteur': return h['auteur'] or ''
        return h['date'] or timezone.now()
    filtered_historique = sorted(filtered_historique, key=get_sort_key, reverse=(ordre != 'asc'))
    # Pagination de l'historique détaillé filtré
    page = int(request.GET.get('page', 1))
    page_size = 20
    total_items = len(filtered_historique)
    total_pages = (total_items + page_size - 1) // page_size
    start = (page - 1) * page_size
    end = start + page_size
    historique_page = filtered_historique[start:end]
    context = {
        'vehicules': vehicules,
        'selected_vehicule': selected_vehicule,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'km_actuel': km_actuel,
        'missions_count': missions_count,
        'distance_totale': distance_totale,
        'conso_moyenne': conso_moyenne,
        'cout_total_carburant': cout_total_carburant,
        'incidents': incidents,
        'prochain_entretien': prochain_entretien,
        'alertes': alertes,
        'entretiens_list': entretiens_list,
        'ravs_list': ravs_list,
        'stats_mois': stats_mois,
        'historique_detaille': historique_page,
        'all_historique_detaille': filtered_historique,
        'page': page,
        'total_pages': total_pages,
        'total_items': total_items,
        'page_size': page_size,
        'page_range': range(1, total_pages+1),
    }
    
    # Gestion des exports
    export_format = request.GET.get('export')
    export_historique = request.GET.get('export_historique')
    
    if export_format:
        if export_historique:
            # Export de l'historique détaillé
            if export_format == 'pdf':
                return export_historique_detaille_pdf(request, context)
            elif export_format == 'excel':
                return export_historique_detaille_excel(request, context)
            elif export_format == 'csv':
                return export_historique_detaille_csv(request, context)
        else:
            # Export du rapport principal
            if export_format == 'pdf':
                return export_rapport_vehicule_advanced_pdf(request, context)
            elif export_format == 'excel':
                return export_rapport_vehicule_advanced_excel(request, context)
    
    return render(request, 'rapport/rapport_vehicule_advanced.html', context)

def export_rapport_vehicule_advanced_pdf(request, context):
    # from weasyprint import HTML  # Temporairement commenté pour le déploiement
    from django.template.loader import render_to_string
    from django.utils import timezone
    
    # Déterminer l'orientation selon le nombre de données
    total_items = len(context.get('all_historique_detaille', []))
    orientation = 'landscape' if total_items > 15 else 'portrait'
    
    # Créer un contexte enrichi pour l'export
    export_context = {
        **context,
        'is_export': True,
        'orientation': orientation,
        'date_export': timezone.now().strftime('%d/%m/%Y à %H:%M'),
        'user_export': request.user.get_full_name() or request.user.username,
    }
    
    html_string = render_to_string('rapport/rapport_vehicule_advanced_pdf.html', export_context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    response = HttpResponse(content_type='application/pdf')
    vehicule = context.get("selected_vehicule")
    immatriculation = vehicule.immatriculation if vehicule else "vehicule"
    response['Content-Disposition'] = f'attachment; filename="rapport_vehicule_avance_{immatriculation}.pdf"'
    html.write_pdf(response)
    return response

def export_rapport_vehicule_advanced_excel(request, context):
    import xlsxwriter
    import io
    from django.utils import timezone
    import os
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Rapport Véhicule Avancé')
    # Insérer le logo si possible
    logo_path = os.path.join('static', 'img', 'logo.png')
    if os.path.exists(logo_path):
        worksheet.insert_image('A1', logo_path, {'x_scale': 0.5, 'y_scale': 0.5, 'x_offset': 0, 'y_offset': 0})
        start_row = 5
    else:
        start_row = 1
    # Styles sobres
    header_format = workbook.add_format({'bold': True,'font_size': 11,'font_color': 'white','bg_color': '#4F81BD','border': 1,'align': 'center','valign': 'vcenter'})
    subheader_format = workbook.add_format({'bold': True,'font_size': 10,'bg_color': '#D9E1F2','border': 1,'align': 'center'})
    data_format = workbook.add_format({'font_size': 10,'border': 1,'align': 'left','bg_color': '#FFFFFF'})
    number_format = workbook.add_format({'font_size': 10,'border': 1,'align': 'right','num_format': '#,##0','bg_color': '#FFFFFF'})
    currency_format = workbook.add_format({'font_size': 10,'border': 1,'align': 'right','num_format': '#,##0.00 $','bg_color': '#FFFFFF'})
    v = context.get('selected_vehicule')
    vehicule_info = f"{v.immatriculation} - {v.marque} {v.modele}" if v else 'Véhicule non sélectionné'
    worksheet.merge_range(start_row, 0, start_row, 8, f'RAPPORT VÉHICULE AVANCÉ - {vehicule_info}', header_format)
    worksheet.merge_range(start_row+1, 0, start_row+1, 8, f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M")} par {request.user.get_full_name() or request.user.username}', data_format)
    worksheet.merge_range(start_row+3, 0, start_row+3, 8, 'SYNTHÈSE DU VÉHICULE', subheader_format)
    headers = ['Kilométrage actuel', 'Missions', 'Km parcourus', 'Conso. (L/100km)', 'Coût carburant', 'Incidents', 'Prochain entretien', 'Alertes']
    for col, h in enumerate(headers):
        worksheet.write(start_row+4, col, h, header_format)
    worksheet.write(start_row+5, 0, context.get('km_actuel', 0), number_format)
    worksheet.write(start_row+5, 1, context.get('missions_count', 0), number_format)
    worksheet.write(start_row+5, 2, context.get('distance_totale', 0), number_format)
    worksheet.write(start_row+5, 3, context.get('conso_moyenne', '-'), data_format)
    worksheet.write(start_row+5, 4, context.get('cout_total_carburant', 0), currency_format)
    worksheet.write(start_row+5, 5, context.get('incidents', 0), number_format)
    prochain_entretien = context.get('prochain_entretien')
    if prochain_entretien:
        worksheet.write(start_row+5, 6, prochain_entretien.date_entretien.strftime('%d/%m/%Y'), data_format)
    else:
        worksheet.write(start_row+5, 6, 'Aucun', data_format)
    alertes = context.get('alertes', [])
    worksheet.write(start_row+5, 7, ', '.join(alertes) if alertes else 'Aucune', data_format)
    row = start_row+8
    worksheet.merge_range(f'A{row+1}:E{row+1}', 'HISTORIQUE DES ENTRETIENS (10 derniers)', subheader_format)
    row += 2
    entretien_headers = ['Date', 'Type', 'Kilométrage', 'Coût', 'Observations']
    for col, h in enumerate(entretien_headers):
        worksheet.write(row, col, h, header_format)
    row += 1
    for e in context.get('entretiens_list', []):
        worksheet.write(row, 0, e.date_entretien.strftime('%d/%m/%Y') if e.date_entretien else '-', data_format)
        worksheet.write(row, 1, e.get_type_entretien_display(), data_format)
        worksheet.write(row, 2, e.kilometrage, number_format)
        worksheet.write(row, 3, e.cout, currency_format)
        worksheet.write(row, 4, e.observations or '-', data_format)
        row += 1
    row += 2
    worksheet.merge_range(f'A{row}:E{row}', 'HISTORIQUE DES RAVITAILLEMENTS (10 derniers)', subheader_format)
    row += 1
    ravitaillement_headers = ['Date', 'Kilométrage', 'Litres', 'Coût', 'Station']
    for col, h in enumerate(ravitaillement_headers):
        worksheet.write(row, col, h, header_format)
    row += 1
    for r in context.get('ravs_list', []):
        worksheet.write(row, 0, r.date_ravitaillement.strftime('%d/%m/%Y') if r.date_ravitaillement else '-', data_format)
        worksheet.write(row, 1, r.kilometrage_apres, number_format)
        worksheet.write(row, 2, r.litres, number_format)
        worksheet.write(row, 3, r.cout_total, currency_format)
        worksheet.write(row, 4, r.station or '-', data_format)
        row += 1
    for col in range(9):
        worksheet.set_column(col, col, 15)
    workbook.close()
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    vehicule = context.get("selected_vehicule")
    immatriculation = vehicule.immatriculation if vehicule else "vehicule"
    response['Content-Disposition'] = f'attachment; filename=rapport_vehicule_avance_{immatriculation}.xlsx'
    return response

def export_historique_detaille_pdf(request, context):
    export_all = request.GET.get('export_all')
    historique = context['historique_detaille']
    if export_all:
        historique = context.get('all_historique_detaille', historique)
    
    # Déterminer l'orientation selon le nombre de données
    orientation = 'landscape' if len(historique) > 20 else 'portrait'
    
    # from weasyprint import HTML  # Temporairement commenté pour le déploiement
    from django.template.loader import render_to_string
    from django.utils import timezone
    
    # Créer un contexte enrichi pour l'export
    export_context = {
        **context,
        'historique_detaille': historique,
        'orientation': orientation,
        'date_export': timezone.now().strftime('%d/%m/%Y à %H:%M'),
        'user_export': request.user.get_full_name() or request.user.username,
        'total_items': len(historique),
        'export_all': export_all,
    }
    
    html_string = render_to_string('rapport/partials/historique_detaille_export.html', export_context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    response = HttpResponse(content_type='application/pdf')
    vehicule = context.get("selected_vehicule")
    immatriculation = vehicule.immatriculation if vehicule else "vehicule"
    response['Content-Disposition'] = f'attachment; filename="historique_{immatriculation}.pdf"'
    html.write_pdf(response)
    return response

def export_historique_detaille_excel(request, context):
    export_all = request.GET.get('export_all')
    historique = context['historique_detaille']
    if export_all:
        historique = context.get('all_historique_detaille', historique)
    import xlsxwriter
    import io
    from django.utils import timezone
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 11,
        'font_color': 'white',
        'bg_color': '#4F81BD',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 13,
        'font_color': 'white',
        'bg_color': '#4F81BD',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    data_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'left',
        'bg_color': '#FFFFFF'
    })
    number_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'right',
        'num_format': '#,##0',
        'bg_color': '#FFFFFF'
    })
    currency_format = workbook.add_format({
        'font_size': 10,
        'border': 1,
        'align': 'right',
        'num_format': '#,##0.00 $',
        'bg_color': '#FFFFFF'
    })
    worksheet = workbook.add_worksheet('Historique détaillé')
    v = context.get("selected_vehicule")
    vehicule_info = f"{v.immatriculation} - {v.marque} {v.modele}" if v else 'Véhicule non sélectionné'
    worksheet.merge_range('A1:G1', f'HISTORIQUE DÉTAILLÉ - {vehicule_info}', title_format)
    worksheet.merge_range('A2:G2', f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M")} par {request.user.get_full_name() or request.user.username}', data_format)
    if export_all:
        worksheet.merge_range('A3:G3', f'Export complet : {len(historique)} événements', data_format)
    else:
        worksheet.merge_range('A3:G3', f'Export de la page courante : {len(historique)} événements', data_format)
    headers = ['Date', 'Type', 'Libellé', 'Kilométrage', 'Coût', 'Auteur', 'Observations']
    for col, h in enumerate(headers):
        worksheet.write(5, col, h, header_format)
    for i, h in enumerate(historique):
        row = 6 + i
        worksheet.write(row, 0, h['date'].strftime('%d/%m/%Y') if h['date'] else '-', data_format)
        worksheet.write(row, 1, h['type'].upper(), data_format)
        worksheet.write(row, 2, h['libelle'], data_format)
        worksheet.write(row, 3, h['km'] if h['km'] != '-' else '-', number_format)
        if h['cout'] and h['cout'] != '-':
            try:
                cout_value = float(h['cout'])
                worksheet.write(row, 4, cout_value, currency_format)
            except:
                worksheet.write(row, 4, h['cout'], data_format)
        else:
            worksheet.write(row, 4, '-', data_format)
        worksheet.write(row, 5, h['auteur'] if h['auteur'] else '-', data_format)
        worksheet.write(row, 6, h['obs'] if h['obs'] else '-', data_format)
    column_widths = [12, 15, 40, 12, 12, 20, 30]
    for col, width in enumerate(column_widths):
        worksheet.set_column(col, col, width)
    worksheet.autofilter(5, 0, 5 + len(historique) - 1, 6)
    workbook.close()
    output.seek(0)
    vehicule = context.get("selected_vehicule")
    immatriculation = vehicule.immatriculation if vehicule else "vehicule"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=historique_{immatriculation}.xlsx'
    return response

def export_historique_detaille_csv(request, context):
    import csv
    from django.http import HttpResponse
    export_all = request.GET.get('export_all')
    historique = context['historique_detaille']
    if export_all:
        historique = context.get('all_historique_detaille', historique)
    vehicule = context.get("selected_vehicule")
    immatriculation = vehicule.immatriculation if vehicule else "vehicule"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=historique_{immatriculation}.csv'
    writer = csv.writer(response)
    writer.writerow(['Date', 'Type', 'Libellé', 'Kilométrage', 'Coût', 'Auteur', 'Observations'])
    for h in historique:
        writer.writerow([
            h['date'].strftime('%d/%m/%Y') if h['date'] else '-',
            h['type'],
            h['libelle'],
            h['km'] if h['km'] else '-',
            h['cout'] if h['cout'] else '-',
            h['auteur'] if h['auteur'] else '-',
            h['obs'] if h['obs'] else '-',
        ])
    return response

def generate_chauffeur_charts_data(evaluations):
    """Génère les données pour les graphiques du rapport d'évaluation des chauffeurs."""
    import json
    
    # Données pour le graphique des scores
    scores_data = {
        'labels': [],
        'scores': [],
        'classifications': []
    }
    
    # Données pour le graphique de répartition des classifications
    classification_counts = {
        'Excellent': 0,
        'Bon': 0,
        'Moyen': 0,
        'À améliorer': 0,
        'Critique': 0
    }
    
    # Données pour le graphique des performances par critère
    criteres_data = {
        'productivite': [],
        'efficacite': [],
        'rentabilite': [],
        'regularite': []
    }
    
    for eval_ in evaluations:
        chauffeur_name = eval_['chauffeur']['get_full_name']
        score_total = eval_.get('score_total', 0)
        classification = eval_.get('classification', '')
        score_details = eval_.get('score_details', {})
        
        # Données pour le graphique des scores
        scores_data['labels'].append(chauffeur_name)
        scores_data['scores'].append(score_total)
        scores_data['classifications'].append(classification)
        
        # Comptage des classifications
        if classification in classification_counts:
            classification_counts[classification] += 1
        
        # Données pour les critères
        criteres_data['productivite'].append(score_details.get('productivite', 0))
        criteres_data['efficacite'].append(score_details.get('efficacite', 0))
        criteres_data['rentabilite'].append(score_details.get('rentabilite', 0))
        criteres_data['regularite'].append(score_details.get('regularite', 0))
    
    # Calcul des moyennes par critère
    moyennes_criteres = {}
    for critere, valeurs in criteres_data.items():
        if valeurs:
            moyennes_criteres[critere] = sum(valeurs) / len(valeurs)
        else:
            moyennes_criteres[critere] = 0
    
    return {
        'scores_data': scores_data,
        'classification_counts': classification_counts,
        'moyennes_criteres': moyennes_criteres
    }

def rapport_evaluation_chauffeurs_advanced(request):
    """Vue avancée du rapport d'évaluation des chauffeurs avec graphiques."""
    # Récupération des paramètres de filtre
    chauffeur_id = request.GET.get('chauffeur')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Récupération de la liste des chauffeurs pour le menu déroulant
    chauffeurs = Utilisateur.objects.filter(role='chauffeur')
    if not request.user.is_superuser:
        chauffeurs = chauffeurs.filter(etablissement=request.user.etablissement)
    
    # Requête de base pour les courses terminées uniquement, et distance raisonnable
    courses = Course.objects.filter(statut='terminee', distance_parcourue__lte=1000)
    
    # Filtrage par chauffeur
    if chauffeur_id:
        courses = courses.filter(chauffeur_id=chauffeur_id)
    
    # Filtrage par date
    if date_debut:
        courses = courses.filter(date_depart__date__gte=date_debut)
    if date_fin:
        courses = courses.filter(date_fin__date__lte=date_fin)
    
    # Calcul des statistiques des chauffeurs avec les véhicules associés
    stats_chauffeurs = courses.values(
        'chauffeur__id',
        'chauffeur__first_name', 
        'chauffeur__last_name',
        'chauffeur__date_joined',
        'vehicule__immatriculation',
        'vehicule__marque',
        'vehicule__modele'
    ).annotate(
        missions_terminees=Count('id'),
        distance_totale=Coalesce(Sum('distance_parcourue'), 0),
        duree_moyenne=Avg(ExpressionWrapper(
            F('date_fin') - F('date_depart'),
            output_field=DurationField()
        ))
    ).order_by('-missions_terminees')
    
    # Récupération des dépenses de carburant et d'entretien
    for chauffeur in stats_chauffeurs:
        vehicule_id = chauffeur.get('vehicule__id')
        if vehicule_id:
            # Dépenses de carburant et consommation
            ravitaillements = Ravitaillement.objects.filter(
                vehicule_id=vehicule_id,
                date_ravitaillement__date__range=[date_debut, date_fin] if date_debut and date_fin else None
            ).aggregate(
                total_carburant=Coalesce(Sum('cout_total'), 0),
                total_litres=Coalesce(Sum('litres'), 0)
            )
            chauffeur['depenses_carburant'] = ravitaillements['total_carburant']
            chauffeur['litres_consommes'] = ravitaillements['total_litres']
            
            # Dépenses d'entretien
            depenses_entretien = Entretien.objects.filter(
                vehicule_id=vehicule_id,
                date__date__range=[date_debut, date_fin] if date_debut and date_fin else None
            ).aggregate(
                total_entretien=Coalesce(Sum('cout'), 0)
            )
            chauffeur['depenses_entretien'] = depenses_entretien['total_entretien']
            
            # Coût total
            chauffeur['cout_total'] = chauffeur['depenses_carburant'] + chauffeur['depenses_entretien']
        else:
            chauffeur['depenses_carburant'] = 0
            chauffeur['depenses_entretien'] = 0
            chauffeur['cout_total'] = 0
    
    # Récupération des destinations fréquentes par chauffeur
    destinations_par_chauffeur = {}
    for chauffeur in stats_chauffeurs:
        destinations = (
            courses.filter(chauffeur_id=chauffeur['chauffeur__id'])
                  .values('destination')
                  .annotate(count=Count('id'))
                  .order_by('-count')[:3]  # Top 3 destinations
        )
        destinations_par_chauffeur[chauffeur['chauffeur__id']] = [
            f"{d['destination']} ({d['count']}x)" for d in destinations
        ]
    
    # Récupération des notes existantes depuis la session
    notes_chauffeurs = request.session.get('notes_chauffeurs', {})
    
    # Calcul des statistiques supplémentaires et du scoring pour chaque chauffeur
    evaluations = []
    scores_data = []  # Pour calculer les moyennes de scoring
    
    for chauffeur in stats_chauffeurs:
        # Calcul du nombre de jours prestés
        jours_prestes = courses.filter(chauffeur__id=chauffeur['chauffeur__id']).values('date_depart__date').distinct().count()
        # Calcul de l'ancienneté en jours
        anciennete = (timezone.now().date() - chauffeur['chauffeur__date_joined'].date()).days
        # Calcul de la distance moyenne par mission
        distance_moyenne = chauffeur['distance_totale'] / chauffeur['missions_terminees'] if chauffeur['missions_terminees'] > 0 else 0
        # Calcul du nombre de missions par jour
        missions_par_jour = chauffeur['missions_terminees'] / jours_prestes if jours_prestes > 0 else 0
        # Date de la dernière mission terminée
        last_mission = courses.filter(chauffeur__id=chauffeur['chauffeur__id']).order_by('-date_fin').first()
        date_mission = last_mission.date_fin if last_mission and last_mission.date_fin else None
        
        # Calcul de la consommation moyenne
        conso_moyenne = (chauffeur.get('litres_consommes', 0) * 100) / chauffeur['distance_totale'] if chauffeur['distance_totale'] > 0 else 0
        # Calcul du coût par km
        cout_km = chauffeur['distance_totale'] / chauffeur.get('cout_total', 1) if chauffeur.get('cout_total', 0) > 0 else 0
        
        # === SYSTÈME DE SCORING AVANCÉ ===
        score_total = 0
        score_details = {}
        
        # 1. Score de productivité (40% du total)
        if missions_par_jour > 0:
            score_productivite = min(100, (missions_par_jour / 2) * 100)  # 2 missions/jour = 100%
        else:
            score_productivite = 0
        score_details['productivite'] = score_productivite
        score_total += score_productivite * 0.4
        
        # 2. Score d'efficacité énergétique (25% du total)
        if conso_moyenne > 0:
            if conso_moyenne <= 6:  # Excellente consommation
                score_efficacite = 100
            elif conso_moyenne <= 8:  # Bonne consommation
                score_efficacite = 80
            elif conso_moyenne <= 10:  # Consommation moyenne
                score_efficacite = 60
            elif conso_moyenne <= 12:  # Consommation élevée
                score_efficacite = 40
            else:  # Consommation très élevée
                score_efficacite = 20
        else:
            score_efficacite = 0
        score_details['efficacite'] = score_efficacite
        score_total += score_efficacite * 0.25
        
        # 3. Score de rentabilité (20% du total)
        if cout_km > 0:
            if cout_km <= 50:  # Très rentable
                score_rentabilite = 100
            elif cout_km <= 100:  # Rentable
                score_rentabilite = 80
            elif cout_km <= 150:  # Moyennement rentable
                score_rentabilite = 60
            elif cout_km <= 200:  # Peu rentable
                score_rentabilite = 40
            else:  # Non rentable
                score_rentabilite = 20
        else:
            score_rentabilite = 0
        score_details['rentabilite'] = score_rentabilite
        score_total += score_rentabilite * 0.20
        
        # 4. Score de régularité (15% du total)
        if jours_prestes > 0:
            taux_regularite = (jours_prestes / 30) * 100  # Basé sur 30 jours
            score_regularite = min(100, taux_regularite)
        else:
            score_regularite = 0
        score_details['regularite'] = score_regularite
        score_total += score_regularite * 0.15
        
        # Classification du chauffeur
        if score_total >= 85:
            classification = "Excellent"
            badge_class = "badge-success"
        elif score_total >= 70:
            classification = "Bon"
            badge_class = "badge-primary"
        elif score_total >= 55:
            classification = "Moyen"
            badge_class = "badge-warning"
        elif score_total >= 40:
            classification = "À améliorer"
            badge_class = "badge-danger"
        else:
            classification = "Critique"
            badge_class = "badge-dark"
        
        # Recommandations basées sur les scores
        recommandations = []
        if score_details['productivite'] < 60:
            recommandations.append("Augmenter le nombre de missions par jour")
        if score_details['efficacite'] < 60:
            recommandations.append("Améliorer la conduite pour réduire la consommation")
        if score_details['rentabilite'] < 60:
            recommandations.append("Optimiser les coûts d'exploitation")
        if score_details['regularite'] < 60:
            recommandations.append("Améliorer l'assiduité")
        
        chauffeur_id = chauffeur['chauffeur__id']
        evaluations.append({
            'chauffeur': {
                'id': chauffeur_id,
                'first_name': chauffeur['chauffeur__first_name'],
                'last_name': chauffeur['chauffeur__last_name'],
                'get_full_name': f"{chauffeur['chauffeur__first_name']} {chauffeur['chauffeur__last_name']}"
            },
            'vehicule': {
                'immatriculation': chauffeur.get('vehicule__immatriculation', 'Non affecté'),
                'marque': chauffeur.get('vehicule__marque', 'Inconnue'),
                'modele': chauffeur.get('vehicule__modele', '')
            },
            'nb_courses': chauffeur['missions_terminees'],
            'nb_courses_terminees': chauffeur['missions_terminees'],
            'nb_jours_prestes': jours_prestes,
            'distance_totale': chauffeur['distance_totale'],
            'anciennete': anciennete,
            'distance_moyenne': distance_moyenne,
            'missions_par_jour': missions_par_jour,
            'depenses_carburant': chauffeur.get('depenses_carburant', 0),
            'depenses_entretien': chauffeur.get('depenses_entretien', 0),
            'cout_total': chauffeur.get('cout_total', 0),
            'cout_km': cout_km,
            'conso_moyenne': conso_moyenne,
            'top_destinations': destinations_par_chauffeur.get(chauffeur_id, []),
            'notes': notes_chauffeurs.get(str(chauffeur_id), ''),
            'date_mission': date_mission,
            # Nouvelles métriques de scoring
            'score_total': round(score_total, 1),
            'score_details': score_details,
            'classification': classification,
            'badge_class': badge_class,
            'recommandations': recommandations,
            'tendance': 'stable'  # À implémenter avec historique
        })
        
        scores_data.append(score_total)
    
    # Calcul des moyennes globales
    if date_debut and date_fin:
        try:
            date_debut_dt = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d').date()
            total_jours = (date_fin_dt - date_debut_dt).days + 1
        except (ValueError, TypeError):
            total_jours = 1
    else:
        # Si pas de date spécifiée, on prend la période depuis la première mission
        premiere_mission = courses.order_by('date_depart').first()
        if premiere_mission:
            total_jours = (timezone.now().date() - premiere_mission.date_depart.date()).days + 1
        else:
            total_jours = 1
    
    total_missions = sum(eval_['nb_courses'] for eval_ in evaluations)
    total_distance = sum(eval_['distance_totale'] for eval_ in evaluations)
    total_jours_prestes = sum(eval_['nb_jours_prestes'] for eval_ in evaluations if eval_['nb_jours_prestes'] > 0)
    
    # Calcul des totaux pour les moyennes globales
    total_litres_consommes = sum(eval_.get('litres_consommes', 0) for eval_ in evaluations)
    total_cout = sum(eval_.get('cout_total', 0) for eval_ in evaluations)
    
    # Calcul de la moyenne des scores
    score_moyen = sum(scores_data) / len(scores_data) if scores_data else 0
    
    moyennes = {
        'missions_par_jour': total_missions / total_jours if total_jours > 0 else 0,
        'distance_moyenne': total_distance / total_missions if total_missions > 0 else 0,
        'conso_moyenne': (total_litres_consommes * 100) / total_distance if total_distance > 0 else 0,
        'cout_km': total_cout / total_distance if total_distance > 0 else 0,
        'total_chauffeurs': len(evaluations),
        'total_distance': total_distance,
        'moyenne_jours_prestes': total_jours_prestes / len(evaluations) if evaluations else 0,
        'score_moyen': round(score_moyen, 1)
    }
    
    # Génération des données pour les graphiques
    charts_data = generate_chauffeur_charts_data(evaluations)
    
    context = {
        'evaluations': evaluations,
        'chauffeurs': chauffeurs,
        'selected_chauffeur': int(chauffeur_id) if chauffeur_id else None,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'moyennes': moyennes,
        'charts_data': charts_data,
        'request': request
    }
    
    return render(request, 'rapport/rapport_evaluation_chauffeurs_advanced.html', context)

def generate_vehicule_charts_data(stats_vehicules):
    """Génère les données pour les graphiques du rapport d'évaluation des véhicules."""
    import json
    
    # Données pour le graphique des scores
    scores_data = {
        'labels': [],
        'scores': [],
        'classifications': []
    }
    
    # Données pour le graphique de répartition des classifications
    classification_counts = {
        'Excellent': 0,
        'Bon': 0,
        'Moyen': 0,
        'À améliorer': 0,
        'Critique': 0
    }
    
    # Données pour le graphique des performances par critère
    criteres_data = {
        'fiabilite': [],
        'efficacite': [],
        'rentabilite': [],
        'utilisation': []
    }
    
    for stat in stats_vehicules:
        vehicule_name = stat['vehicule'].immatriculation
        score_total = stat.get('score_total', 0)
        classification = stat.get('classification', '')
        score_details = stat.get('score_details', {})
        
        # Données pour le graphique des scores
        scores_data['labels'].append(vehicule_name)
        scores_data['scores'].append(score_total)
        scores_data['classifications'].append(classification)
        
        # Comptage des classifications
        if classification in classification_counts:
            classification_counts[classification] += 1
        
        # Données pour les critères
        criteres_data['fiabilite'].append(score_details.get('fiabilite', 0))
        criteres_data['efficacite'].append(score_details.get('efficacite', 0))
        criteres_data['rentabilite'].append(score_details.get('rentabilite', 0))
        criteres_data['utilisation'].append(score_details.get('utilisation', 0))
    
    # Calcul des moyennes par critère
    moyennes_criteres = {}
    for critere, valeurs in criteres_data.items():
        if valeurs:
            moyennes_criteres[critere] = sum(valeurs) / len(valeurs)
        else:
            moyennes_criteres[critere] = 0
    
    return {
        'scores_data': scores_data,
        'classification_counts': classification_counts,
        'moyennes_criteres': moyennes_criteres
    }

def generate_mission_charts_data(courses):
    """Génère les données pour les graphiques du rapport des missions."""
    import json
    
    # Données pour le graphique des scores
    scores_data = {
        'labels': [],
        'scores': [],
        'classifications': []
    }
    
    # Données pour le graphique de répartition des classifications
    classification_counts = {
        'Excellent': 0,
        'Bon': 0,
        'Moyen': 0,
        'À améliorer': 0,
        'Critique': 0
    }
    
    # Données pour le graphique des performances par critère
    criteres_data = {
        'ponctualite': [],
        'efficacite': [],
        'rentabilite': [],
        'qualite': []
    }
    
    # Données pour le graphique des statuts
    statut_counts = {}
    
    for course in courses:
        # Données pour le graphique des scores
        mission_id = f"Mission {course.id}"
        score_total = getattr(course, 'score_total', 0)
        classification = getattr(course, 'classification', '')
        
        scores_data['labels'].append(mission_id)
        scores_data['scores'].append(score_total)
        scores_data['classifications'].append(classification)
        
        # Comptage des classifications
        if classification in classification_counts:
            classification_counts[classification] += 1
        
        # Données pour les critères
        score_details = getattr(course, 'score_details', {})
        criteres_data['ponctualite'].append(score_details.get('ponctualite', 0))
        criteres_data['efficacite'].append(score_details.get('efficacite', 0))
        criteres_data['rentabilite'].append(score_details.get('rentabilite', 0))
        criteres_data['qualite'].append(score_details.get('qualite', 0))
        
        # Comptage des statuts
        statut = course.get_statut_display()
        if statut not in statut_counts:
            statut_counts[statut] = 0
        statut_counts[statut] += 1
    
    # Calcul des moyennes par critère
    moyennes_criteres = {}
    for critere, valeurs in criteres_data.items():
        if valeurs:
            moyennes_criteres[critere] = sum(valeurs) / len(valeurs)
        else:
            moyennes_criteres[critere] = 0
    
    return {
        'scores_data': scores_data,
        'classification_counts': classification_counts,
        'statut_counts': statut_counts,
        'moyennes_criteres': moyennes_criteres
    }

def rapport_missions_advanced(request):
    """Vue avancée du rapport des missions avec graphiques."""
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    chauffeur_id = request.GET.get('chauffeur')
    vehicule_id = request.GET.get('vehicule')
    demandeur_id = request.GET.get('demandeur')
    export_format = request.GET.get('export')

    courses = Course.objects.all()
    
    if not request.user.is_superuser:
        courses = courses.filter(vehicule__etablissement=request.user.etablissement)
    
    if date_debut:
        courses = courses.filter(date_depart__date__gte=date_debut)
    if date_fin:
        courses = courses.filter(date_depart__date__lte=date_fin)
    statut_filter = request.GET.get('statut')
    if statut_filter:
        courses = courses.filter(statut=statut_filter)
    if chauffeur_id:
        courses = courses.filter(chauffeur_id=chauffeur_id)
    if vehicule_id:
        courses = courses.filter(vehicule_id=vehicule_id)
    if demandeur_id:
        courses = courses.filter(demandeur_id=demandeur_id)

    # Calcul de la distance en km et de la durée en heures
    courses = courses.annotate(
        distance_km=ExpressionWrapper(
            F('distance_parcourue'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    )
    
    courses = courses.annotate(
        duree_heures=Value(None, output_field=DecimalField(max_digits=10, decimal_places=2))
    )
    
    # === SYSTÈME DE SCORING AVANCÉ POUR MISSIONS ===
    missions_scored = []
    scores_data = []  # Pour calculer les moyennes de scoring
    
    for course in courses:
        # Calculer la durée en Python pour chaque cours
        if course.date_depart and course.date_fin:
            duree = (course.date_fin - course.date_depart).total_seconds() / 3600.0
            course.duree_heures = round(duree, 2)
        else:
            course.duree_heures = 0
        
        # === CALCUL DU SCORE DE LA MISSION ===
        score_total = 0
        score_details = {}
        
        # 1. Score de ponctualité (30% du total)
        # Basé sur le respect des horaires et la durée de la mission
        score_ponctualite = 100
        if course.date_depart and course.date_fin:
            duree_mission = course.duree_heures
            distance_mission = float(course.distance_parcourue) if course.distance_parcourue else 0
            
            # Calcul de la vitesse moyenne (km/h)
            if duree_mission > 0:
                vitesse_moyenne = distance_mission / duree_mission
                # Score basé sur la vitesse moyenne (objectif : 60-80 km/h)
                if 60 <= vitesse_moyenne <= 80:
                    score_ponctualite = 100
                elif 50 <= vitesse_moyenne < 60 or 80 < vitesse_moyenne <= 90:
                    score_ponctualite = 80
                elif 40 <= vitesse_moyenne < 50 or 90 < vitesse_moyenne <= 100:
                    score_ponctualite = 60
                elif vitesse_moyenne < 40 or vitesse_moyenne > 100:
                    score_ponctualite = 40
            else:
                score_ponctualite = 50  # Score neutre si pas de durée
        else:
            score_ponctualite = 30  # Pénalité si pas de dates
        
        # Pénalité pour missions annulées ou en retard
        if course.statut == 'annulee':
            score_ponctualite -= 50
        elif course.statut == 'en_retard':
            score_ponctualite -= 30
        
        score_ponctualite = max(0, score_ponctualite)
        score_details['ponctualite'] = score_ponctualite
        score_total += score_ponctualite * 0.30
        
        # 2. Score d'efficacité (25% du total)
        # Basé sur la distance parcourue et l'optimisation du trajet
        score_efficacite = 100
        if course.distance_parcourue:
            distance = float(course.distance_parcourue)
            # Score basé sur la distance (missions plus longues = meilleur score)
            if distance >= 100:  # Mission longue
                score_efficacite = 100
            elif distance >= 50:  # Mission moyenne
                score_efficacite = 80
            elif distance >= 20:  # Mission courte
                score_efficacite = 60
            else:  # Très courte mission
                score_efficacite = 40
        else:
            score_efficacite = 0
        
        # Bonus pour missions terminées
        if course.statut == 'terminee':
            score_efficacite += 20
            score_efficacite = min(100, score_efficacite)
        
        score_details['efficacite'] = score_efficacite
        score_total += score_efficacite * 0.25
        
        # 3. Score de rentabilité (25% du total)
        # Basé sur le coût par km et la valeur de la mission
        score_rentabilite = 100
        if course.distance_parcourue:
            distance = float(course.distance_parcourue)
            cout_km = 0.5  # Coût fixe de 0.5$/km
            cout_total = distance * cout_km
            
            # Score basé sur la rentabilité (missions plus rentables = meilleur score)
            if cout_total <= 25:  # Mission très rentable
                score_rentabilite = 100
            elif cout_total <= 50:  # Mission rentable
                score_rentabilite = 80
            elif cout_total <= 100:  # Mission moyennement rentable
                score_rentabilite = 60
            else:  # Mission coûteuse
                score_rentabilite = 40
        else:
            score_rentabilite = 50
        
        score_details['rentabilite'] = score_rentabilite
        score_total += score_rentabilite * 0.25
        
        # 4. Score de qualité (20% du total)
        # Basé sur la satisfaction client et la qualité du service
        score_qualite = 100
        
        # Pénalités pour différents statuts
        if course.statut == 'annulee':
            score_qualite -= 80
        elif course.statut == 'en_retard':
            score_qualite -= 40
        elif course.statut == 'en_cours':
            score_qualite -= 20
        elif course.statut == 'terminee':
            score_qualite += 10  # Bonus pour mission terminée
            score_qualite = min(100, score_qualite)
        
        # Bonus pour missions avec passagers
        if hasattr(course, 'nombre_passagers') and course.nombre_passagers and course.nombre_passagers > 0:
            score_qualite += 10
            score_qualite = min(100, score_qualite)
        
        score_qualite = max(0, score_qualite)
        score_details['qualite'] = score_qualite
        score_total += score_qualite * 0.20
        
        # Classification de la mission
        if score_total >= 85:
            classification = "Excellent"
            badge_class = "badge-success"
        elif score_total >= 70:
            classification = "Bon"
            badge_class = "badge-primary"
        elif score_total >= 55:
            classification = "Moyen"
            badge_class = "badge-warning"
        elif score_total >= 40:
            classification = "À améliorer"
            badge_class = "badge-danger"
        else:
            classification = "Critique"
            badge_class = "badge-dark"
        
        # Recommandations basées sur les scores
        recommandations = []
        if score_details['ponctualite'] < 60:
            recommandations.append("Améliorer la ponctualité et optimiser les trajets")
        if score_details['efficacite'] < 60:
            recommandations.append("Optimiser les distances et les itinéraires")
        if score_details['rentabilite'] < 60:
            recommandations.append("Réduire les coûts d'exploitation")
        if score_details['qualite'] < 60:
            recommandations.append("Améliorer la qualité du service")
        if course.statut == 'annulee':
            recommandations.append("Analyser les causes d'annulation")
        
        # Ajouter les métriques calculées à la mission
        course.score_total = round(score_total, 1)
        course.score_details = score_details
        course.classification = classification
        course.badge_class = badge_class
        course.recommandations = recommandations
        course.cout_km = 0.5  # Coût fixe
        course.cout_total = float(course.distance_parcourue * 0.5) if course.distance_parcourue else 0
        
        missions_scored.append(course)
        scores_data.append(score_total)
    
    total_missions = len(missions_scored)
    total_distance = sum(float(c.distance_km) for c in missions_scored if c.distance_km is not None)
    total_duree = sum(float(c.duree_heures) for c in missions_scored if c.duree_heures is not None)
    # Calcul du coût basé sur la distance (0.5$ par km)
    COUT_PAR_KM = Decimal('0.5')  # 0.5$ par km
    total_cout = float(sum(Decimal(str(c.distance_km)) * COUT_PAR_KM for c in missions_scored if c.distance_km is not None))
    
    # Calcul de la moyenne des scores
    score_moyen = sum(scores_data) / len(scores_data) if scores_data else 0
    
    # Génération des données pour les graphiques
    charts_data = generate_mission_charts_data(missions_scored)
    
    # Calcul des statistiques par statut
    stats_par_statut = {}
    for statut_choice in Course.STATUS_CHOICES:
        statut_code = statut_choice[0]
        missions_statut = [c for c in missions_scored if c.statut == statut_code]
        count = len(missions_statut)
        distance = sum(float(m.distance_parcourue) for m in missions_statut if m.distance_parcourue is not None)
        duree = sum(float(m.duree_heures) for m in missions_statut if m.duree_heures is not None)
        cout = float(sum(Decimal(str(m.distance_parcourue)) * Decimal('0.5') for m in missions_statut if m.distance_parcourue is not None))
        
        stats_par_statut[statut_code] = {
            'count': count,
            'distance': distance,
            'duree': duree,
            'cout': cout,
            'label': dict(Course.STATUS_CHOICES)[statut_code]
        }

    if total_missions > 0:
        moyennes = {
            'distance_moyenne': total_distance / total_missions,
            'duree_moyenne': total_duree / total_missions,
            'cout_moyen': total_cout / total_missions,
            'cout_km': total_cout / total_distance if total_distance > 0 else 0,
            'score_moyen': round(score_moyen, 1)
        }
    else:
        moyennes = {
            'distance_moyenne': 0,
            'duree_moyenne': timedelta(),
            'cout_moyen': 0,
            'cout_km': 0,
            'score_moyen': 0
        }

    # Préparation du contexte
    context = {
        'courses': missions_scored,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'selected_statut': statut_filter,
        'selected_chauffeur': chauffeur_id,
        'selected_vehicule': vehicule_id,
        'selected_demandeur': demandeur_id,
        'total_missions': total_missions,
        'total_distance': total_distance,
        'total_duree': total_duree,
        'total_cout': total_cout,
        'score_moyen': round(score_moyen, 1),
        'stats_par_statut': stats_par_statut,
        'moyennes': moyennes,
        'charts_data': charts_data,
        'statut_choices': Course.STATUS_CHOICES,
    }
    
    return render(request, 'rapport/rapport_missions_advanced.html', context)
