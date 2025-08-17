from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.template.loader import get_template
from core.views import send_message as send_chat_message
from core.models import Course, ActionTraceur, Vehicule
from ravitaillement.models import Ravitaillement
from entretien.models import Entretien
from .forms import DemarrerMissionForm, TerminerMissionForm
from notifications.utils import notify_user, send_sms, send_whatsapp
from core.utils import get_latest_vehicle_kilometrage
import datetime
import openpyxl
from openpyxl.styles import Font
import tempfile
import os
import logging
# from weasyprint import HTML  # Temporairement commenté pour le déploiement
from .models import CommentaireRapportChauffeur, CommentaireRapportDemandeur
from django.contrib.admin.views.decorators import staff_member_required

@login_required
def dashboard(request):
    """Vue pour le tableau de bord du chauffeur"""
    # Vérifier que l'utilisateur est bien un chauffeur, un admin ou un superuser
    if request.user.role != 'chauffeur' and request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas les droits pour accéder à cette page.")
        return redirect('home')
    
    # Récupérer les missions assignées au chauffeur connecté avec leurs relations
    missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(chauffeur=request.user).filter(
        Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')
    ).order_by('-date_validation')
    
    # Si l'utilisateur est admin ou superuser, montrer toutes les missions
    if request.user.role == 'admin' or request.user.is_superuser:
        missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(
            Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')
        ).order_by('-date_validation')
    
    # Filtres
    statut = request.GET.get('statut')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if statut:
        missions = missions.filter(statut=statut)
    
    if date_debut:
        date_debut = datetime.datetime.strptime(date_debut, '%Y-%m-%d').date()
        missions = missions.filter(date_validation__date__gte=date_debut)
    
    if date_fin:
        date_fin = datetime.datetime.strptime(date_fin, '%Y-%m-%d').date()
        missions = missions.filter(date_validation__date__lte=date_fin)
    
    # Pagination
    paginator = Paginator(missions, 12)  # 12 missions par page
    page_number = request.GET.get('page')
    missions_page = paginator.get_page(page_number)
    
    # Statistiques
    if request.user.role == 'admin' or request.user.is_superuser:
        stats = {
            'total': Course.objects.filter(Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')).count(),
            'a_effectuer': Course.objects.filter(statut='validee').count(),
            'en_cours': Course.objects.filter(statut='en_cours').count(),
            'terminees': Course.objects.filter(statut='terminee').count(),
        }
    else:
        stats = {
            'total': Course.objects.filter(chauffeur=request.user).count(),
            'a_effectuer': Course.objects.filter(chauffeur=request.user, statut='validee').count(),
            'en_cours': Course.objects.filter(chauffeur=request.user, statut='en_cours').count(),
            'terminees': Course.objects.filter(chauffeur=request.user, statut='terminee').count(),
        }
    
    context = {
        'missions': missions_page,
        'stats': stats,
    }
    
    return render(request, 'chauffeur/dashboard.html', context)

@login_required
def detail_mission(request, mission_id):
    """Vue pour afficher les détails d'une mission"""
    # Vérifier que l'utilisateur est bien un chauffeur, un admin ou un superuser
    if request.user.role != 'chauffeur' and request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas les droits pour accéder à cette page.")
        return redirect('home')
    
    # Si l'utilisateur est admin ou superuser, permettre l'accès à toutes les missions
    if request.user.role == 'admin' or request.user.is_superuser:
        mission = get_object_or_404(Course.objects.select_related('demandeur', 'vehicule', 'dispatcher'), id=mission_id)
    else:
        mission = get_object_or_404(Course.objects.select_related('demandeur', 'vehicule', 'dispatcher'), id=mission_id, chauffeur=request.user)
    
    # Vérifier que la mission est bien dans un état valide
    if mission.statut not in ['validee', 'en_cours', 'terminee']:
        messages.error(request, "Vous n'avez pas accès à cette mission.")
        return redirect('chauffeur:dashboard')
    
    # Récupérer l'historique des actions liées à cette mission
    historique = ActionTraceur.objects.filter(
        Q(details__icontains=f"Demande #{mission.id}") |
        Q(details__icontains=f"Course {mission.id}") |
        Q(details__icontains=f"Mission {mission.id}")
    ).order_by('-date_action')
    
    context = {
        'mission': mission,
        'historique': historique,
    }
    
    return render(request, 'chauffeur/detail_mission.html', context)

@login_required
def demarrer_mission(request, mission_id):
    """Vue pour démarrer une mission"""
    # Vérifier que l'utilisateur est bien un chauffeur, un admin ou un superuser
    if request.user.role != 'chauffeur' and request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas les droits pour accéder à cette page.")
        return redirect('home')
    
    # Si l'utilisateur est admin ou superuser, permettre l'accès à toutes les missions
    if request.user.role == 'admin' or request.user.is_superuser:
        mission = get_object_or_404(Course, id=mission_id, statut='validee')
    else:
        mission = get_object_or_404(Course, id=mission_id, chauffeur=request.user, statut='validee')
    
    vehicule = mission.vehicule
    
    if request.method == 'POST':
        form = DemarrerMissionForm(request.POST, vehicule=vehicule)
        if form.is_valid():
            kilometrage_depart = form.cleaned_data['kilometrage_depart']
            
            # Utiliser la fonction utilitaire pour obtenir le dernier kilométrage enregistré
            dernier_kilometrage = get_latest_vehicle_kilometrage(vehicule)
            
            if kilometrage_depart < dernier_kilometrage:
                messages.error(request, f"Le kilométrage de départ ({kilometrage_depart} km) ne peut pas être inférieur au dernier kilométrage enregistré ({dernier_kilometrage} km).")
                return render(request, 'chauffeur/demarrer_mission.html', {'mission': mission, 'form': form})
            
            # Mettre à jour la mission
            mission.statut = 'en_cours'
            mission.kilometrage_depart = kilometrage_depart
            mission.date_depart = timezone.now()
            mission.save()
            
            # Créer une entrée dans l'historique des actions
            commentaire = form.cleaned_data['commentaire']
            action_details = f"Mission {mission.id} démarrée - Kilométrage de départ: {mission.kilometrage_depart} km"
            if commentaire:
                action_details += f" - Commentaire: {commentaire}"
            
            ActionTraceur.objects.create(
                utilisateur=request.user,
                action="Démarrage de mission",
                details=action_details
            )
            
            # Message pour le demandeur
            demandeur_title = f"Votre course #{mission.id} a démarré"
            demandeur_message = f"Votre course de {mission.point_embarquement} à {mission.destination} a démarré. Chauffeur: {mission.chauffeur.get_full_name()}"
            
            # Notification interne au demandeur
            notify_user(
                mission.demandeur,
                demandeur_title,
                demandeur_message,
                notification_type='all',
                course=mission
            )
            
            # Notification par SMS au demandeur
            if mission.demandeur.telephone:
                sms_message = f"{demandeur_title}\n{demandeur_message}"
                send_sms(mission.demandeur.telephone, sms_message)
            
            # Notification par WhatsApp au demandeur
            if mission.demandeur.telephone:
                whatsapp_message = f"*{demandeur_title}*\n\n{demandeur_message}"
                send_whatsapp(mission.demandeur.telephone, whatsapp_message)
            
            messages.success(request, f'La mission #{mission.id} a été démarrée avec succès.')
            return redirect('chauffeur:detail_mission', mission.id)
    else:
        # Initialiser le formulaire avec le dernier kilométrage connu du véhicule
        initial_data = {}
        if vehicule:
            initial_data['kilometrage_depart'] = get_latest_vehicle_kilometrage(vehicule)
        form = DemarrerMissionForm(vehicule=vehicule, initial=initial_data)
    
    context = {
        'mission': mission,
        'form': form,
    }
    
    return render(request, 'chauffeur/demarrer_mission.html', context)

@login_required
def terminer_mission(request, mission_id):
    """Vue pour terminer une mission"""
    # Vérifier que l'utilisateur est bien un chauffeur, un admin ou un superuser
    if request.user.role != 'chauffeur' and request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, "Vous n'avez pas les droits pour accéder à cette page.")
        return redirect('home')
    
    try:
        # Si l'utilisateur est admin ou superuser, permettre l'accès à toutes les missions
        if request.user.role == 'admin' or request.user.is_superuser:
            # Utiliser select_related pour optimiser la requête
            mission = Course.objects.select_related('demandeur', 'vehicule').get(id=mission_id, statut='en_cours')
        else:
            # Utiliser select_related pour optimiser la requête
            mission = Course.objects.select_related('demandeur', 'vehicule').get(id=mission_id, chauffeur=request.user, statut='en_cours')
    except Course.DoesNotExist:
        # Vérifier si la mission existe mais n'est pas en cours
        try:
            if request.user.role == 'admin' or request.user.is_superuser:
                mission_status = Course.objects.values_list('statut', flat=True).get(id=mission_id)
            else:
                mission_status = Course.objects.values_list('statut', flat=True).get(id=mission_id, chauffeur=request.user)
            messages.error(request, f"La mission #{mission_id} ne peut pas être terminée car son statut actuel est '{mission_status}'")
        except Course.DoesNotExist:
            messages.error(request, f"La mission #{mission_id} n'existe pas ou n'est pas assignée à ce chauffeur.")
        return redirect('chauffeur:dashboard')
    
    if request.method == 'POST':
        form = TerminerMissionForm(request.POST, kilometrage_depart=mission.kilometrage_depart)
        if form.is_valid():
            kilometrage_fin = form.cleaned_data['kilometrage_fin']
            
            # Vérifier que le kilométrage n'est pas inférieur au dernier kilométrage enregistré
            dernier_kilometrage = 0
            
            # Vérifier le dernier ravitaillement
            dernier_ravitaillement = Ravitaillement.objects.filter(
                vehicule=mission.vehicule
            ).order_by('-date_ravitaillement').first()
            if dernier_ravitaillement and dernier_ravitaillement.kilometrage_apres > dernier_kilometrage:
                dernier_kilometrage = dernier_ravitaillement.kilometrage_apres
            
            # Vérifier la dernière course terminée
            derniere_course = Course.objects.filter(
                vehicule=mission.vehicule,
                statut='terminee'
            ).exclude(id=mission.id).order_by('-date_fin').first()
            if derniere_course and derniere_course.kilometrage_fin > dernier_kilometrage:
                dernier_kilometrage = derniere_course.kilometrage_fin
                
            # Vérifier le dernier entretien
            dernier_entretien = Entretien.objects.filter(
                vehicule=mission.vehicule,
                kilometrage_apres__isnull=False
            ).order_by('-date_entretien').first()
            if dernier_entretien and dernier_entretien.kilometrage_apres > dernier_kilometrage:
                dernier_kilometrage = dernier_entretien.kilometrage_apres
            
            if kilometrage_fin < dernier_kilometrage:
                messages.error(request, f"Le kilométrage d'arrivée ({kilometrage_fin} km) ne peut pas être inférieur au dernier kilométrage enregistré ({dernier_kilometrage} km).")
                return render(request, 'chauffeur/terminer_mission.html', {'mission': mission, 'form': form})
            
            if kilometrage_fin < mission.kilometrage_depart:
                messages.error(request, f"Le kilométrage d'arrivée ({kilometrage_fin} km) ne peut pas être inférieur au kilométrage de départ ({mission.kilometrage_depart} km).")
                return render(request, 'chauffeur/terminer_mission.html', {'mission': mission, 'form': form})
            
            # Mettre à jour la mission
            mission.statut = 'terminee'
            mission.kilometrage_fin = kilometrage_fin
            mission.date_fin = timezone.now()
            mission.distance_parcourue = mission.kilometrage_fin - mission.kilometrage_depart
            mission.save()

            # Mettre à jour le kilométrage du véhicule si le kilométrage de fin de mission est plus élevé
            if mission.kilometrage_fin is not None and mission.vehicule and mission.vehicule.kilometrage_actuel < mission.kilometrage_fin:
                mission.vehicule.kilometrage_actuel = mission.kilometrage_fin
                mission.vehicule.save(update_fields=['kilometrage_actuel'])
                messages.info(request, f"Le kilométrage du véhicule {mission.vehicule.immatriculation} a été mis à jour à {mission.vehicule.kilometrage_actuel} km suite à la fin de la mission.")
            
            # Vérifier si un entretien est nécessaire
            vehicule = mission.vehicule
            if vehicule.entretien_necessaire(mission.kilometrage_fin):
                # Notification pour l'entretien nécessaire
                entretien_title = f"Entretien nécessaire pour {vehicule.immatriculation}"
                entretien_message = f"Le véhicule {vehicule.immatriculation} a atteint {mission.kilometrage_fin} km. Un entretien est nécessaire (dernier entretien à {vehicule.kilometrage_dernier_entretien} km)."
                
                # Notification interne aux administrateurs
                User = get_user_model()
                admins = User.objects.filter(role='admin')
                for admin in admins:
                    notify_user(
                        admin,
                        entretien_title,
                        entretien_message,
                        notification_type='all'
                    )
                
                # Notification par SMS aux administrateurs
                for admin in admins:
                    if admin.telephone:
                        send_sms(admin.telephone, f"{entretien_title}\n{entretien_message}")
                
                # Notification par WhatsApp aux administrateurs
                for admin in admins:
                    if admin.telephone:
                        send_whatsapp(admin.telephone, f"*{entretien_title}*\n\n{entretien_message}")

            # Créer une entrée dans l'historique des actions
            commentaire = form.cleaned_data['commentaire']
            action_details = f"Mission {mission.id} terminée - Kilométrage d'arrivée: {mission.kilometrage_fin} km - Distance parcourue: {mission.distance_parcourue} km"
            if commentaire:
                action_details += f" - Commentaire: {commentaire}"
            
            ActionTraceur.objects.create(
                utilisateur=request.user,
                action="Fin de mission",
                details=action_details
            )
            
            # Message pour le demandeur
            demandeur_title = f"Votre course #{mission.id} est terminée"
            demandeur_message = f"Votre course de {mission.point_embarquement} à {mission.destination} est terminée. Distance parcourue: {mission.distance_parcourue} km."
            
            # Notification interne au demandeur via le système de notification
            notify_user(
                mission.demandeur,
                demandeur_title,
                demandeur_message,
                notification_type='all',
                course=mission
            )
            
            # Envoi d'un message dans le chat interne
            try:
                # Créer une requête factice pour send_message
                from django.http import HttpRequest
                
                # Créer une requête factice
                fake_request = HttpRequest()
                fake_request.user = request.user  # L'expéditeur est l'utilisateur actuel (le chauffeur)
                fake_request.method = 'POST'
                fake_request.POST = {
                    'recipient_id': str(mission.demandeur.id),
                    'content': f"🚗 Mission #{mission.id} terminée !\n\n" \
                              f"📍 De: {mission.point_embarquement}\n" \
                              f"🏁 À: {mission.destination}\n" \
                              f"📏 Distance: {mission.distance_parcourue} km\n" \
                              f"⏱️ Terminée le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}\n\n" \
                              f"Merci d'avoir fait confiance à notre service !"
                }
                
                # Appeler la vue send_message
                response = send_chat_message(fake_request)
                
                # Vérifier si l'envoi a réussi
                if response.status_code != 200:
                    print(f"Erreur lors de l'envoi du message dans le chat: {response.content}")
                    
            except Exception as e:
                print(f"Erreur lors de l'envoi de la notification de fin de mission dans le chat: {str(e)}")
            
            # Notification par SMS au demandeur (désactivée pour le moment)
            if mission.demandeur.telephone:
                sms_message = f"{demandeur_title}\n{demandeur_message}"
                send_sms(mission.demandeur.telephone, sms_message)
            
            # Notification par WhatsApp au demandeur (désactivée pour le moment)
            if mission.demandeur.telephone:
                whatsapp_message = f"*{demandeur_title}*\n\n{demandeur_message}\n\nMerci d'avoir utilisé notre service."
                send_whatsapp(mission.demandeur.telephone, whatsapp_message)
            
            # Mettre à jour la distance journalière
            from .models import DistanceJournaliere
            date_jour = timezone.now().date()
            distance_jour, created = DistanceJournaliere.objects.get_or_create(
                chauffeur=request.user,
                date=date_jour,
                defaults={
                    'distance_totale': mission.distance_parcourue,
                    'nombre_courses': 1
                }
            )
            if not created:
                distance_jour.distance_totale += mission.distance_parcourue
                distance_jour.nombre_courses += 1
                distance_jour.save()

            # Synchronisation du kilométrage centralisé (mise à jour du kilometrage_actuel du véhicule)
            if mission.vehicule and mission.kilometrage_fin is not None:
                vehicule = mission.vehicule
                if vehicule.kilometrage_actuel is None or vehicule.kilometrage_actuel < mission.kilometrage_fin:
                    vehicule.kilometrage_actuel = mission.kilometrage_fin
                    vehicule.save(update_fields=["kilometrage_actuel"])

            messages.success(request, f'La mission #{mission.id} a été terminée avec succès.')
            return redirect('chauffeur:detail_mission', mission.id)
    else:
        form = TerminerMissionForm(kilometrage_depart=mission.kilometrage_depart)
    
    context = {
        'mission': mission,
        'form': form,
    }
    
    return render(request, 'chauffeur/terminer_mission.html', context)

@login_required
def missions_pdf(request):
    """
    Exporte la liste des missions au format PDF avec les filtres appliqués.
    """
    try:
        # Récupérer les missions selon le rôle
        if request.user.role == 'admin' or request.user.is_superuser:
            missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(
                Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')
            ).order_by('-date_validation')
        else:
            missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(
                chauffeur=request.user
            ).filter(Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')).order_by('-date_validation')
        
        # Appliquer les filtres
        statut = request.GET.get('statut')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        
        if statut:
            missions = missions.filter(statut=statut)
        
        if date_debut:
            try:
                date_debut = datetime.datetime.strptime(date_debut, '%Y-%m-%d').date()
                missions = missions.filter(date_validation__date__gte=date_debut)
            except (ValueError, TypeError):
                pass  # Ignorer si la date est mal formatée
        
        if date_fin:
            try:
                date_fin = datetime.datetime.strptime(date_fin, '%Y-%m-%d').date()
                missions = missions.filter(date_validation__date__lte=date_fin)
            except (ValueError, TypeError):
                pass  # Ignorer si la date est mal formatée
        
        # Récupérer le département de l'utilisateur
        departement = None
        if hasattr(request.user, 'etablissement'):
            departement = request.user.etablissement.nom if request.user.etablissement else None
        
        # Préparer les données pour le template
        context = {
            'request': request,  # Ajout de la requête au contexte
            'missions': missions,
            'date_export': datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            'departement': departement,  # Ajout du département
            'filters': {
                'statut': statut,
                'date_debut': date_debut.strftime("%d/%m/%Y") if date_debut and hasattr(date_debut, 'strftime') else date_debut,
                'date_fin': date_fin.strftime("%d/%m/%Y") if date_fin and hasattr(date_fin, 'strftime') else date_fin,
            },
            'now': datetime.datetime.now(),  # Ajout de la date actuelle pour le template
            'BASE_DIR': os.path.abspath(os.curdir),
        }
        
        # Rendre le template HTML
        template = get_template('chauffeur/missions_pdf.html')
        html = template.render(context)
        
        # Vérifier si l'export PDF est activé
        from .deployment_config import is_feature_enabled, get_pdf_placeholder_response
        
        if not is_feature_enabled('ENABLE_PDF_EXPORT'):
            return get_pdf_placeholder_response()
        
        # Créer un fichier PDF temporaire
        fd, path = tempfile.mkstemp(suffix=".pdf")
        try:
            # Temporairement désactivé - weasyprint non disponible
            # HTML(string=html).write_pdf(path)
            # with open(path, "rb") as f:
            #     pdf = f.read()
            return get_pdf_placeholder_response()
        finally:
            os.close(fd)
            try:
                os.remove(path)
            except:
                pass
        
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error("Erreur lors de la génération du PDF: %s", str(e), exc_info=True)
        messages.error(request, f"Une erreur est survenue lors de la génération du PDF: {str(e)}")
        return redirect('chauffeur:dashboard')

@login_required
def missions_excel(request):
    """
    Exporte la liste des missions au format Excel avec les filtres appliqués.
    """
    # Récupérer les missions selon le rôle
    if request.user.role == 'admin' or request.user.is_superuser:
        missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(
            Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')
        ).order_by('-date_validation')
    else:
        missions = Course.objects.select_related('demandeur', 'vehicule', 'dispatcher').filter(
            chauffeur=request.user
        ).filter(Q(statut='validee') | Q(statut='en_cours') | Q(statut='terminee')).order_by('-date_validation')
    
    # Appliquer les filtres comme dans la vue dashboard
    statut = request.GET.get('statut')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if statut:
        missions = missions.filter(statut=statut)
    
    if date_debut:
        try:
            date_debut = datetime.datetime.strptime(date_debut, '%Y-%m-%d').date()
            missions = missions.filter(date_validation__date__gte=date_debut)
        except (ValueError, TypeError):
            pass  # Ignorer si la date est mal formatée
    
    if date_fin:
        try:
            date_fin = datetime.datetime.strptime(date_fin, '%Y-%m-%d').date()
            missions = missions.filter(date_validation__date__lte=date_fin)
        except (ValueError, TypeError):
            pass  # Ignorer si la date est mal formatée

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Missions"

        # En-têtes
        headers = [
            "ID", "Date de validation", "Demandeur", "Destination", "Point d'embarquement", 
            "Motif", "Nombre de passagers", "Statut", "Chauffeur", "Véhicule", "Dispatcher", 
            "Date départ", "Date arrivée", "Km départ", "Km arrivée", "Distance parcourue"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Données
        for row, mission in enumerate(missions, start=2):
            try:
                ws.cell(row=row, column=1, value=mission.id)
                ws.cell(row=row, column=2, value=mission.date_validation.strftime("%d/%m/%Y %H:%M") if mission.date_validation else "")
                ws.cell(row=row, column=3, value=mission.demandeur.get_full_name() if mission.demandeur else "")
                ws.cell(row=row, column=4, value=mission.destination or "")
                ws.cell(row=row, column=5, value=mission.point_embarquement or "")
                ws.cell(row=row, column=6, value=mission.motif or "")
                ws.cell(row=row, column=7, value=mission.nombre_passagers or "")
                ws.cell(row=row, column=8, value=mission.get_statut_display())
                ws.cell(row=row, column=9, value=mission.chauffeur.get_full_name() if mission.chauffeur else "")
                ws.cell(row=row, column=10, value=f"{mission.vehicule.immatriculation} - {mission.vehicule.marque} {mission.vehicule.modele}" if mission.vehicule else "")
                ws.cell(row=row, column=11, value=mission.dispatcher.get_full_name() if mission.dispatcher else "")
                ws.cell(row=row, column=12, value=mission.date_depart.strftime("%d/%m/%Y %H:%M") if mission.date_depart else "")
                ws.cell(row=row, column=13, value=mission.date_fin.strftime("%d/%m/%Y %H:%M") if mission.date_fin else "")
                ws.cell(row=row, column=14, value=mission.kilometrage_depart or "")
                ws.cell(row=row, column=15, value=mission.kilometrage_fin or "")
                ws.cell(row=row, column=16, value=mission.distance_parcourue or "")
            except Exception as e:
                # En cas d'erreur sur une mission, on passe à la suivante
                logger = logging.getLogger(__name__)
                logger.error("Erreur lors de l'exportation de la mission %s: %s", mission.id, str(e), exc_info=True)
                continue

        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Génération de la réponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=missions.xlsx'
        wb.save(response)
        return response
        
    except Exception as e:
        # En cas d'erreur lors de la création du fichier
        logger = logging.getLogger(__name__)
        logger.error("Erreur lors de la génération du fichier Excel: %s", str(e), exc_info=True)
        
        # Retourner une réponse d'erreur
        messages.error(request, f"Une erreur est survenue lors de l'exportation: {str(e)}")
        return redirect('chauffeur:dashboard')

@login_required
def mission_pdf(request, mission_id):
    mission = get_object_or_404(Course, id=mission_id)
    template = get_template('chauffeur/mission_pdf.html')
    
    # Préparer le contexte avec les variables nécessaires
    context = {
        'mission': mission,
        'request': request,
        'now': timezone.now()
    }
    
    html = template.render(context)

    # Vérifier si l'export PDF est activé
    from .deployment_config import is_feature_enabled, get_pdf_placeholder_response
    
    if not is_feature_enabled('ENABLE_PDF_EXPORT'):
        return get_pdf_placeholder_response()
    
    # Temporairement désactivé - weasyprint non disponible
    return get_pdf_placeholder_response()
    
    # fd, path = tempfile.mkstemp(suffix=".pdf")
    # try:
    #     HTML(string=html).write_pdf(path)
    #     with open(path, "rb") as f:
    #         pdf = f.read()
    # finally:
    #     os.close(fd)
    #     os.remove(path)

    # response = HttpResponse(pdf, content_type='application/pdf')
    # response['Content-Disposition'] = f'filename="mission_{mission_id}.pdf"'
    # return response

@login_required
def mission_excel(request, mission_id):
    mission = get_object_or_404(Course, id=mission_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mission {mission.id}"

    # En-têtes
    ws['A1'] = "Champ"
    ws['B1'] = "Valeur"
    ws['A1'].font = ws['B1'].font = Font(bold=True)

    # Données
    data = [
        ("ID", mission.id),
        ("Date de validation", mission.date_validation.strftime("%d/%m/%Y %H:%M") if mission.date_validation else ""),
        ("Demandeur", mission.demandeur.get_full_name() if mission.demandeur else ""),
        ("Destination", mission.destination),
        ("Point d'embarquement", mission.point_embarquement),
        ("Motif", mission.motif),
        ("Nombre de passagers", mission.nombre_passagers),
        ("Statut", mission.get_statut_display()),
        ("Chauffeur", mission.chauffeur.get_full_name() if mission.chauffeur else ""),
        ("Véhicule", f"{mission.vehicule.immatriculation} - {mission.vehicule.marque} {mission.vehicule.modele}" if mission.vehicule else ""),
        ("Dispatcher", mission.dispatcher.get_full_name() if mission.dispatcher else ""),
        ("Date départ", mission.date_depart.strftime("%d/%m/%Y %H:%M") if mission.date_depart else ""),
        ("Date arrivée", mission.date_fin.strftime("%d/%m/%Y %H:%M") if mission.date_fin else ""),
        ("Km départ", mission.kilometrage_depart),
        ("Km arrivée", mission.kilometrage_fin),
        ("Distance parcourue", mission.distance_parcourue),
    ]
    for i, (champ, valeur) in enumerate(data, start=2):
        ws[f'A{i}'] = champ
        ws[f'B{i}'] = valeur

    # Génération de la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=mission_{mission.id}.xlsx'
    wb.save(response)
    return response

@login_required
def get_vehicule_kilometrage(request):
    vehicule_id = request.GET.get('vehicule_id')
    if not vehicule_id:
        return JsonResponse({'error': 'Aucun ID de véhicule fourni'}, status=400)
    try:
        vehicule = Vehicule.objects.get(id=vehicule_id)
        dernier_kilometrage = 0
        
        # Vérifier d'abord le dernier ravitaillement
        dernier_ravitaillement = Ravitaillement.objects.filter(
            vehicule=vehicule
        ).order_by('-date_ravitaillement').first()
        if dernier_ravitaillement and dernier_ravitaillement.kilometrage_apres > 0:
            dernier_kilometrage = dernier_ravitaillement.kilometrage_apres
        
        # Vérifier ensuite la dernière course terminée
        derniere_course = Course.objects.filter(
            vehicule=vehicule, 
            statut='terminee'
        ).order_by('-date_fin').first()
        if derniere_course and derniere_course.kilometrage_fin > dernier_kilometrage:
            dernier_kilometrage = derniere_course.kilometrage_fin
            
        # Vérifier enfin le dernier entretien
        dernier_entretien = Entretien.objects.filter(
            vehicule=vehicule,
            kilometrage_apres__isnull=False
        ).order_by('-date_entretien').first()
        if dernier_entretien and dernier_entretien.kilometrage_apres > dernier_kilometrage:
            dernier_kilometrage = dernier_entretien.kilometrage_apres
            
        return JsonResponse({
            'success': True,
            'kilometrage': dernier_kilometrage,
            'immatriculation': vehicule.immatriculation
        })
    except Vehicule.DoesNotExist:
        return JsonResponse({'error': 'Véhicule non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def rapport_chauffeur(request):
    """Vue pour afficher un rapport détaillé du chauffeur connecté"""
    from django.db.models import Sum, Avg, Count
    import datetime
    user = request.user
    if user.role != 'chauffeur' and not user.is_superuser and user.role != 'admin':
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')

    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    statut = request.GET.get('statut')
    vehicule_id = request.GET.get('vehicule')
    
    missions = Course.objects.filter(chauffeur=user)
    
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    if statut:
        missions = missions.filter(statut=statut)
    if vehicule_id:
        missions = missions.filter(vehicule_id=vehicule_id)
        
    missions = missions.order_by('-date_validation')

    # Statistiques globales
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    jours_travailles = missions.values('date_validation__date').distinct().count()
    missions_terminees = missions.filter(statut='terminee').count()
    conso_moyenne = None
    cout_total = 0
    if total_km > 0:
        # Consommation moyenne sur les ravitaillements du chauffeur
        ravs = Ravitaillement.objects.filter(chauffeur=user)
        litres = ravs.aggregate(l=Sum('litres'))['l'] or 0
        conso_moyenne = round((litres * 100) / total_km, 2) if litres > 0 else None
        cout_total = ravs.aggregate(c=Sum('cout_total'))['c'] or 0
    # Incidents (exemple: missions avec commentaire d'incident)
    incidents = 0  # Champ 'commentaires' n'existe pas sur Course
    # Score simple (à améliorer selon critères)
    score = 10
    if conso_moyenne and conso_moyenne > 8:
        score -= 1
    # incidents est toujours 0 ici
    # if incidents > 0:
    #     score -= incidents
    # Statistiques par mois
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    
    # Destinations fréquentes
    destinations_frequentes = missions.values('destination').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Statistiques par statut
    stats_par_statut = missions.values('statut').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Récupérer la liste des véhicules pour le filtre
    vehicules = Vehicule.objects.all().order_by('immatriculation')
    # Pagination
    paginator = Paginator(missions, 12)  # 12 missions par page
    page_number = request.GET.get('page')
    missions_page = paginator.get_page(page_number)
    
    # Préparer les données pour le template
    context = {
        'missions': missions_page,
        'total_missions': total_missions,
        'total_km': total_km,
        'jours_travailles': jours_travailles,
        'missions_terminees': missions_terminees,
        'conso_moyenne': conso_moyenne,
        'cout_total': cout_total,
        'incidents': incidents,
        'score': score,
        'stats_mois': stats_mois,
        'destinations_frequentes': destinations_frequentes,
        'stats_par_statut': stats_par_statut,
        'vehicules': vehicules,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    # Récupérer le département principal du chauffeur
    departement = None
    if hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    # Gestion POST pour commentaire
    if request.method == 'POST' and 'commentaire' in request.POST:
        texte = request.POST.get('commentaire', '').strip()
        if texte:
            # Sécuriser la récupération des bornes de date
            first_mission = missions.order_by('date_validation').first()
            last_mission = missions.order_by('-date_validation').first()
            date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
            date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
            CommentaireRapportChauffeur.objects.create(
                chauffeur=user,
                auteur=request.user,
                date_debut=date_debut_val or timezone.now().date(),
                date_fin=date_fin_val or timezone.now().date(),
                texte=texte
            )
            messages.success(request, "Commentaire enregistré.")
            return redirect(request.path + f'?date_debut={date_debut}&date_fin={date_fin}')
    # Récupérer les commentaires existants pour la période
    first_mission = missions.order_by('date_validation').first()
    last_mission = missions.order_by('-date_validation').first()
    date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
    date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
    commentaires = CommentaireRapportChauffeur.objects.filter(
        chauffeur=user,
        date_debut=date_debut_val or timezone.now().date(),
        date_fin=date_fin_val or timezone.now().date(),
    ).order_by('-date_commentaire')
    return render(request, 'chauffeur/rapport_chauffeur.html', {
        'missions': missions_page,
        'total_missions': total_missions,
        'total_km': total_km,
        'jours_travailles': jours_travailles,
        'missions_terminees': missions_terminees,
        'conso_moyenne': conso_moyenne,
        'cout_total': cout_total,
        'incidents': incidents,
        'score': score,
        'stats_mois': stats_mois,
        'destinations_frequentes': destinations_frequentes,
        'stats_par_statut': stats_par_statut,
        'vehicules': vehicules,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'commentaires': commentaires,
        'departement': departement,
    })

@login_required
def rapport_chauffeur_pdf(request):
    from django.db.models import Sum, Avg, Count
    import datetime
    user = request.user
    if user.role != 'chauffeur' and not user.is_superuser and user.role != 'admin':
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    missions = Course.objects.filter(chauffeur=user)
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    missions = missions.order_by('-date_validation')
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    jours_travailles = missions.values('date_validation__date').distinct().count()
    conso_moyenne = None
    cout_total = 0
    if total_km > 0:
        ravs = Ravitaillement.objects.filter(chauffeur=user)
        litres = ravs.aggregate(l=Sum('litres'))['l'] or 0
        conso_moyenne = round((litres * 100) / total_km, 2) if litres > 0 else None
        cout_total = ravs.aggregate(c=Sum('cout_total'))['c'] or 0
    incidents = missions.filter(motif__icontains='incident').count()
    score = 10
    if conso_moyenne and conso_moyenne > 8:
        score -= 1
    if incidents > 0:
        score -= incidents
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    departement = None
    if hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    
    # Gestion des dates par défaut avec vérification de sécurité
    if not date_debut and missions.exists():
        first_mission = missions.order_by('date_validation').first()
        date_debut = first_mission.date_validation.date() if first_mission and first_mission.date_validation else None
    
    if not date_fin and missions.exists():
        last_mission = missions.order_by('-date_validation').first()
        date_fin = last_mission.date_validation.date() if last_mission and last_mission.date_validation else None
    
    commentaires = CommentaireRapportChauffeur.objects.filter(
        chauffeur=user,
        date_debut=date_debut,
        date_fin=date_fin,
    ).order_by('-date_commentaire')
    context = {
        'missions': missions,
        'total_missions': total_missions,
        'total_km': total_km,
        'jours_travailles': jours_travailles,
        'conso_moyenne': conso_moyenne,
        'cout_total': cout_total,
        'incidents': incidents,
        'score': score,
        'stats_mois': stats_mois,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'commentaires': commentaires,
        'departement': departement,
        'user': user,
        'now': datetime.datetime.now(),
    }
    template = get_template('chauffeur/rapport_chauffeur_pdf.html')
    html = template.render(context)
    # Vérifier si l'export PDF est activé
    from .deployment_config import is_feature_enabled, get_pdf_placeholder_response
    
    if not is_feature_enabled('ENABLE_PDF_EXPORT'):
        return get_pdf_placeholder_response()
    
    # Temporairement désactivé - weasyprint non disponible
    return get_pdf_placeholder_response()
    
    # fd, path = tempfile.mkstemp(suffix=".pdf")
    # try:
    #     HTML(string=html).write_pdf(path)
    #     with open(path, "rb") as f:
    #         pdf = f.read()
    # finally:
    #     os.close(fd)
    #     os.remove(path)
    # response = HttpResponse(pdf, content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; filename=rapport_chauffeur_{user.username}.pdf'
    # return response

@login_required
def rapport_chauffeur_excel(request):
    from django.db.models import Sum, Avg, Count
    import datetime
    user = request.user
    if user.role != 'chauffeur' and not user.is_superuser and user.role != 'admin':
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    missions = Course.objects.filter(chauffeur=user)
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    missions = missions.order_by('-date_validation')
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    jours_travailles = missions.values('date_validation__date').distinct().count()
    conso_moyenne = None
    cout_total = 0
    if total_km > 0:
        ravs = Ravitaillement.objects.filter(chauffeur=user)
        litres = ravs.aggregate(l=Sum('litres'))['l'] or 0
        conso_moyenne = round((litres * 100) / total_km, 2) if litres > 0 else None
        cout_total = ravs.aggregate(c=Sum('cout_total'))['c'] or 0
    incidents = missions.filter(motif__icontains='incident').count()
    score = 10
    if conso_moyenne and conso_moyenne > 8:
        score -= 1
    if incidents > 0:
        score -= incidents
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    departement = None
    if hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    
    # Gestion des dates par défaut avec vérification de sécurité
    if not date_debut and missions.exists():
        first_mission = missions.order_by('date_validation').first()
        date_debut = first_mission.date_validation.date() if first_mission and first_mission.date_validation else None
    
    if not date_fin and missions.exists():
        last_mission = missions.order_by('-date_validation').first()
        date_fin = last_mission.date_validation.date() if last_mission and last_mission.date_validation else None
    
    commentaires = CommentaireRapportChauffeur.objects.filter(
        chauffeur=user,
        date_debut=date_debut,
        date_fin=date_fin,
    ).order_by('-date_commentaire')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Chauffeur"
    ws['A1'] = "Rapport Chauffeur"
    ws['A2'] = f"Nom : {user.get_full_name()}"
    ws['A3'] = f"Département : {departement}"
    ws['A4'] = f"Période : {date_debut or '...'} au {date_fin or '...'}"
    ws['A5'] = f"Date d'édition : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = ws['A3'].font = ws['A4'].font = ws['A5'].font = Font(bold=True)
    ws.append([])
    ws.append(["Statistiques globales"])
    ws.append(["Missions", total_missions])
    ws.append(["Kilomètres parcourus", total_km])
    ws.append(["Jours travaillés", jours_travailles])
    ws.append(["Consommation moyenne (L/100km)", conso_moyenne or '-'])
    ws.append(["Coût total", cout_total])
    ws.append(["Score", f"{score}/10"])
    ws.append([])
    ws.append(["Statistiques par mois"])
    ws.append(["Mois", "Missions", "Kilomètres"])
    for stat in stats_mois:
        ws.append([
            stat['mois'].strftime('%B %Y') if stat['mois'] else '',
            stat['missions'],
            stat['km']
        ])
    ws.append([])
    ws.append(["Incidents"])
    if incidents > 0:
        for mission in missions:
            if mission.motif and 'incident' in mission.motif.lower():
                ws.append([f"Mission #{mission.id} ({mission.date_validation.strftime('%d/%m/%Y') if mission.date_validation else ''})", mission.motif])
    else:
        ws.append(["Aucun incident signalé."])
    ws.append([])
    ws.append(["Commentaires / Appréciations"])
    for c in commentaires:
        ws.append([f"Par {c.auteur.username if c.auteur else 'Anonyme'} le {c.date_commentaire.strftime('%d/%m/%Y %H:%M')}", c.texte])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_chauffeur_{user.username}.xlsx'
    wb.save(response)
    return response

@login_required
def rapport_demandeur(request):
    from django.db.models import Sum, Avg, Count
    import datetime
    user = request.user
    User = get_user_model()
    # Si admin/superuser, permettre de choisir le demandeur à afficher
    demandeur_id = request.GET.get('demandeur_id')
    if user.role in ['admin'] or user.is_superuser:
        # Inclure tous les utilisateurs qui peuvent être demandeurs (pas seulement le rôle 'demandeur')
        demandeurs = User.objects.filter(role__in=['demandeur', 'admin', 'dispatch'])
        if demandeur_id:
            try:
                demandeur = User.objects.get(id=demandeur_id)
            except User.DoesNotExist:
                demandeur = None
        else:
            demandeur = None
    else:
        demandeurs = None
        demandeur = user
    if not (user.role == 'demandeur' or user.is_superuser or user.role == 'admin'):
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')

    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    # On filtre dès le départ les missions aberrantes
    missions = Course.objects.filter(distance_parcourue__lte=1000)
    if demandeur:
        missions = missions.filter(demandeur=demandeur)
    elif user.role == 'demandeur':
        missions = missions.filter(demandeur=user)
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    missions = missions.order_by('-date_validation')

    # Statistiques globales
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    missions_terminees = missions.filter(statut='terminee').count()
    missions_en_cours = missions.filter(statut='en_cours').count()
    missions_validees = missions.filter(statut='validee').count()
    
    cout_par_km = 0.5  # Coût estimé par km (à ajuster selon vos besoins)
    cout_total_estime = total_km * cout_par_km
    
    destinations_frequentes = missions.values('destination').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    chauffeurs_frequents = missions.values('chauffeur__first_name', 'chauffeur__last_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    
    departement = None
    if demandeur and hasattr(demandeur, 'etablissement') and demandeur.etablissement:
        departement = demandeur.etablissement.nom
    elif hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    
    # Gestion POST pour commentaire
    if request.method == 'POST' and 'commentaire' in request.POST:
        texte = request.POST.get('commentaire', '').strip()
        if texte:
            first_mission = missions.order_by('date_validation').first()
            last_mission = missions.order_by('-date_validation').first()
            date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
            date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
            CommentaireRapportDemandeur.objects.create(
                demandeur=demandeur or user,
                auteur=request.user,
                date_debut=date_debut_val or timezone.now().date(),
                date_fin=date_fin_val or timezone.now().date(),
                texte=texte
            )
            messages.success(request, "Commentaire enregistré.")
            params = f'?date_debut={date_debut}&date_fin={date_fin}'
            if demandeur_id:
                params += f'&demandeur_id={demandeur_id}'
            return redirect(request.path + params)
    first_mission = missions.order_by('date_validation').first()
    last_mission = missions.order_by('-date_validation').first()
    date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
    date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
    commentaires = CommentaireRapportDemandeur.objects.filter(
        demandeur=demandeur or user,
        date_debut=date_debut_val or timezone.now().date(),
        date_fin=date_fin_val or timezone.now().date(),
    ).order_by('-date_commentaire')
    
    # Pagination
    paginator = Paginator(missions, 12)  # 12 missions par page
    page_number = request.GET.get('page')
    missions_page = paginator.get_page(page_number)
    
    context = {
        'missions': missions_page,
        'total_missions': total_missions,
        'total_km': total_km,
        'missions_terminees': missions_terminees,
        'missions_en_cours': missions_en_cours,
        'missions_validees': missions_validees,
        'cout_total_estime': cout_total_estime,
        'destinations_frequentes': destinations_frequentes,
        'chauffeurs_frequents': chauffeurs_frequents,
        'stats_mois': stats_mois,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'commentaires': commentaires,
        'departement': departement,
        'demandeurs': demandeurs,
        'demandeur_id': demandeur_id,
    }
    
    return render(request, 'chauffeur/rapport_demandeur.html', context)

@login_required
def rapport_demandeur_pdf(request):
    """Export PDF du rapport demandeur"""
    from django.db.models import Sum, Count
    import datetime
    user = request.user
    User = get_user_model()
    demandeur_id = request.GET.get('demandeur_id')
    if user.role in ['admin'] or user.is_superuser:
        if demandeur_id:
            try:
                demandeur = User.objects.get(id=demandeur_id)
            except User.DoesNotExist:
                demandeur = None
        else:
            demandeur = None
    else:
        demandeur = user
    if user.role != 'demandeur' and not user.is_superuser and user.role != 'admin':
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if not date_debut or date_debut in ["None", "", None]:
        date_debut = None
    if not date_fin or date_fin in ["None", "", None]:
        date_fin = None
    missions = Course.objects.all()
    if demandeur:
        missions = missions.filter(demandeur=demandeur)
    elif user.role == 'demandeur':
        missions = missions.filter(demandeur=user)
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    missions = missions.order_by('-date_validation')
    
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    missions_terminees = missions.filter(statut='terminee').count()
    missions_en_cours = missions.filter(statut='en_cours').count()
    missions_validees = missions.filter(statut='validee').count()
    
    cout_par_km = 0.5
    cout_total_estime = total_km * cout_par_km
    
    destinations_frequentes = missions.values('destination').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    chauffeurs_frequents = missions.values('chauffeur__first_name', 'chauffeur__last_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    
    departement = None
    if demandeur and hasattr(demandeur, 'etablissement') and demandeur.etablissement:
        departement = demandeur.etablissement.nom
    elif hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    
    first_mission = missions.order_by('date_validation').first()
    last_mission = missions.order_by('-date_validation').first()
    date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
    date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
    commentaires = CommentaireRapportDemandeur.objects.filter(
        demandeur=demandeur or user,
        date_debut=date_debut_val or timezone.now().date(),
        date_fin=date_fin_val or timezone.now().date(),
    ).order_by('-date_commentaire')
    
    # Encoder le logo en base64 pour le PDF
    import base64
    from django.conf import settings
    logo_base64 = ''
    try:
        logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_ips_co.png'
        with open(logo_path, 'rb') as logo_file:
            logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
    except Exception as e:
        print(f"Erreur lors du chargement du logo: {e}")
    
    context = {
        'missions': missions,
        'total_missions': total_missions,
        'total_km': total_km,
        'missions_terminees': missions_terminees,
        'missions_en_cours': missions_en_cours,
        'missions_validees': missions_validees,
        'cout_total_estime': cout_total_estime,
        'destinations_frequentes': destinations_frequentes,
        'chauffeurs_frequents': chauffeurs_frequents,
        'stats_mois': stats_mois,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'commentaires': commentaires,
        'departement': departement,
        'user': demandeur or user,
        'now': datetime.datetime.now(),
        'logo_base64': logo_base64,
    }
    
    template = get_template('chauffeur/rapport_demandeur_pdf.html')
    html = template.render(context)
    # Vérifier si l'export PDF est activé
    from .deployment_config import is_feature_enabled, get_pdf_placeholder_response
    
    if not is_feature_enabled('ENABLE_PDF_EXPORT'):
        return get_pdf_placeholder_response()
    
    # Temporairement désactivé - weasyprint non disponible
    return get_pdf_placeholder_response()
    
    # fd, path = tempfile.mkstemp(suffix=".pdf")
    # try:
    #     HTML(string=html).write_pdf(path)
    #     with open(path, "rb") as f:
    #         pdf = f.read()
    # finally:
    #     os.close(fd)
    #         os.remove(path)
    
    # response = HttpResponse(pdf, content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; filename=rapport_demandeur_{(demandeur or user).username}.pdf'
    # return response

@login_required
def rapport_demandeur_excel(request):
    """Export Excel du rapport demandeur"""
    from django.db.models import Sum, Count
    import datetime
    user = request.user
    User = get_user_model()
    demandeur_id = request.GET.get('demandeur_id')
    if user.role in ['admin'] or user.is_superuser:
        if demandeur_id:
            try:
                demandeur = User.objects.get(id=demandeur_id)
            except User.DoesNotExist:
                demandeur = None
        else:
            demandeur = None
    else:
        demandeur = user
    if user.role != 'demandeur' and not user.is_superuser and user.role != 'admin':
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if not date_debut or date_debut in ["None", "", None]:
        date_debut = None
    if not date_fin or date_fin in ["None", "", None]:
        date_fin = None
    missions = Course.objects.all()
    if demandeur:
        missions = missions.filter(demandeur=demandeur)
    elif user.role == 'demandeur':
        missions = missions.filter(demandeur=user)
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)
    missions = missions.order_by('-date_validation')
    
    total_missions = missions.count()
    total_km = missions.aggregate(total=Sum('distance_parcourue'))['total'] or 0
    missions_terminees = missions.filter(statut='terminee').count()
    missions_en_cours = missions.filter(statut='en_cours').count()
    missions_validees = missions.filter(statut='validee').count()
    
    cout_par_km = 0.5
    cout_total_estime = total_km * cout_par_km
    
    destinations_frequentes = missions.values('destination').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    chauffeurs_frequents = missions.values('chauffeur__first_name', 'chauffeur__last_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    from django.db.models.functions import TruncMonth
    stats_mois = missions.annotate(mois=TruncMonth('date_validation')).values('mois').annotate(
        missions=Count('id'),
        km=Sum('distance_parcourue'),
    ).order_by('mois')
    
    departement = None
    if demandeur and hasattr(demandeur, 'etablissement') and demandeur.etablissement:
        departement = demandeur.etablissement.nom
    elif hasattr(user, 'etablissement') and user.etablissement:
        departement = user.etablissement.nom
    
    first_mission = missions.order_by('date_validation').first()
    last_mission = missions.order_by('-date_validation').first()
    date_debut_val = date_debut or (first_mission.date_validation.date() if first_mission and first_mission.date_validation else None)
    date_fin_val = date_fin or (last_mission.date_validation.date() if last_mission and last_mission.date_validation else None)
    commentaires = CommentaireRapportDemandeur.objects.filter(
        demandeur=demandeur or user,
        date_debut=date_debut_val or timezone.now().date(),
        date_fin=date_fin_val or timezone.now().date(),
    ).order_by('-date_commentaire')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Demandeur"
    ws['A1'] = "Rapport Demandeur"
    ws['A2'] = f"Nom : {(demandeur or user).get_full_name()}"
    ws['A3'] = f"Département : {departement}"
    ws['A4'] = f"Période : {date_debut or '...'} au {date_fin or '...'}"
    ws['A5'] = f"Date d'édition : {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = ws['A3'].font = ws['A4'].font = ws['A5'].font = Font(bold=True)
    ws.append([])
    ws.append(["Statistiques globales"])
    ws.append(["Missions totales", total_missions])
    ws.append(["Missions terminées", missions_terminees])
    ws.append(["Missions en cours", missions_en_cours])
    ws.append(["Missions validées", missions_validees])
    ws.append(["Kilomètres parcourus", total_km])
    ws.append(["Coût total estimé", cout_total_estime])
    ws.append([])
    ws.append(["Destinations les plus fréquentes"])
    ws.append(["Destination", "Nombre de missions"])
    for dest in destinations_frequentes:
        ws.append([dest['destination'], dest['count']])
    ws.append([])
    ws.append(["Chauffeurs les plus utilisés"])
    ws.append(["Chauffeur", "Nombre de missions"])
    for chauffeur in chauffeurs_frequents:
        nom_complet = f"{chauffeur['chauffeur__first_name']} {chauffeur['chauffeur__last_name']}"
        ws.append([nom_complet, chauffeur['count']])
    ws.append([])
    ws.append(["Statistiques par mois"])
    ws.append(["Mois", "Missions", "Kilomètres"])
    for stat in stats_mois:
        ws.append([
            stat['mois'].strftime('%B %Y') if stat['mois'] else '',
            stat['missions'],
            stat['km']
        ])
    ws.append([])
    ws.append(["Commentaires / Appréciations"])
    for c in commentaires:
        ws.append([f"Par {c.auteur.username if c.auteur else 'Anonyme'} le {c.date_commentaire.strftime('%d/%m/%Y %H:%M')}", c.texte])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_demandeur_{(demandeur or user).username}.xlsx'
    wb.save(response)
    return response

@login_required
def rapport_missions_par_demandeur(request):
    """Rapport global : nombre de missions par demandeur (admin uniquement)"""
    from django.db.models import Count, Sum
    import datetime
    user = request.user
    if not (user.is_superuser or user.role == 'admin'):
        messages.error(request, "Vous n'avez pas accès à ce rapport.")
        return redirect('home')

    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    missions = Course.objects.all()
    if date_debut:
        missions = missions.filter(date_validation__date__gte=date_debut)
    if date_fin:
        missions = missions.filter(date_validation__date__lte=date_fin)

    # Pagination sur les missions (12 par page)
    from django.core.paginator import Paginator
    page_number = request.GET.get('page')
    paginator = Paginator(missions, 12)
    missions_page = paginator.get_page(page_number)

    # Agrégation par demandeur
    stats_demandeurs = missions.values('demandeur__id', 'demandeur__first_name', 'demandeur__last_name', 'demandeur__username').annotate(
        nb_missions=Count('id'),
        km_total=Sum('distance_parcourue'),
    ).order_by('-nb_missions')

    context = {
        'stats_demandeurs': stats_demandeurs,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'chauffeur/rapport_missions_par_demandeur.html', context)
